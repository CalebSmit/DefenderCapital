"""
Portfolio Risk Dashboard
======================================================
Run with:  streamlit run dashboard/app.py
"""

import sys, time, warnings, json
from pathlib import Path
import numpy as np
import pandas as pd
import plotly.graph_objects as go
import plotly.express as px
from plotly.subplots import make_subplots
import streamlit as st
import streamlit.components.v1 as components

# ── path setup ──────────────────────────────────────────────────────────────
ROOT = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(ROOT))

from engine.utils import EXPORTS_DIR, get_portfolio_path
from engine.data_loader import load_portfolio
from engine.market_data import fetch_market_data
from engine.risk_metrics import compute_all_metrics
from engine.monte_carlo import run_simulation, compute_multihorizon_es
from engine.stress_testing import run_all_stress_tests
from engine.backtesting import run_backtest
from engine.report_generator import generate_html_report, write_results_to_excel

warnings.filterwarnings("ignore")

# ── page config ─────────────────────────────────────────────────────────────
st.set_page_config(
    page_title="Risk Model",
    page_icon="D",
    layout="wide",
    initial_sidebar_state="expanded",
)

# ════════════════════════════════════════════════════════════════════════════
# AUTH GATE  — must run before any other content
# ════════════════════════════════════════════════════════════════════════════
from dashboard.login_page import render_auth_gate, logout, current_user
from auth.portfolio_store import (
    save_portfolio as _db_save_portfolio,
    load_portfolio as _db_load_portfolio,
    has_portfolio  as _db_has_portfolio,
    holdings_to_excel_bytes,
    excel_path_to_holdings,
)

if not render_auth_gate():
    st.stop()

# ── auto-load saved portfolio on first login this session ────────────────────
_dcm_user = current_user()
# HIGH-8 FIX: Namespace session state by user ID
_user_key = str(_dcm_user["id"]) if _dcm_user else "anon"
_portfolio_loaded_key = f"dcm_portfolio_loaded_{_user_key}"

if _dcm_user and not st.session_state.get(_portfolio_loaded_key):
    _saved = _db_load_portfolio(_dcm_user["id"])
    if _saved and _saved["holdings"]:
        import hashlib, tempfile as _tmpmod
        _h_bytes = holdings_to_excel_bytes(
            _saved["portfolio_name"],
            _saved["short_name"],
            _saved["holdings"],
            _saved["settings"],
        )
        _h_hash = hashlib.md5(_h_bytes).hexdigest()
        if st.session_state.get("uploaded_portfolio_hash") != _h_hash:
            _tmp_dir = Path(_tmpmod.gettempdir()) / "dcm_uploads"
            _tmp_dir.mkdir(exist_ok=True)
            _tmp_path = _tmp_dir / f"saved_{_dcm_user['id']}.xlsx"
            _tmp_path.write_bytes(_h_bytes)
            st.session_state["uploaded_portfolio_path"] = str(_tmp_path)
            st.session_state["uploaded_portfolio_hash"] = _h_hash
            _load_all_needs_clear = True
        else:
            _load_all_needs_clear = False
    else:
        _load_all_needs_clear = False
    st.session_state[_portfolio_loaded_key] = True
    if _load_all_needs_clear:
        st.rerun()

# ── brand colours (Broadsheet palette) ────────────────────────────────────
DCM_CANVAS    = "#faf8f4"
DCM_SURFACE   = "#ffffff"
DCM_SURFACE2  = "#f5f2eb"
DCM_BORDER    = "#e2ddd3"
DCM_GOLD      = "#c9a84c"
DCM_RED       = "#c62828"
DCM_GREEN     = "#1b7a3d"
DCM_BLUE      = "#2962ff"
DCM_INK       = "#1a1a1a"
DCM_MUTED     = "#6b6560"
DCM_DIM       = "#a9a29a"
# Legacy aliases
DCM_TEXT      = DCM_INK
DCM_ICE       = DCM_BLUE
DCM_LIGHT     = DCM_SURFACE
DCM_GRAY      = DCM_MUTED
DCM_VOID      = DCM_CANVAS

# ── Plotly brand template ────────────────────────────────────────────────
import plotly.io as pio

_DCM_TEMPLATE = go.layout.Template(
    layout=go.Layout(
        font=dict(family="JetBrains Mono, monospace", size=11, color=DCM_INK),
        title=dict(font=dict(size=14, color=DCM_INK), x=0.0, xanchor="left"),
        paper_bgcolor=DCM_CANVAS,
        plot_bgcolor=DCM_SURFACE,
        colorway=[DCM_BLUE, DCM_GOLD, DCM_GREEN, DCM_RED,
                  "#6a1b9a", "#e65100", "#00838f", "#ad1457", "#558b2f", "#283593"],
        margin=dict(t=30, b=50, l=60, r=20),
        xaxis=dict(
            gridcolor="rgba(226,221,211,0.6)", linecolor=DCM_BORDER,
            tickfont=dict(size=10, color=DCM_MUTED),
            title_font=dict(size=10, color=DCM_MUTED), zeroline=False,
        ),
        yaxis=dict(
            gridcolor="rgba(226,221,211,0.6)", linecolor=DCM_BORDER,
            tickfont=dict(size=10, color=DCM_MUTED),
            title_font=dict(size=10, color=DCM_MUTED), zeroline=False,
        ),
        legend=dict(font=dict(size=10, color=DCM_MUTED),
                    bgcolor="rgba(255,255,255,0.95)", bordercolor=DCM_BORDER),
        hoverlabel=dict(bgcolor=DCM_SURFACE, font_color=DCM_INK, font_size=11,
                        bordercolor=DCM_GOLD),
        hovermode="x unified",
    )
)
pio.templates["dcm"] = _DCM_TEMPLATE
pio.templates.default = "dcm"

PLOTLY_CONFIG = dict(
    displayModeBar=True,
    modeBarButtonsToRemove=["pan2d", "select2d", "lasso2d", "autoScale2d", "toggleSpikelines"],
    displaylogo=False,
)

# ── global CSS — Broadsheet ──────────────────────────────────────────────────
st.markdown("""
<style>
  /* ── Fonts ───────────────────────────────────────────────────────────── */
  @import url('https://fonts.googleapis.com/css2?family=Cormorant+Garamond:wght@300;400;600;700&family=Source+Sans+3:wght@300;400;500;600&family=JetBrains+Mono:wght@400;500;600&display=swap');

  /* ── Design Tokens ───────────────────────────────────────────────────── */
  :root {
    --canvas:      #faf8f4;
    --surface:     #ffffff;
    --surface-alt: #f5f2eb;
    --border:      #e2ddd3;
    --border-hover:#c9a84c;

    --ink:         #1a1a1a;
    --ink-muted:   #6b6560;
    --ink-dim:     #a9a29a;

    --gain:        #1b7a3d;
    --gain-bg:     rgba(27,122,61,0.06);
    --loss:        #c62828;
    --loss-bg:     rgba(198,40,40,0.06);
    --warn:        #e65100;
    --warn-bg:     rgba(230,81,0,0.06);

    --accent:      #c9a84c;
    --accent-hover:#b8943f;
    --accent-bg:   rgba(201,168,76,0.08);

    --radius:      5px;
    --ease:        cubic-bezier(0.4,0,0.2,1);
    --transition:  all 0.3s cubic-bezier(0.4,0,0.2,1);
    --font-display:'Cormorant Garamond', 'Georgia', serif;
    --font-data:   'JetBrains Mono', 'Consolas', monospace;
    --font-ui:     'Source Sans 3', 'Segoe UI', sans-serif;
  }

  /* ── Global Canvas ───────────────────────────────────────────────────── */
  html, body, [class*="css"], .main, .stApp {
    font-family: var(--font-ui) !important;
    color: var(--ink) !important;
  }
  .stApp {
    background: var(--canvas) !important;
    background-image: url("data:image/svg+xml,%3Csvg viewBox='0 0 256 256' xmlns='http://www.w3.org/2000/svg'%3E%3Cfilter id='n'%3E%3CfeTurbulence type='fractalNoise' baseFrequency='0.9' numOctaves='4' stitchTiles='stitch'/%3E%3C/filter%3E%3Crect width='100%25' height='100%25' filter='url(%23n)' opacity='0.02'/%3E%3C/svg%3E") !important;
  }

  /* ── Page Entrance Animation ─────────────────────────────────────────── */
  .main .block-container {
    padding-top: 2rem !important;
  }
  @keyframes slideUp {
    from { opacity: 0; transform: translateY(20px); }
    to   { opacity: 1; transform: translateY(0); }
  }
  .main .block-container h1 {
    animation: slideUp 0.6s var(--ease) both;
  }
  .main .block-container [data-testid="stCaptionContainer"] {
    animation: slideUp 0.6s var(--ease) 0.08s both;
  }
  .main .block-container [data-testid="stHorizontalBlock"] {
    animation: slideUp 0.5s var(--ease) 0.15s both;
  }
  .main .block-container .section-header {
    animation: slideUp 0.5s var(--ease) 0.2s both;
  }
  .main .block-container [data-testid="stPlotlyChart"] {
    animation: slideUp 0.5s var(--ease) 0.25s both;
  }

  /* ── Typography ──────────────────────────────────────────────────────── */
  h1 {
    font-family: var(--font-display) !important;
    font-weight: 300 !important; letter-spacing: -0.04em !important;
    color: var(--ink) !important; font-size: 3.2rem !important;
    line-height: 1.1 !important;
  }
  h2 {
    font-family: var(--font-display) !important;
    font-weight: 700 !important; font-size: 1.4rem !important;
    color: var(--ink) !important;
  }
  .stCaption, [data-testid="stCaptionContainer"] {
    font-family: var(--font-data) !important;
    font-size: 10px !important; color: var(--ink-dim) !important;
    text-transform: uppercase; letter-spacing: 0.12em;
  }
  code, .stCodeBlock {
    font-family: var(--font-data) !important; font-size: 12px !important;
    background: var(--surface-alt) !important; color: var(--ink) !important;
  }

  /* ── Sidebar ─────────────────────────────────────────────────────────── */
  [data-testid="stSidebar"] {
    background: var(--surface-alt) !important;
    border-right: 1px solid var(--border) !important;
    box-shadow: none !important;
    position: relative;
  }
  [data-testid="stSidebar"]::before {
    content: '';
    position: absolute; inset: 0; pointer-events: none;
    background: repeating-linear-gradient(
      135deg,
      var(--accent) 0, var(--accent) 1px,
      transparent 1px, transparent 8px
    );
    opacity: 0.03;
  }
  [data-testid="stSidebar"] .stSelectbox label {
    color: var(--ink-dim) !important; font-weight: 500;
    text-transform: uppercase; font-size: 9px; letter-spacing: 0.1em;
    font-family: var(--font-data) !important;
  }
  [data-testid="stSidebar"] .stRadio > label {
    font-size: 9px !important; text-transform: uppercase;
    letter-spacing: 0.08em; color: var(--ink-dim) !important;
    font-family: var(--font-data) !important;
  }
  [data-testid="stSidebar"] .stRadio [role="radiogroup"] label {
    padding: 10px 14px !important; border-radius: var(--radius) !important;
    transition: var(--transition); margin-bottom: 2px !important;
    font-family: var(--font-ui) !important;
    font-weight: 500 !important; font-size: 13px !important;
    border-left: 2px solid transparent;
  }
  [data-testid="stSidebar"] .stRadio [role="radiogroup"] label:hover {
    background: var(--accent-bg) !important;
    color: var(--ink) !important;
    transform: translateX(4px);
    border-left-color: var(--accent) !important;
  }
  [data-testid="stSidebar"] hr {
    border-color: var(--border) !important; margin: 16px 0 !important;
  }
  [data-testid="stSidebar"] .stButton > button {
    background: transparent !important;
    border: 1px solid var(--accent) !important;
    color: var(--accent) !important; font-weight: 600 !important;
    font-family: var(--font-data) !important; font-size: 12px !important;
    letter-spacing: 0.04em;
    transition: var(--transition);
  }
  [data-testid="stSidebar"] .stButton > button:hover {
    background: var(--accent) !important;
    color: var(--surface) !important;
    transform: translateY(-1px);
    box-shadow: 0 4px 12px rgba(201,168,76,0.2);
  }

  /* ── Metric Cards ────────────────────────────────────────────────────── */
  .metric-card {
    background: var(--surface); border-radius: var(--radius);
    padding: 20px 18px; text-align: left;
    border: 1px solid var(--border);
    border-top: 3px solid var(--accent);
    transition: var(--transition);
    box-shadow: 0 1px 3px rgba(0,0,0,0.04), 0 4px 16px rgba(0,0,0,0.03);
    animation: slideUp 0.5s var(--ease) both;
  }
  .metric-card:hover {
    transform: translateY(-2px);
    box-shadow: 0 4px 20px rgba(0,0,0,0.08);
    border-top-color: var(--border-hover);
  }
  .metric-card .label {
    font-family: var(--font-data); font-size: 9px; color: var(--ink-dim);
    text-transform: uppercase; font-weight: 500;
    letter-spacing: 0.12em; margin-bottom: 10px; line-height: 1.3;
  }
  .metric-card .value {
    font-family: var(--font-data); font-size: 28px; font-weight: 600;
    color: var(--ink); margin: 6px 0;
    letter-spacing: -0.01em; font-variant-numeric: tabular-nums; line-height: 1.2;
  }
  .metric-card .sub {
    font-family: var(--font-data); font-size: 10px; color: var(--ink-dim);
    line-height: 1.4;
  }

  /* ── Section Headers ─────────────────────────────────────────────────── */
  .section-header {
    margin: 36px 0 18px; padding: 0;
  }
  .section-header span {
    font-family: var(--font-display); font-size: 18px; font-weight: 700;
    color: var(--ink); letter-spacing: -0.01em;
    display: block; margin-bottom: 8px;
  }
  .section-rule {
    height: 1px;
    background: linear-gradient(90deg, var(--accent), transparent 50%);
  }

  /* ── Badges ──────────────────────────────────────────────────────────── */
  .pass-badge {
    background: var(--gain-bg); color: var(--gain); padding: 4px 12px;
    border-radius: 20px; font-size: 11px; font-weight: 600;
    font-family: var(--font-data);
    display: inline-flex; align-items: center; gap: 4px;
  }
  .fail-badge {
    background: var(--loss-bg); color: var(--loss); padding: 4px 12px;
    border-radius: 20px; font-size: 11px; font-weight: 600;
    font-family: var(--font-data);
    display: inline-flex; align-items: center; gap: 4px;
  }
  .warn-badge {
    background: var(--warn-bg); color: var(--warn); padding: 4px 12px;
    border-radius: 20px; font-size: 11px; font-weight: 600;
    font-family: var(--font-data);
    display: inline-flex; align-items: center; gap: 4px;
  }

  /* ── Widget Overrides ────────────────────────────────────────────────── */
  div[data-testid="stHorizontalBlock"] { gap: 12px; }
  [data-testid="stDataFrame"] {
    border-radius: var(--radius) !important; overflow: hidden;
    border: 1px solid var(--border) !important;
    transition: var(--transition);
  }
  [data-testid="stExpander"] {
    background: var(--surface) !important;
    border: 1px solid var(--border) !important;
    border-radius: var(--radius) !important;
    transition: var(--transition);
  }
  [data-testid="stExpander"]:hover {
    border-color: var(--accent) !important;
  }
  [data-testid="stExpander"] summary {
    font-weight: 600 !important; color: var(--ink) !important;
    font-family: var(--font-ui) !important;
  }

  .stButton > button[kind="primary"] {
    background: var(--accent) !important; color: #ffffff !important;
    border: none !important; font-weight: 600 !important;
    font-family: var(--font-data) !important; letter-spacing: 0.04em;
    transition: var(--transition);
  }
  .stButton > button[kind="primary"]:hover {
    background: var(--accent-hover) !important;
    transform: translateY(-1px);
    box-shadow: 0 4px 16px rgba(201,168,76,0.25);
  }
  .stSelectbox > div > div {
    border-color: var(--border) !important;
    border-radius: var(--radius) !important;
    transition: var(--transition);
  }
  .stSelectbox > div > div:hover,
  .stSelectbox > div > div:focus-within {
    border-color: var(--accent) !important;
  }

  [data-testid="stDownloadButton"] > button {
    border: 1px solid var(--border) !important;
    border-radius: var(--radius) !important;
    transition: var(--transition); font-weight: 500;
    font-family: var(--font-data) !important;
    color: var(--ink-muted) !important;
  }
  [data-testid="stDownloadButton"] > button:hover {
    border-color: var(--accent) !important;
    color: var(--accent) !important;
    transform: translateY(-1px);
  }

  /* ── Success/Error/Info ──────────────────────────────────────────────── */
  [data-testid="stAlert"] {
    border-radius: var(--radius) !important;
    font-family: var(--font-data) !important;
  }

  /* ── Utility Bar ─────────────────────────────────────────────────────── */
  .util-bar {
    display: flex;
    justify-content: flex-end;
    align-items: center;
    gap: 8px;
    margin: -4px 0 20px;
  }
  .util-btn {
    display: inline-flex;
    align-items: center;
    gap: 6px;
    padding: 7px 18px;
    min-width: 100px;
    justify-content: center;
    border-radius: 20px;
    border: 1.5px solid var(--border);
    background: var(--surface);
    color: var(--ink-muted) !important;
    font-family: var(--font-data);
    font-size: 11px;
    font-weight: 600;
    letter-spacing: 0.08em;
    text-transform: uppercase;
    text-decoration: none !important;
    cursor: pointer;
    transition: var(--transition);
    box-shadow: 0 1px 3px rgba(0,0,0,0.04);
    white-space: nowrap;
  }
  .util-btn svg {
    flex-shrink: 0;
    opacity: 0.7;
    transition: opacity 0.2s ease;
  }
  .util-btn:hover {
    border-color: var(--accent);
    color: var(--accent) !important;
    background: var(--accent-bg);
    transform: translateY(-1px);
    box-shadow: 0 3px 12px rgba(201,168,76,0.14);
  }
  .util-btn:hover svg {
    opacity: 1;
  }
  .util-btn:active {
    transform: translateY(0);
    box-shadow: 0 1px 4px rgba(201,168,76,0.10);
  }

  /* ── Hide Streamlit Branding ─────────────────────────────────────────── */
  footer { visibility: hidden; }
  #MainMenu { visibility: hidden; }
  header[data-testid="stHeader"] { background: var(--canvas) !important; }

  /* ── Sector Holdings Tables ────────────────────────────────────────── */
  .sector-block {
    margin-bottom: 36px;
    animation: slideUp 0.5s var(--ease) both;
  }
  .sector-title {
    font-family: var(--font-display);
    font-size: 20px;
    font-weight: 700;
    color: var(--ink);
    margin: 0 0 4px 0;
    display: flex;
    align-items: center;
    gap: 10px;
  }
  .sector-title .swatch {
    width: 14px;
    height: 14px;
    border-radius: 3px;
    display: inline-block;
    flex-shrink: 0;
  }
  .sector-subtitle {
    font-family: var(--font-data);
    font-size: 10px;
    color: var(--ink-dim);
    text-transform: uppercase;
    letter-spacing: 0.1em;
    margin-bottom: 10px;
  }
  .sector-table {
    width: 100%;
    border-collapse: collapse;
    font-family: var(--font-data);
    font-size: 12px;
    border: 1px solid var(--border);
    border-radius: var(--radius);
    overflow: hidden;
    background: var(--surface);
  }
  .sector-table thead th {
    background: var(--surface-alt);
    color: var(--ink-muted);
    font-weight: 600;
    font-size: 9px;
    text-transform: uppercase;
    letter-spacing: 0.08em;
    padding: 10px 12px;
    text-align: right;
    border-bottom: 1px solid var(--border);
  }
  .sector-table thead th:first-child,
  .sector-table thead th:nth-child(2) {
    text-align: left;
  }
  .sector-table tbody td {
    padding: 8px 12px;
    border-bottom: 1px solid rgba(226,221,211,0.5);
    color: var(--ink);
    text-align: right;
    font-variant-numeric: tabular-nums;
    transition: background 0.2s var(--ease);
  }
  .sector-table tbody td:first-child {
    text-align: left;
    font-weight: 600;
    color: var(--ink);
  }
  .sector-table tbody td:nth-child(2) {
    text-align: left;
    color: var(--ink-muted);
    font-family: var(--font-ui);
    font-size: 12px;
  }
  .sector-table tbody tr:hover td {
    background: var(--accent-bg);
  }
  .sector-table tfoot td {
    padding: 10px 12px;
    font-weight: 700;
    color: var(--ink);
    border-top: 2px solid var(--border);
    background: var(--surface-alt);
    text-align: right;
  }
  .sector-table tfoot td:first-child {
    text-align: left;
  }

  /* ── Sector vs S&P 500 Comparison Table ────────────────────────────── */
  .comparison-table {
    width: 100%;
    border-collapse: collapse;
    font-family: var(--font-data);
    font-size: 12px;
    border: 1px solid var(--border);
    border-radius: var(--radius);
    overflow: hidden;
    background: var(--surface);
    margin-bottom: 8px;
  }
  .comparison-table thead th {
    background: var(--surface-alt);
    color: var(--ink-muted);
    font-weight: 600;
    font-size: 9px;
    text-transform: uppercase;
    letter-spacing: 0.08em;
    padding: 10px 14px;
    text-align: right;
    border-bottom: 1px solid var(--border);
  }
  .comparison-table thead th:first-child {
    text-align: left;
  }
  .comparison-table tbody td {
    padding: 9px 14px;
    border-bottom: 1px solid rgba(226,221,211,0.5);
    color: var(--ink);
    text-align: right;
    font-variant-numeric: tabular-nums;
    transition: background 0.2s var(--ease);
  }
  .comparison-table tbody td:first-child {
    text-align: left;
    font-weight: 500;
  }
  .comparison-table tbody tr:hover td {
    background: var(--accent-bg);
  }
  .comparison-table tbody tr.separator-row td {
    border-top: 2px solid var(--border);
  }
  .comparison-table tfoot td {
    padding: 10px 14px;
    font-weight: 700;
    color: var(--ink);
    border-top: 2px solid var(--border);
    background: var(--surface-alt);
    text-align: right;
  }
  .comparison-table tfoot td:first-child {
    text-align: left;
  }
  .delta-pos { color: var(--gain); }
  .delta-neg { color: var(--loss); }

  /* ── Responsive ──────────────────────────────────────────────────────── */
  @media (max-width: 768px) {
    .metric-card .value { font-size: 22px; }
    .metric-card .label { font-size: 8px; }
    h1 { font-size: 2.2rem !important; }
  }
</style>
""", unsafe_allow_html=True)

# ════════════════════════════════════════════════════════════════════════════
# TEMPLATE GENERATION
# ════════════════════════════════════════════════════════════════════════════

def _generate_template_bytes() -> bytes:
    """Generate a clean, annotated Excel template as bytes for download."""
    import io
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    wb = openpyxl.Workbook()

    # ── Holdings sheet ────────────────────────────────────────────────────
    ws_h = wb.active
    ws_h.title = "Holdings"

    hdr_font  = Font(bold=True, color="FFFFFF", size=11)
    hdr_fill  = PatternFill("solid", fgColor="1B2A4A")
    note_fill = PatternFill("solid", fgColor="FFF8DC")
    note_font = Font(italic=True, color="7B6B4A", size=9)
    thin      = Side(border_style="thin", color="C0C0C0")
    cell_border = Border(left=thin, right=thin, top=thin, bottom=thin)

    # Title row
    ws_h.merge_cells("A1:C1")
    title_cell = ws_h["A1"]
    title_cell.value = "Portfolio Holdings — DCM Risk Model Template"
    title_cell.font  = Font(bold=True, size=13, color="1B2A4A")
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    title_cell.fill = PatternFill("solid", fgColor="EEF2F8")
    ws_h.row_dimensions[1].height = 28

    # Instruction row
    ws_h.merge_cells("A2:C2")
    instr = ws_h["A2"]
    instr.value = "Fill in each row below with a stock holding. Ticker must be a valid US exchange symbol (e.g. AAPL, MSFT). Delete these example rows before uploading."
    instr.font  = note_font
    instr.fill  = note_fill
    instr.alignment = Alignment(wrap_text=True, horizontal="left")
    ws_h.row_dimensions[2].height = 32

    # Header row
    headers = [("Ticker", 14), ("Shares Held", 16), ("Cost Basis", 16)]
    for col_idx, (header, width) in enumerate(headers, start=1):
        cell = ws_h.cell(row=3, column=col_idx, value=header)
        cell.font   = hdr_font
        cell.fill   = hdr_fill
        cell.alignment = Alignment(horizontal="center")
        cell.border = cell_border
        ws_h.column_dimensions[get_column_letter(col_idx)].width = width

    # Example holdings rows
    examples = [
        ("AAPL",  150, 145.50),
        ("MSFT",  100, 280.00),
        ("GOOGL",  50, 2800.00),
        ("AMZN",   80, 155.00),
        ("NVDA",  200, 480.00),
        ("JPM",   120, 165.00),
        ("V",      90, 210.00),
        ("JNJ",   110, 160.00),
    ]
    alt_fill = PatternFill("solid", fgColor="F5F8FF")
    for r_idx, (ticker, shares, cost) in enumerate(examples, start=4):
        fill = alt_fill if r_idx % 2 == 0 else None
        for col_idx, val in enumerate([ticker, shares, cost], start=1):
            cell = ws_h.cell(row=r_idx, column=col_idx, value=val)
            cell.border = cell_border
            if fill:
                cell.fill = fill
            if col_idx == 3:
                cell.number_format = '"$"#,##0.00'
            cell.font = Font(color="555555", italic=True, size=10)

    # Note about cost basis
    ws_h.merge_cells(f"A{4+len(examples)+1}:C{4+len(examples)+1}")
    note2 = ws_h.cell(row=4+len(examples)+1, column=1,
                      value="↑ Cost Basis = price per share paid (used for unrealised P&L display only — not for risk calculations). Delete example rows above before uploading.")
    note2.font  = note_font
    note2.fill  = note_fill
    note2.alignment = Alignment(wrap_text=True)

    # ── Settings sheet ────────────────────────────────────────────────────
    ws_s = wb.create_sheet("Settings")

    ws_s.merge_cells("A1:B1")
    ws_s["A1"].value = "Portfolio Settings (optional — safe to leave as defaults)"
    ws_s["A1"].font  = Font(bold=True, size=12, color="1B2A4A")
    ws_s["A1"].fill  = PatternFill("solid", fgColor="EEF2F8")
    ws_s["A1"].alignment = Alignment(horizontal="center")
    ws_s.row_dimensions[1].height = 26

    ws_s.cell(row=2, column=1, value="Parameter").font = Font(bold=True)
    ws_s.cell(row=2, column=2, value="Value").font     = Font(bold=True)
    ws_s.cell(row=2, column=3, value="Description").font = Font(bold=True)
    for col, w in [(1, 30), (2, 22), (3, 55)]:
        ws_s.column_dimensions[get_column_letter(col)].width = w

    settings_rows = [
        ("benchmark_ticker",        "SPY",      "Benchmark for beta/alpha calc (e.g. SPY, QQQ, IWM)"),
        ("risk_free_rate",          "auto",     "'auto' fetches current 3-month T-bill; or enter a decimal e.g. 0.04"),
        ("confidence_level_1",      0.95,       "Primary VaR confidence level"),
        ("confidence_level_2",      0.99,       "Secondary VaR confidence level"),
        ("es_confidence_level",     0.975,      "Primary Expected Shortfall confidence (FRTB standard = 0.975)"),
        ("lookback_years",          2,          "Years of historical price data used for all calculations"),
        ("simulation_paths",        10000,      "Number of Monte Carlo paths (higher = slower but more accurate)"),
        ("simulation_days",         252,        "Trading days to simulate forward (252 = 1 year)"),
        ("portfolio_name",          "My Portfolio", "Full name shown in reports and dashboard header"),
        ("portfolio_short_name",    "PORT",     "Short name / ticker used in compact displays"),
        ("covariance_mode",         "ledoit_wolf", "'ledoit_wolf' (default) or 'ewma' for time-varying covariance"),
        ("ewma_lambda",             0.94,       "EWMA decay factor (only used if covariance_mode = ewma)"),
        ("mc_shock_distribution",   "normal",   "'normal' or 'student_t' for fat-tailed Monte Carlo shocks"),
        ("stress_custom_drawdown",  -0.20,      "Custom uniform drawdown scenario (negative fraction, e.g. -0.20)"),
        ("report_title",            "Portfolio Risk Report", "Title shown on generated PDF/HTML reports"),
    ]
    for r_idx, (param, val, desc) in enumerate(settings_rows, start=3):
        ws_s.cell(row=r_idx, column=1, value=param).font = Font(size=10)
        ws_s.cell(row=r_idx, column=2, value=val).font   = Font(size=10, color="1B2A4A", bold=True)
        ws_s.cell(row=r_idx, column=3, value=desc).font  = Font(size=9, italic=True, color="666666")
        if r_idx % 2 == 0:
            for col in range(1, 4):
                ws_s.cell(row=r_idx, column=col).fill = PatternFill("solid", fgColor="F5F8FF")

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()


# ════════════════════════════════════════════════════════════════════════════
# DATA LOADING  (always fresh — no caching)
# ════════════════════════════════════════════════════════════════════════════

def _load_all():
    """Load portfolio, market data, and run all analytics. Returns a dict."""
    excel_path = None
    if "uploaded_portfolio_path" in st.session_state:
        excel_path = Path(st.session_state["uploaded_portfolio_path"])
    load_result = load_portfolio(excel_path=excel_path)
    cfg         = load_result.settings
    md          = fetch_market_data(load_result)
    metrics     = compute_all_metrics(md, settings=cfg)
    sim         = run_simulation(
        md, n_paths=cfg.simulation_paths, n_days=cfg.simulation_days, seed=42,
        es_confidence_level=cfg.es_confidence_level,
        shock_distribution=cfg.mc_shock_distribution, df=cfg.mc_df,
    )
    stress      = run_all_stress_tests(
        md, custom_drawdown=cfg.stress_custom_drawdown, metrics=metrics,
    )
    # Multi-horizon ES table (1-day, 10-day, 21-day at 95/97.5/99%)
    try:
        mh_es = compute_multihorizon_es(
            md, horizons=[1, 10, 21],
            confidence_levels=[0.95, cfg.es_confidence_level, 0.99],
            n_paths=min(cfg.simulation_paths, 2000), seed=42,
        )
    except Exception:
        mh_es = None
    # Backtesting (parametric rolling VaR, min 252-day training window)
    try:
        backtest = run_backtest(md, confidence=0.95, min_periods=252)
    except Exception:
        backtest = None
    return dict(load_result=load_result, md=md, metrics=metrics,
                sim=sim, stress=stress, mh_es=mh_es, backtest=backtest)


def get_data(force_refresh: bool = False):
    return _load_all()


# ════════════════════════════════════════════════════════════════════════════
# AI CHATBOT WIDGET — dynamic system prompt + data serialisation
# ════════════════════════════════════════════════════════════════════════════

def _generate_system_prompt(data: dict) -> str:
    """Build a fully dynamic system prompt from the live dashboard state.

    Updates automatically whenever methodology settings or portfolio change.
    """
    metrics     = data["metrics"]
    load_result = data["load_result"]
    settings    = load_result.settings
    stress_obj  = data.get("stress")

    # Collect stress scenario names
    stress_names: list[str] = []
    try:
        for s in (stress_obj.all_scenarios if hasattr(stress_obj, "all_scenarios") else []):
            if hasattr(s, "name"):
                stress_names.append(s.name)
    except Exception:
        pass

    cov_label = (
        "Ledoit-Wolf shrinkage (reduces estimation error for high-dimensional covariance matrices)"
        if settings.covariance_mode == "ledoit_wolf"
        else f"Exponentially Weighted Moving Average (lambda={getattr(settings, 'ewma_lambda', 0.94)})"
    )
    rfr = getattr(settings, "risk_free_rate", "auto")
    rfr_label = rfr if rfr and rfr.lower() != "auto" else "auto-fetched from FRED 3-month T-Bill"
    mc_dist = settings.mc_shock_distribution
    mc_extra = (
        f" (Student-t df={getattr(settings, 'mc_df', 5)} — heavier tails than normal)"
        if mc_dist == "student_t" else ""
    )

    # Top-5 by weight for prompt grounding
    top5 = sorted(load_result.holdings, key=lambda h: -h.weight)[:5]
    top5_str = ", ".join(f"{h.ticker} ({h.weight:.1%})" for h in top5)

    lines = [
        f"You are an expert AI portfolio risk analyst embedded in the Defender Capital Management"
        f" (DCM) portfolio risk dashboard. Your role is to help interpret the live data shown on"
        f" each dashboard page and answer questions about portfolio risk, methodology, and analytics.",
        "",
        f"## Current Portfolio: {settings.portfolio_name}",
        f"Data as of {pd.Timestamp.now().strftime('%B %d, %Y')}."
        f" All data is fetched fresh on each page load.",
        f"Holdings: {len(load_result.holdings)} positions."
        f" Top 5 by weight: {top5_str}.",
        "",
        "## Analytics Methodology (current configuration)",
        f"- **Covariance estimation**: {cov_label}",
        f"- **Return history lookback**: {settings.lookback_years} year(s) of daily returns",
        f"- **Benchmark**: {settings.benchmark_ticker}",
        f"- **Risk-free rate**: {rfr_label}",
        f"- **VaR confidence levels**: {settings.confidence_level_1*100:.0f}% and {settings.confidence_level_2*100:.0f}%",
        f"- **Expected Shortfall (CVaR) confidence**: {settings.es_confidence_level*100:.1f}%",
        f"- **VaR methods computed**: Parametric (variance-covariance with Cornish-Fisher expansion),"
        f" Historical simulation (full-repricing), Monte Carlo simulation",
        f"- **Monte Carlo**: {settings.simulation_paths:,} paths, {settings.simulation_days}-day horizon,"
        f" shock distribution: {mc_dist}{mc_extra}",
        "",
        "## Per-Stock Metrics Available",
        "Annualized volatility, beta, alpha, Sharpe ratio, Sortino ratio, Calmar ratio, max drawdown,"
        " max drawdown duration, skewness, kurtosis, 30-day rolling vol, 90-day rolling vol,"
        " annualized return, component VaR (95%), marginal VaR (95%), risk contribution %.",
        "",
        "## Portfolio-Level Metrics Available",
        "All per-stock metrics aggregated, plus: HHI concentration index, effective number of bets (1/HHI),"
        " diversification ratio, average pairwise correlation, PCA factor decomposition"
        f" (n factors explaining 90% of variance = {getattr(metrics, 'n_pca_factors_90pct', 'N/A')}),"
        " multi-horizon Expected Shortfall (1-day, 10-day, 21-day at 95/97.5/99%).",
        "",
        "## Dashboard Pages",
        "1. **Portfolio Overview** — KPI summary (value, return, vol, Sharpe, beta);"
        " sector allocation bar chart vs S&P 500 benchmark; top-15 holdings table;"
        " weight distribution bar chart; full holdings P&L table.",
        "2. **Holdings by Sector** — Sector grouping with weight totals;"
        " sector pie chart; intra-sector weight breakdown bars.",
        "3. **Risk Dashboard** — VaR/CVaR comparison table (3 methods × 2 confidence levels);"
        " component VaR waterfall chart; risk contribution donut; correlation heatmap;"
        " rolling volatility chart (30d/90d); drawdown chart; return distribution histogram.",
        f"4. **Monte Carlo** — {settings.simulation_paths:,}-path simulation fan chart;"
        " multi-horizon ES table; probability of loss analysis; projected P&L distribution.",
        f"5. **Stress Tests** — Named scenario P&L analysis"
        f" ({', '.join(stress_names[:6]) + ('…' if len(stress_names) > 6 else '') if stress_names else 'historical crash scenarios'});"
        " per-holding stress impact breakdown; severity classification vs ES benchmarks.",
        "6. **Stock Analysis** — Per-ticker deep-dive page: rolling vol, drawdown chart,"
        " return histogram, beta/alpha, Sharpe/Sortino, risk contribution vs portfolio.",
        "7. **Backtesting** — Rolling parametric VaR backtesting overlaid with actual returns;"
        " VaR breach count/rate; Kupiec POF test; conditional and unconditional coverage.",
        "8. **Reports & Export** — Generate standalone HTML risk report; export full metrics to Excel.",
        "9. **Settings** — Configure methodology: covariance mode, confidence levels, MC paths,"
        " benchmark, lookback period, shock distribution, custom stress drawdown, etc.",
        "",
        "## Response Guidelines",
        "- **Lead with the answer**, then explain. Be concise — 3–6 sentences for most queries.",
        "- **Cite specific numbers** from the [CONTEXT DATA] block injected before each question.",
        "- Flag risk concentrations, anomalies, or actionable observations proactively.",
        "- If a methodology question is asked, explain the exact configured approach (above).",
        "- Use **bold** for key figures, bullet lists for comparisons or rankings.",
        "- If a value is not in the provided context, say so — do not fabricate data.",
        "- When the user asks about a specific page, describe what that page shows using the"
        " page descriptions above.",
    ]
    return "\n".join(lines)


def _serialize_portfolio(data: dict) -> str:
    """Serialise live portfolio data + dynamic system prompt into a JSON string
    for injection into the chatbot widget HTML."""
    try:
        metrics     = data["metrics"]
        load_result = data["load_result"]
        settings    = load_result.settings
        stress_obj  = data.get("stress")

        # Index stock metrics by ticker
        sm_map: dict = {}
        for sm in (metrics.stock_metrics or []):
            sm_map[sm.ticker] = sm

        def _f(v, decimals=6):
            """Safe float round — handles NaN/inf."""
            try:
                x = float(v or 0)
                if not (x == x) or abs(x) == float('inf'):  # NaN / inf check
                    return 0.0
                return round(x, decimals)
            except Exception:
                return 0.0

        # Holdings
        holdings_out = []
        for h in load_result.holdings:
            sm = sm_map.get(h.ticker)
            holdings_out.append({
                "ticker":              h.ticker,
                "name":               h.company_name,
                "sector":             h.sector,
                "industry":           getattr(h, "industry", ""),
                "weight":             _f(h.weight),
                "market_value":       _f(h.market_value, 2),
                "current_price":      _f(h.current_price, 4),
                "shares_held":        int(h.shares_held),
                "unrealized_pct":     _f(getattr(h, "unrealized_pct", 0.0)),
                "annualized_vol":     _f(sm.annualized_vol   if sm else 0),
                "annualized_return":  _f(sm.annualized_return if sm else 0),
                "beta":               _f(sm.beta             if sm else 0, 4),
                "alpha":              _f(sm.alpha            if sm else 0),
                "sharpe":             _f(sm.sharpe           if sm else 0, 4),
                "sortino":            _f(sm.sortino          if sm else 0, 4),
                "max_drawdown":       _f(sm.max_drawdown     if sm else 0, 4),
                "component_var_95":   _f(sm.component_var_95 if sm else 0),
                "marginal_var_95":    _f(sm.marginal_var_95  if sm else 0),
                "risk_contribution_pct": _f(sm.risk_contribution_pct if sm else 0, 4),
            })

        # Portfolio summary
        v95 = metrics.var_95
        v99 = metrics.var_99
        portfolio_out = {
            "name":                 settings.portfolio_name,
            "total_value":          _f(metrics.total_value, 2),
            "annualized_return":    _f(metrics.annualized_return),
            "annualized_vol":       _f(metrics.annualized_vol),
            "sharpe":               _f(metrics.sharpe, 4),
            "sortino":              _f(metrics.sortino, 4),
            "calmar":               _f(metrics.calmar, 4),
            "beta":                 _f(metrics.beta, 4),
            "alpha":                _f(metrics.alpha),
            "max_drawdown":         _f(metrics.max_drawdown, 4),
            "var_95_hist":          _f(v95.historical_var),
            "var_95_param":         _f(v95.parametric_var),
            "var_95_mc":            _f(v95.mc_var),
            "var_99_hist":          _f(v99.historical_var),
            "eff_num_bets":         _f(metrics.eff_num_bets, 2),
            "hhi":                  _f(metrics.hhi),
            "diversification_ratio": _f(metrics.diversification_ratio, 4),
            "avg_corr":             _f(metrics.avg_pairwise_corr, 4),
        }

        # Sector exposures
        sector_exp: dict = {}
        for h in load_result.holdings:
            s = h.sector
            if s not in sector_exp:
                sector_exp[s] = {"weight": 0.0, "count": 0}
            sector_exp[s]["weight"] = round(sector_exp[s]["weight"] + h.weight, 8)
            sector_exp[s]["count"] += 1

        # Stress scenarios
        stress_out = []
        try:
            all_sc = stress_obj.all_scenarios if hasattr(stress_obj, "all_scenarios") else []
            for s in all_sc:
                pnl_abs = getattr(s, "portfolio_loss_usd", None)
                pnl_pct = getattr(s, "portfolio_loss_pct", 0.0)
                if pnl_abs is None:
                    pnl_abs = _f(pnl_pct) * _f(metrics.total_value, 2)
                stress_out.append({
                    "scenario": getattr(s, "name", "?"),
                    "pnl_pct":  _f(pnl_pct),
                    "pnl_abs":  _f(pnl_abs, 2),
                })
        except Exception:
            pass

        # Methodology
        methodology = {
            "covariance_mode":     settings.covariance_mode,
            "confidence_level_1":  f"{settings.confidence_level_1*100:.0f}%",
            "confidence_level_2":  f"{settings.confidence_level_2*100:.0f}%",
            "es_confidence_level": f"{settings.es_confidence_level*100:.1f}%",
            "simulation_paths":    settings.simulation_paths,
            "simulation_days":     settings.simulation_days,
            "mc_shock_distribution": settings.mc_shock_distribution,
            "lookback_years":      settings.lookback_years,
            "benchmark":          settings.benchmark_ticker,
        }

        payload = {
            "portfolio":        portfolio_out,
            "holdings":         holdings_out,
            "sector_exposures": sector_exp,
            "stress_summary":   stress_out,
            "methodology":      methodology,
            "systemPrompt":     _generate_system_prompt(data),
        }
        return json.dumps(payload, allow_nan=False)

    except Exception as exc:  # noqa: BLE001
        fallback = {
            "portfolio": {}, "holdings": [], "sector_exposures": {},
            "stress_summary": [], "methodology": {},
            "systemPrompt": (
                "You are an expert portfolio risk analyst AI assistant."
                f" (Error building context: {exc})"
            ),
        }
        return json.dumps(fallback)


def render_chatbot_widget(data: dict) -> None:
    """Inject the floating AI chatbot widget with live portfolio data.

    Reads chatbot_widget.html, injects the serialised portfolio JSON,
    and renders it as a 0-height Streamlit component whose JS self-appends
    to the parent Streamlit page DOM.
    """
    try:
        widget_path = Path(__file__).parent / "chatbot_widget.html"
        html_src    = widget_path.read_text(encoding="utf-8")
        portfolio_json = _serialize_portfolio(data)
        html_src = html_src.replace(
            "window.__PORTFOLIO_DATA__ = {};",
            f"window.__PORTFOLIO_DATA__ = {portfolio_json};",
            1,
        )
        components.html(html_src, height=0)
    except Exception:
        pass  # never crash the dashboard over the chatbot


def _load_settings_quick() -> dict:
    """Load just settings from Excel as a plain dict."""
    defaults = {
        "portfolio_name": "Defender Capital Management",
        "portfolio_short_name": "DCM",
        "report_title": "Portfolio Risk Report",
    }
    excel_path = (
        Path(st.session_state["uploaded_portfolio_path"])
        if "uploaded_portfolio_path" in st.session_state
        else None
    )
    try:
        lr = load_portfolio(excel_path=excel_path)
        s = lr.settings
        return {
            "portfolio_name": getattr(s, "portfolio_name", defaults["portfolio_name"]),
            "portfolio_short_name": getattr(s, "portfolio_short_name", defaults["portfolio_short_name"]),
            "report_title": getattr(s, "report_title", defaults["report_title"]),
        }
    except Exception:
        return defaults


def _fetch_sector_etf_prices() -> pd.DataFrame:
    """Fetch 13 months of price history for all sector ETFs."""
    import yfinance as yf
    import importlib, engine.report_generator as _rg_mod
    importlib.reload(_rg_mod)
    SECTOR_ETF_MAP = _rg_mod.SECTOR_ETF_MAP
    etf_tickers = list(SECTOR_ETF_MAP.values())
    end = pd.Timestamp.now().strftime("%Y-%m-%d")
    start = (pd.Timestamp.now() - pd.DateOffset(months=13)).strftime("%Y-%m-%d")
    raw = yf.download(etf_tickers, start=start, end=end, auto_adjust=True,
                       progress=False, actions=False, group_by="ticker", threads=True)
    frames = {}
    if isinstance(raw.columns, pd.MultiIndex):
        for t in etf_tickers:
            try:
                s = raw[t]["Close"].dropna()
                if len(s) > 5:
                    frames[t] = s
            except Exception:
                pass
    elif "Close" in raw.columns and len(etf_tickers) == 1:
        frames[etf_tickers[0]] = raw["Close"].dropna()
    return pd.DataFrame(frames)


def _period_return(prices: pd.Series, months: int) -> float | None:
    """Simple return over the last N months of a price series."""
    if prices is None or len(prices) < 2:
        return None
    end_price = prices.iloc[-1]
    cutoff = prices.index[-1] - pd.DateOffset(months=months)
    past = prices.loc[:cutoff]
    if len(past) == 0:
        return None
    start_price = past.iloc[-1]
    if start_price <= 0:
        return None
    return (end_price / start_price) - 1.0


def _sector_portfolio_return(md, sector: str, months: int) -> float | None:
    """Weight-weighted return for portfolio holdings in a given sector over N months."""
    holdings_in = [h for h in md.holdings if h.sector == sector and h.ticker in md.prices.columns]
    if not holdings_in:
        return None
    total_mv = sum(h.market_value for h in holdings_in)
    if total_mv <= 0:
        return None
    weighted_ret = 0.0
    for h in holdings_in:
        ret = _period_return(md.prices[h.ticker].dropna(), months)
        if ret is None:
            continue
        w = h.market_value / total_mv
        weighted_ret += w * ret
    return weighted_ret


# ════════════════════════════════════════════════════════════════════════════
# HELPERS
# ════════════════════════════════════════════════════════════════════════════

def metric_card(label: str, value: str, sub: str = "", color: str = DCM_INK,
                trend: str | None = None, icon: str = "") -> str:
    border_color = DCM_GOLD
    if color == DCM_RED:
        border_color = DCM_RED
    elif color == DCM_GREEN:
        border_color = DCM_GREEN
    elif color == "#8b0000":
        border_color = "#c62828"
    trend_html = ""
    if trend == "up":
        trend_html = f'<span style="color:{DCM_GREEN};font-size:13px;margin-left:6px">&#9650;</span>'
    elif trend == "down":
        trend_html = f'<span style="color:{DCM_RED};font-size:13px;margin-left:6px">&#9660;</span>'
    return f"""
<div class="metric-card" style="border-top-color:{border_color}">
  <div class="label">{label}</div>
  <div class="value" style="color:{color}">{value}{trend_html}</div>
  <div class="sub">{sub}</div>
</div>"""

def section_header(text: str):
    st.markdown(f'<div class="section-header"><span>{text}</span><div class="section-rule"></div></div>',
                unsafe_allow_html=True)

def fmt_pct(v: float, decimals: int = 1) -> str:
    return f"{v*100:.{decimals}f}%"

def fmt_dollar(v: float) -> str:
    return f"${abs(v):,.0f}" if abs(v) < 1e6 else f"${abs(v)/1e6:.2f}M"

# yfinance sector names — these are the canonical names returned by yfinance .info["sector"]
SECTOR_COLORS = {
    "Technology": "#2962ff", "Financial Services": "#c9a84c",
    "Healthcare": "#ad1457", "Consumer Cyclical": "#e65100",
    "Consumer Defensive": "#558b2f", "Energy": "#ef6c00",
    "Industrials": "#00838f", "Basic Materials": "#1b7a3d",
    "Communication Services": "#6a1b9a", "Real Estate": "#c62828",
    "Utilities": "#283593", "Other": "#78909c",
}

# SP500_SECTOR_WEIGHTS and SECTOR_ETF_MAP imported lazily from engine.report_generator

# ════════════════════════════════════════════════════════════════════════════
# SIDEBAR
# ════════════════════════════════════════════════════════════════════════════

_quick_settings = _load_settings_quick()

with st.sidebar:
    st.markdown(f"""
<div style="padding:8px 0 20px">
  <h2 style="font-family:'Cormorant Garamond',Georgia,serif;color:{DCM_GOLD};margin:0;
             font-size:22px;font-weight:700;letter-spacing:-0.02em">
    {_quick_settings["portfolio_name"].upper()}
  </h2>
  <p style="font-family:'JetBrains Mono',monospace;font-size:8px;margin:6px 0 0;
            text-transform:uppercase;letter-spacing:0.2em;color:{DCM_DIM};font-weight:400">
    Risk Analytics
  </p>
</div>
""", unsafe_allow_html=True)
    st.divider()

    # ── Sectioned navigation ──────────────────────────────────────────
    if "_nav_page" not in st.session_state:
        st.session_state["_nav_page"] = "Portfolio Overview"

    _NAV_SECTIONS = [
        ("OVERVIEW",     ["Portfolio Overview"]),
        ("SECTORS",      ["Holdings by Sector"]),
        ("RISK",         ["Risk Dashboard", "Monte Carlo", "Stress Tests", "Stock Analysis"]),
        ("BACKTESTING",  ["Backtesting"]),
        ("REPORTS",      ["Reports & Export"]),
    ]

    # Callback factory: when a radio in one section changes, clear the others
    def _make_nav_cb(section_key):
        def _cb():
            chosen = st.session_state.get(section_key)
            if chosen is not None:
                st.session_state["_nav_page"] = chosen
                # Reset other section radios to None
                for _sn, _ in _NAV_SECTIONS:
                    k = f"_navsec_{_sn}"
                    if k != section_key:
                        st.session_state[k] = None
        return _cb

    for _sec_name, _sec_pages in _NAV_SECTIONS:
        st.markdown(
            f'<p style="font-size:9px;text-transform:uppercase;letter-spacing:0.15em;'
            f'color:{DCM_DIM};margin:16px 0 0;font-weight:700;font-family:\'JetBrains Mono\','
            f'monospace;padding-bottom:4px;'
            f'border-bottom:2px solid;'
            f'border-image:linear-gradient(to right,{DCM_GOLD},transparent) 1'
            f'">{_sec_name}</p>',
            unsafe_allow_html=True,
        )
        _sec_key = f"_navsec_{_sec_name}"
        # Determine current value for this section's radio
        _cur = st.session_state["_nav_page"]
        _sec_val = _cur if _cur in _sec_pages else None
        _sec_idx = _sec_pages.index(_sec_val) if _sec_val in _sec_pages else None
        st.radio(
            _sec_name,
            options=_sec_pages,
            index=_sec_idx,
            key=_sec_key,
            on_change=_make_nav_cb(_sec_key),
            label_visibility="collapsed",
        )

    page = st.session_state["_nav_page"]

    st.divider()
    refresh_clicked = st.button("Refresh Data", use_container_width=True)
    if refresh_clicked:
        st.rerun()

    # ── Portfolio upload ───────────────────────────────────────────────
    st.divider()

    # Download template
    st.download_button(
        "📥 Download Excel Template",
        data=_generate_template_bytes(),
        file_name="dcm_portfolio_template.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        use_container_width=True,
        help="Download a pre-formatted Excel template. Fill in your holdings and upload it below.",
    )

    # Manual entry
    if st.button("✏️ Enter Holdings Manually", use_container_width=True,
                 help="Enter your portfolio holdings directly without an Excel file"):
        st.session_state["_util_page"] = "Manual Entry"
        st.rerun()

    st.caption("— or upload a filled template —")

    uploaded_file = st.file_uploader(
        "Upload Portfolio (.xlsx)",
        type=["xlsx"],
        help="Upload the filled Excel template (Holdings sheet: Ticker, Shares Held, Cost Basis).",
        label_visibility="collapsed",
    )
    if uploaded_file is not None:
        import hashlib, tempfile
        file_bytes = uploaded_file.read()
        file_hash = hashlib.md5(file_bytes).hexdigest()
        if st.session_state.get("uploaded_portfolio_hash") != file_hash:
            tmp_dir = Path(tempfile.gettempdir()) / "dcm_uploads"
            tmp_dir.mkdir(exist_ok=True)
            tmp_path = tmp_dir / "uploaded_portfolio.xlsx"
            tmp_path.write_bytes(file_bytes)
            st.session_state["uploaded_portfolio_path"] = str(tmp_path)
            st.session_state["uploaded_portfolio_hash"] = file_hash
            st.rerun()

    if "uploaded_portfolio_hash" in st.session_state:
        st.caption("✓ Using uploaded portfolio")
        if st.button("Reset to default", use_container_width=True):
            st.session_state.pop("uploaded_portfolio_path", None)
            st.session_state.pop("uploaded_portfolio_hash", None)
            st.rerun()

    # ── Save portfolio to database ──────────────────────────────────────
    _active_user = current_user()
    if _active_user:
        st.divider()

        # Save button
        _port_path = st.session_state.get("uploaded_portfolio_path")
        if _port_path and Path(_port_path).exists():
            if st.button("💾 Save Portfolio", use_container_width=True,
                         help="Save your current portfolio to your account so it loads automatically next time."):
                try:
                    _h_list, _h_name, _h_short, _h_sets = excel_path_to_holdings(_port_path)
                    _ok, _msg = _db_save_portfolio(
                        _active_user["id"], _h_name, _h_short, _h_list, _h_sets
                    )
                    if _ok:
                        st.success("Portfolio saved to your account.")
                    else:
                        st.error(_msg)
                except Exception as _exc:
                    st.error(f"Save error: {_exc}")
        else:
            st.caption("Upload or enter a portfolio to enable saving.")

        # Download saved copy
        if _db_has_portfolio(_active_user["id"]):
            _saved_p = _db_load_portfolio(_active_user["id"])
            if _saved_p:
                st.download_button(
                    "⬇ Download My Saved Portfolio",
                    data=holdings_to_excel_bytes(
                        _saved_p["portfolio_name"],
                        _saved_p["short_name"],
                        _saved_p["holdings"],
                        _saved_p["settings"],
                    ),
                    file_name=f"{_saved_p['short_name']}_portfolio.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    use_container_width=True,
                )

        # User info + logout
        st.divider()
        _uname = _active_user.get("full_name") or _active_user["username"]
        st.markdown(
            f'<div style="font-family:\'JetBrains Mono\',monospace;font-size:9px;'
            f'color:{DCM_DIM};margin-bottom:6px;text-transform:uppercase;letter-spacing:0.1em">'
            f'Signed in as<br>'
            f'<span style="color:{DCM_GOLD};font-weight:700">{_uname}</span>'
            f'</div>',
            unsafe_allow_html=True,
        )
        if st.button("Sign Out", use_container_width=True):
            logout()

    st.markdown(f"""
<div style="font-family:'JetBrains Mono',monospace;font-size:9px;color:{DCM_DIM};
            margin-top:20px;padding-top:12px;border-top:1px solid {DCM_BORDER}">
  Live data &middot; 2Y lookback &middot; 10K Monte Carlo paths
</div>
""", unsafe_allow_html=True)

# ════════════════════════════════════════════════════════════════════════════
# TOP-RIGHT UTILITY BUTTONS  (Getting Started + Settings)
# ════════════════════════════════════════════════════════════════════════════

# Handle utility button navigation via query params
_qp_nav = st.query_params.get("_dcm_nav", None)
if _qp_nav == "guide":
    st.session_state["_util_page"] = "Getting Started"
    st.query_params.clear()
    st.rerun()
elif _qp_nav == "settings":
    st.session_state["_util_page"] = "Settings"
    st.query_params.clear()
    st.rerun()

_util_page = st.session_state.get("_util_page", None)

# Polished equal-width pill buttons rendered as HTML anchor links
st.markdown("""
<div class="util-bar">
  <a href="?_dcm_nav=guide" target="_self" class="util-btn">
    <svg xmlns="http://www.w3.org/2000/svg" width="14" height="14" viewBox="0 0 24 24"
         fill="none" stroke="currentColor" stroke-width="2" stroke-linecap="round"
         stroke-linejoin="round">
      <path d="M2 3h6a4 4 0 0 1 4 4v14a3 3 0 0 0-3-3H2z"/>
      <path d="M22 3h-6a4 4 0 0 0-4 4v14a3 3 0 0 1 3-3h7z"/>
    </svg>
    Guide
  </a>
  <a href="?_dcm_nav=settings" target="_self" class="util-btn">
    <svg xmlns="http://www.w3.org/2000/svg" width="14" height="14" viewBox="0 0 24 24"
         fill="none" stroke="currentColor" stroke-width="2" stroke-linecap="round"
         stroke-linejoin="round">
      <circle cx="12" cy="12" r="3"/>
      <path d="M19.4 15a1.65 1.65 0 0 0 .33 1.82l.06.06a2 2 0 0 1-2.83 2.83l-.06-.06
               a1.65 1.65 0 0 0-1.82-.33 1.65 1.65 0 0 0-1 1.51V21a2 2 0 0 1-4 0v-.09
               A1.65 1.65 0 0 0 9 19.4a1.65 1.65 0 0 0-1.82.33l-.06.06a2 2 0 0 1-2.83-2.83
               l.06-.06A1.65 1.65 0 0 0 4.68 15a1.65 1.65 0 0 0-1.51-1H3a2 2 0 0 1 0-4h.09
               A1.65 1.65 0 0 0 4.6 9a1.65 1.65 0 0 0-.33-1.82l-.06-.06a2 2 0 0 1 2.83-2.83
               l.06.06A1.65 1.65 0 0 0 9 4.68a1.65 1.65 0 0 0 1-1.51V3a2 2 0 0 1 4 0v.09
               a1.65 1.65 0 0 0 1 1.51 1.65 1.65 0 0 0 1.82-.33l.06-.06a2 2 0 0 1 2.83 2.83
               l-.06.06A1.65 1.65 0 0 0 19.4 9a1.65 1.65 0 0 0 1.51 1H21a2 2 0 0 1 0 4h-.09
               a1.65 1.65 0 0 0-1.51 1z"/>
    </svg>
    Settings
  </a>
</div>
""", unsafe_allow_html=True)

# Utility page overrides sidebar selection
if _util_page:
    page = _util_page

# ════════════════════════════════════════════════════════════════════════════
# PAGE 0 — GETTING STARTED
# ════════════════════════════════════════════════════════════════════════════

if page == "Getting Started":
    if st.button("\u2190 Back to Dashboard", key="_gs_back"):
        st.session_state.pop("_util_page", None)
        st.rerun()
    st.title("Getting Started")
    st.caption(f"{_quick_settings['portfolio_name']} Risk Analytics Platform")

    st.markdown(f"""
<div style="font-family:'Source Sans 3',sans-serif;font-size:16px;color:{DCM_INK};
            line-height:1.7;max-width:720px;margin-top:12px">

<p>This platform runs <strong>institutional-grade risk analytics</strong> on any stock portfolio.
Upload your holdings, and the engine fetches live market data, computes risk metrics,
runs Monte Carlo simulations, and stress-tests your portfolio against historical and
hypothetical scenarios.</p>

</div>
""", unsafe_allow_html=True)

    section_header("How It Works")

    c1, c2, c3 = st.columns(3)
    c1.markdown(metric_card("Step 1", "Upload", "Drag-and-drop your .xlsx file in the sidebar, or edit data/portfolio_holdings.xlsx directly.", DCM_GOLD), unsafe_allow_html=True)
    c2.markdown(metric_card("Step 2", "Analyze", "Navigate to any page. Live prices are fetched, and all analytics run automatically.", DCM_BLUE), unsafe_allow_html=True)
    c3.markdown(metric_card("Step 3", "Export", "Generate HTML reports or write results back to Excel from the Reports page.", DCM_GREEN), unsafe_allow_html=True)

    section_header("Required Excel Format")

    st.markdown("""
Your Excel file needs a **Holdings** sheet with these columns:

| Column | Required | Example |
|---|---|---|
| **Ticker** | Yes | AAPL |
| **Shares Held** | Yes | 500 |
| **Cost Basis** | Yes | 145.00 |
| Company Name | No (auto-populated) | Apple Inc. |
| Sector | No (auto-populated) | Technology |

Column names are matched by substring — "Ticker Symbol", "Total Shares", "Per-Share Cost" all work.
    """)

    section_header("Dashboard Pages")

    pages_info = [
        ("Portfolio Overview", "Sector allocation, weight distribution, top holdings, concentration metrics (HHI, ENB)"),
        ("Holdings by Sector", "Complete breakdown of every holding grouped by GICS sector, with per-sector totals and weights"),
        ("Risk Dashboard", "VaR/CVaR at 95% and 99%, Cornish-Fisher VaR, rolling volatility, drawdown analysis, Sharpe/Sortino/Calmar ratios"),
        ("Monte Carlo", "10,000 Cholesky-correlated simulation paths, fan chart, terminal value distribution, probability analysis"),
        ("Stress Tests", "4 historical scenarios (GFC, COVID, Rate Shock, Dot-Com) + 5 hypothetical, with stock-level breakdown"),
        ("Stock Analysis", "Per-holding risk table, component VaR chart (Euler decomposition), risk-return scatter, correlation heatmap"),
        ("Reports & Export", "Generate self-contained HTML reports, write results to Excel, view data quality and session info"),
    ]
    for name, desc in pages_info:
        st.markdown(f"**{name}** — {desc}")

    st.markdown("**Settings** *(gear icon, top-right)* — Configure portfolio name, benchmark, simulation parameters, and other preferences")

    section_header("Quick Start")

    st.markdown("""
1. Use the **Upload Portfolio** widget in the sidebar to upload your Excel file
2. Or edit `data/portfolio_holdings.xlsx` directly and click **Refresh Data**
3. Navigate to **Portfolio Overview** to begin

The engine fetches live prices from Yahoo Finance on every page load. First load may take 15-30 seconds.
    """)

    st.stop()

# ════════════════════════════════════════════════════════════════════════════
# PAGE — MANUAL ENTRY  (no analytics data needed)
# ════════════════════════════════════════════════════════════════════════════

if page == "Manual Entry":
    if st.button("← Back to Dashboard", key="_me_back"):
        st.session_state.pop("_util_page", None)
        st.rerun()

    st.title("Enter Portfolio Manually")
    st.caption("Add your holdings one by one, then click **Load Portfolio** to run all analytics.")

    # ── Instructions ──────────────────────────────────────────────────────
    with st.expander("How to use this form", expanded=False):
        st.markdown("""
**Instructions:**
- Enter each stock holding as a row in the table below.
- **Ticker** — the stock's exchange symbol, e.g. `AAPL`, `MSFT`, `GOOGL`. Must be a valid US-listed ticker.
- **Shares Held** — the number of shares you own (can be fractional, e.g. 1.5).
- **Cost Basis** — the price *per share* you originally paid. Used for unrealised P&L display only; it does not affect any risk calculations.
- Click the **+** row at the bottom of the table to add more holdings.
- When done, click **Load Portfolio** — the tool will fetch live prices and compute all analytics.
- Optionally fill in a portfolio name below.
        """)

    # ── Portfolio name ────────────────────────────────────────────────────
    me_port_name  = st.text_input("Portfolio name (optional)", value="My Portfolio",
                                   help="Shown in the dashboard header and reports")
    me_short_name = st.text_input("Short name / ticker (optional)", value="PORT", max_chars=10,
                                   help="Compact name used in tables, e.g. DCM, PORT")

    # ── Holdings editor ───────────────────────────────────────────────────
    st.markdown("#### Holdings")

    _default_rows = pd.DataFrame({
        "Ticker":      ["AAPL", "MSFT", "GOOGL", "AMZN", ""],
        "Shares Held": [100.0,   80.0,   40.0,   60.0,  0.0],
        "Cost Basis":  [145.50, 280.00, 2800.00, 155.00, 0.0],
    })

    # Load existing manual entry state if available
    if "manual_entry_df" not in st.session_state:
        st.session_state["manual_entry_df"] = _default_rows.copy()

    edited_df = st.data_editor(
        st.session_state["manual_entry_df"],
        num_rows="dynamic",
        use_container_width=True,
        column_config={
            "Ticker": st.column_config.TextColumn(
                "Ticker", help="US-listed stock symbol (e.g. AAPL)", max_chars=10,
                validate=r"^[A-Za-z]{1,10}$",
            ),
            "Shares Held": st.column_config.NumberColumn(
                "Shares Held", help="Number of shares owned", min_value=0.0, format="%.4f",
            ),
            "Cost Basis": st.column_config.NumberColumn(
                "Cost Basis ($)", help="Price per share paid", min_value=0.0, format="$%.2f",
            ),
        },
        hide_index=True,
        key="manual_entry_editor",
    )
    st.session_state["manual_entry_df"] = edited_df

    # ── Validation summary ────────────────────────────────────────────────
    valid_rows = edited_df[
        edited_df["Ticker"].str.strip().str.len().gt(0).fillna(False)
    ].dropna(subset=["Ticker", "Shares Held"]) if not edited_df.empty else pd.DataFrame()
    valid_rows = valid_rows[valid_rows["Shares Held"] > 0] if not valid_rows.empty else valid_rows

    if not valid_rows.empty:
        st.caption(f"✓ {len(valid_rows)} valid holding(s) ready to load: "
                   f"{', '.join(valid_rows['Ticker'].str.upper().tolist()[:8])}"
                   + (" …" if len(valid_rows) > 8 else ""))
    else:
        st.caption("Add at least one holding with a ticker and positive share count to continue.")

    st.divider()

    col_load, col_clear, col_dl = st.columns([3, 1.5, 1.5])

    with col_load:
        load_clicked = st.button(
            "🚀 Load Portfolio & Run Analytics",
            use_container_width=True,
            type="primary",
            disabled=valid_rows.empty,
        )

    with col_clear:
        if st.button("🗑 Clear Table", use_container_width=True):
            st.session_state["manual_entry_df"] = pd.DataFrame({
                "Ticker": [""], "Shares Held": [0.0], "Cost Basis": [0.0],
            })
            st.rerun()

    with col_dl:
        if not valid_rows.empty:
            # Offer download of the table as Excel so they can save it for next time
            import io as _io
            _dl_buf = _io.BytesIO()
            with pd.ExcelWriter(_dl_buf, engine="openpyxl") as _xw:
                valid_rows[["Ticker", "Shares Held", "Cost Basis"]].to_excel(
                    _xw, sheet_name="Holdings", index=False,
                )
                _settings_df = pd.DataFrame([
                    ["portfolio_name",  me_port_name],
                    ["portfolio_short_name", me_short_name],
                    ["benchmark_ticker", "SPY"],
                    ["risk_free_rate", "auto"],
                    ["confidence_level_1", 0.95],
                    ["confidence_level_2", 0.99],
                    ["es_confidence_level", 0.975],
                    ["lookback_years", 2],
                    ["simulation_paths", 10000],
                    ["simulation_days", 252],
                    ["covariance_mode", "ledoit_wolf"],
                    ["stress_custom_drawdown", -0.20],
                ], columns=["Parameter", "Value"])
                _settings_df.to_excel(_xw, sheet_name="Settings", index=False)
            _dl_buf.seek(0)
            st.download_button(
                "💾 Save as Excel",
                data=_dl_buf.getvalue(),
                file_name="my_portfolio.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                use_container_width=True,
                help="Download your current entries as a filled Excel file you can re-upload later",
            )

    if load_clicked and not valid_rows.empty:
        import io as _io, hashlib as _hl, tempfile as _tf
        with st.spinner("Building portfolio file and loading analytics…"):
            # Build an Excel in memory
            _buf = _io.BytesIO()
            with pd.ExcelWriter(_buf, engine="openpyxl") as _xw:
                valid_rows[["Ticker", "Shares Held", "Cost Basis"]].to_excel(
                    _xw, sheet_name="Holdings", index=False,
                )
                _settings_df = pd.DataFrame([
                    ["portfolio_name",  me_port_name],
                    ["portfolio_short_name", me_short_name],
                    ["benchmark_ticker", "SPY"],
                    ["risk_free_rate", "auto"],
                    ["confidence_level_1", 0.95],
                    ["confidence_level_2", 0.99],
                    ["es_confidence_level", 0.975],
                    ["lookback_years", 2],
                    ["simulation_paths", 10000],
                    ["simulation_days", 252],
                    ["covariance_mode", "ledoit_wolf"],
                    ["ewma_lambda", 0.94],
                    ["mc_shock_distribution", "normal"],
                    ["mc_df", 7],
                    ["stress_custom_drawdown", -0.20],
                    ["report_title", f"{me_port_name} — Portfolio Risk Report"],
                ], columns=["Parameter", "Value"])
                _settings_df.to_excel(_xw, sheet_name="Settings", index=False)
            _buf.seek(0)
            _file_bytes = _buf.getvalue()
            _file_hash  = _hl.md5(_file_bytes).hexdigest()

            _tmp_dir = Path(_tf.gettempdir()) / "dcm_uploads"
            _tmp_dir.mkdir(exist_ok=True)
            _tmp_path = _tmp_dir / "uploaded_portfolio.xlsx"
            _tmp_path.write_bytes(_file_bytes)

            st.session_state["uploaded_portfolio_path"] = str(_tmp_path)
            st.session_state["uploaded_portfolio_hash"] = _file_hash
            st.session_state.pop("_util_page", None)
            st.rerun()

    st.stop()


# ════════════════════════════════════════════════════════════════════════════
# PAGE — SETTINGS  (no analytics data needed)
# ════════════════════════════════════════════════════════════════════════════

if page == "Settings":
    if st.button("\u2190 Back to Dashboard", key="_set_back"):
        st.session_state.pop("_util_page", None)
        st.rerun()
    import importlib
    import engine.data_loader as _dl_mod
    importlib.reload(_dl_mod)
    PortfolioSettings = _dl_mod.PortfolioSettings
    save_settings = _dl_mod.save_settings

    st.title("Settings")
    st.caption("Configure portfolio parameters and preferences")

    # Load current settings directly from Excel (no analytics pipeline needed)
    _settings_excel = (
        Path(st.session_state["uploaded_portfolio_path"])
        if "uploaded_portfolio_path" in st.session_state
        else get_portfolio_path()
    )
    try:
        _lr = load_portfolio(excel_path=_settings_excel)
        current = _lr.settings
    except Exception:
        current = PortfolioSettings()

    # ── Organization / Identity ────────────────────────────────────────
    section_header("Organization")

    col1, col2 = st.columns(2)
    with col1:
        new_name = st.text_input(
            "Portfolio / Organization Name",
            value=current.portfolio_name,
            help="Full name displayed in sidebar, report headers, and page captions",
        )
    with col2:
        new_short = st.text_input(
            "Short Name / Abbreviation",
            value=current.portfolio_short_name,
            help="Used in table headers, filenames, and compact labels (e.g. DCM)",
        )

    new_title = st.text_input(
        "Report Title",
        value=current.report_title,
        help="Title displayed at the top of generated HTML reports",
    )

    # ── Market Parameters ──────────────────────────────────────────────
    section_header("Market Parameters")

    col1, col2 = st.columns(2)
    with col1:
        new_benchmark = st.text_input(
            "Benchmark Ticker",
            value=current.benchmark_ticker,
            help="SPY for S&P 500, QQQ for Nasdaq 100, IWM for Russell 2000",
        )
    with col2:
        new_rfr = st.text_input(
            "Risk-Free Rate",
            value=current.risk_free_rate,
            help="'auto' to use 10-Year Treasury, or a decimal like 0.043 for 4.3%",
        )

    # ── Risk Parameters ────────────────────────────────────────────────
    section_header("Risk Parameters")

    col1, col2, col3 = st.columns(3)
    with col1:
        new_cl1 = st.number_input(
            "Confidence Level 1", value=current.confidence_level_1,
            min_value=0.50, max_value=0.999, step=0.01, format="%.2f",
        )
    with col2:
        new_cl2 = st.number_input(
            "Confidence Level 2", value=current.confidence_level_2,
            min_value=0.50, max_value=0.999, step=0.01, format="%.2f",
        )
    with col3:
        new_lookback = st.number_input(
            "Lookback Years", value=current.lookback_years,
            min_value=1, max_value=10, step=1,
        )

    # ── Monte Carlo Simulation ─────────────────────────────────────────
    section_header("Monte Carlo Simulation")

    col1, col2 = st.columns(2)
    with col1:
        new_paths = st.number_input(
            "Simulation Paths", value=current.simulation_paths,
            min_value=1000, max_value=100000, step=1000,
        )
    with col2:
        new_days = st.number_input(
            "Simulation Days", value=current.simulation_days,
            min_value=21, max_value=504, step=21,
        )

    # ── Stress Testing ─────────────────────────────────────────────────
    section_header("Stress Testing")

    _dd_pct = st.slider(
        "Custom Stress Drawdown",
        min_value=-50, max_value=-5,
        value=int(current.stress_custom_drawdown * 100),
        step=5, format="%d%%",
        help="Uniform market drawdown applied in the custom stress scenario",
    )
    new_drawdown = _dd_pct / 100.0

    # ── Advanced ───────────────────────────────────────────────────────
    with st.expander("Advanced Settings"):
        col1, col2 = st.columns(2)
        with col1:
            new_maxpos = st.number_input(
                "Max Position Warning %", value=current.max_position_warning_pct,
                min_value=0.01, max_value=1.0, step=0.01, format="%.2f",
            )
        with col2:
            new_minpts = st.number_input(
                "Min Data Points", value=current.min_data_points,
                min_value=20, max_value=500, step=10,
            )

    # ── Save ───────────────────────────────────────────────────────────
    st.divider()
    if st.button("Save Settings", type="primary", use_container_width=True):
        updated = PortfolioSettings(
            portfolio_name=new_name.strip(),
            portfolio_short_name=new_short.strip(),
            benchmark_ticker=new_benchmark.strip().upper(),
            risk_free_rate=new_rfr.strip(),
            confidence_level_1=new_cl1,
            confidence_level_2=new_cl2,
            lookback_years=int(new_lookback),
            simulation_paths=int(new_paths),
            simulation_days=int(new_days),
            stress_custom_drawdown=new_drawdown,
            report_title=new_title.strip(),
            color_primary=current.color_primary,
            color_secondary=current.color_secondary,
            color_accent=current.color_accent,
            max_position_warning_pct=new_maxpos,
            min_data_points=int(new_minpts),
        )
        try:
            save_settings(updated, excel_path=_settings_excel)
            st.success("Settings saved. Navigate to any page to see changes take effect.")
            st.rerun()
        except PermissionError as e:
            st.error(str(e))
        except Exception as e:
            st.error(f"Failed to save settings: {e}")

    st.stop()

# ════════════════════════════════════════════════════════════════════════════
# LOAD DATA  (with spinner — only for analytics pages)
# ════════════════════════════════════════════════════════════════════════════

with st.spinner("Loading portfolio data and running analytics…"):
    try:
        data = get_data()
    except ConnectionError:
        st.error("No internet connection. Live market data requires a working internet connection to fetch prices from Yahoo Finance.")
        st.stop()
    except ValueError as e:
        st.error(f"Data validation error: {e}")
        st.stop()
    except Exception as e:
        st.error(f"Failed to load data: {e}")
        st.stop()

md       = data["md"]
metrics  = data["metrics"]
sim      = data["sim"]
stress   = data["stress"]
mh_es    = data.get("mh_es")
backtest = data.get("backtest")
settings = data["load_result"].settings

# ── Floating AI chatbot widget (injected once; persists across page navigations)
render_chatbot_widget(data)

# ── Warn about failed tickers ──────────────────────────────────────────────
if md.quality.failed_tickers:
    failed_list = ", ".join(md.quality.failed_tickers)
    st.warning(
        f"**{len(md.quality.failed_tickers)} ticker(s) could not be fetched:** {failed_list}. "
        f"These holdings are excluded from the analysis. Check that the ticker symbols are correct "
        f"and currently listed on a major exchange."
    )

# ════════════════════════════════════════════════════════════════════════════
# PAGE 1 — PORTFOLIO OVERVIEW
# ════════════════════════════════════════════════════════════════════════════

if page == "Portfolio Overview":
    st.title("Portfolio Overview")
    st.caption(f"{settings.portfolio_name} · {pd.Timestamp.now().strftime('%B %d, %Y')}")

    # ── Top KPIs ──────────────────────────────────────────────────────────
    pv   = md.total_portfolio_value
    n    = len(md.holdings)
    ann_ret  = metrics.annualized_return
    ann_vol  = metrics.annualized_vol
    sharpe   = metrics.sharpe
    beta     = metrics.beta

    cols = st.columns(5)
    kpis = [
        ("Portfolio Value",  fmt_dollar(pv),       f"{n} holdings",                DCM_BLUE,  None),
        ("Ann. Return",      fmt_pct(ann_ret),     "2-year history",     DCM_GREEN if ann_ret > 0 else DCM_RED, "up" if ann_ret > 0 else "down"),
        ("Ann. Volatility",  fmt_pct(ann_vol),     "Std dev × √252",              DCM_GOLD,  None),
        ("Sharpe Ratio",     f"{sharpe:.2f}",      "Excess return / vol (Rfr=auto)", DCM_BLUE, "up" if sharpe > 0 else "down"),
        ("Portfolio Beta",   f"{beta:.2f}",        "Relative to SPY",             DCM_BLUE,  None),
    ]
    for col, (lbl, val, sub, clr, trnd) in zip(cols, kpis):
        col.markdown(metric_card(lbl, val, sub, clr, trend=trnd), unsafe_allow_html=True)

    st.write("")

    # ── Sector allocation + top holdings ──────────────────────────────────
    left, right = st.columns([1, 1])

    with left:
        section_header("Sector Allocation vs S&P 500")
        from engine.report_generator import SP500_SECTOR_WEIGHTS as _SP500_SW
        # Portfolio sector weights
        port_sector_wt: dict[str, float] = {}
        for h in md.holdings:
            port_sector_wt[h.sector] = port_sector_wt.get(h.sector, 0.0) + h.weight
        # Union of sectors, sorted by S&P 500 weight descending
        all_sectors = sorted(
            set(port_sector_wt.keys()) | set(_SP500_SW.keys()),
            key=lambda s: _SP500_SW.get(s, 0.0), reverse=True,
        )
        port_vals = [port_sector_wt.get(s, 0.0) * 100 for s in all_sectors]
        sp_vals = [_SP500_SW.get(s, 0.0) * 100 for s in all_sectors]

        fig_sector = go.Figure()
        fig_sector.add_trace(go.Bar(
            name="Portfolio", x=all_sectors, y=port_vals,
            marker_color=[SECTOR_COLORS.get(s, "#78909c") for s in all_sectors],
            hovertemplate="<b>%{x}</b><br>Portfolio: %{y:.1f}%<extra></extra>",
        ))
        fig_sector.add_trace(go.Bar(
            name="S&P 500", x=all_sectors, y=sp_vals,
            marker_color="rgba(180,180,180,0.5)",
            marker_line=dict(color="#999", width=1),
            hovertemplate="<b>%{x}</b><br>S&P 500: %{y:.1f}%<extra></extra>",
        ))
        fig_sector.update_layout(
            barmode="group", bargap=0.25, bargroupgap=0.05,
            yaxis_title="Weight (%)",
            height=340, margin=dict(t=10, b=80, l=50, r=10),
            legend=dict(orientation="h", yanchor="bottom", y=1.02, xanchor="right", x=1),
            xaxis_tickangle=-35, xaxis_tickfont_size=10,
        )
        st.plotly_chart(fig_sector, use_container_width=True, config=PLOTLY_CONFIG)

    with right:
        section_header("Top 15 Holdings by Weight")
        top15 = sorted(md.holdings, key=lambda h: -h.weight)[:15]
        df_top = pd.DataFrame([{
            "Ticker": h.ticker,
            "Company": h.company_name[:22] + "…" if len(h.company_name) > 22 else h.company_name,
            "Sector": h.sector[:20],
            "Price": f"${h.current_price:.2f}",
            "Shares": f"{h.shares_held:,}",
            "Value": f"${h.market_value:,.0f}",
            "Weight": f"{h.weight:.1%}",
        } for h in top15])
        st.dataframe(df_top, hide_index=True, use_container_width=True, height=320)

    # ── Weight bar chart ──────────────────────────────────────────────────
    section_header("Portfolio Weight Distribution")
    sorted_h = sorted(md.holdings, key=lambda h: -h.weight)
    fig_bar = go.Figure(go.Bar(
        x=[h.ticker for h in sorted_h],
        y=[h.weight * 100 for h in sorted_h],
        marker_color=[SECTOR_COLORS.get(h.sector, "#78909c") for h in sorted_h],
        hovertemplate="<b>%{x}</b><br>Weight: %{y:.2f}%<extra></extra>",
    ))
    equal_w = 100 / len(md.holdings)
    fig_bar.add_hline(y=equal_w, line_dash="dash", line_color=DCM_GRAY,
                      annotation_text=f"Equal-weight {equal_w:.1f}%",
                      annotation_position="top right")
    fig_bar.update_layout(
        xaxis_title="Ticker", yaxis_title="Weight (%)",
        height=280, margin=dict(t=10, b=50, l=50, r=20),

        bargap=0.3,
    )
    st.plotly_chart(fig_bar, use_container_width=True, config=PLOTLY_CONFIG)

    # ── Concentration ─────────────────────────────────────────────────────
    section_header("Concentration & Diversification")
    c1, c2, c3, c4 = st.columns(4)
    top5_w  = sum(h.weight for h in sorted(md.holdings, key=lambda h:-h.weight)[:5])
    top10_w = sum(h.weight for h in sorted(md.holdings, key=lambda h:-h.weight)[:10])
    cards = [
        ("HHI Index",          f"{metrics.hhi:.4f}",            "1/n = equal-weight ideal",         DCM_GOLD),
        ("Eff. Num. Bets",     f"{metrics.eff_num_bets:.1f}",   f"of {n} holdings",                  DCM_BLUE),
        ("Top-5 Concentration",f"{top5_w:.1%}",                  "5 largest positions",               DCM_GOLD if top5_w > 0.40 else DCM_GREEN),
        ("Top-10 Concentration",f"{top10_w:.1%}",               "10 largest positions",              DCM_GOLD if top10_w > 0.60 else DCM_GREEN),
    ]
    for col, (lbl, val, sub, clr) in zip([c1,c2,c3,c4], cards):
        col.markdown(metric_card(lbl, val, sub, clr), unsafe_allow_html=True)

# ════════════════════════════════════════════════════════════════════════════
# PAGE 1b — HOLDINGS BY SECTOR
# ════════════════════════════════════════════════════════════════════════════

elif page == "Holdings by Sector":
    import importlib, engine.report_generator as _rg_mod
    importlib.reload(_rg_mod)
    SP500_SECTOR_WEIGHTS = _rg_mod.SP500_SECTOR_WEIGHTS
    SECTOR_ETF_MAP = _rg_mod.SECTOR_ETF_MAP

    st.title("Holdings by Sector")
    st.caption(f"{settings.portfolio_name} \u00b7 {pd.Timestamp.now().strftime('%B %d, %Y')}")

    pv = md.total_portfolio_value
    n  = len(md.holdings)

    # ── Build sector groups ────────────────────────────────────────────
    from collections import defaultdict
    sector_groups: dict[str, list] = defaultdict(list)
    for h in md.holdings:
        sector_groups[h.sector].append(h)

    # Sort sectors by total market value (descending)
    sector_order = sorted(
        sector_groups.keys(),
        key=lambda s: sum(h.market_value for h in sector_groups[s]),
        reverse=True,
    )

    # ── KPI row ────────────────────────────────────────────────────────
    n_sectors = len(sector_order)
    largest_sector = sector_order[0] if sector_order else "N/A"
    largest_sector_wt = (
        sum(h.weight for h in sector_groups[largest_sector])
        if sector_order else 0.0
    )
    smallest_sector = sector_order[-1] if sector_order else "N/A"
    smallest_sector_wt = (
        sum(h.weight for h in sector_groups[smallest_sector])
        if sector_order else 0.0
    )

    cols = st.columns(5)
    kpis = [
        ("Portfolio Value",     fmt_dollar(pv),                     f"{n} holdings",       DCM_BLUE,  None),
        ("Active Sectors",      str(n_sectors),                     "of 11 GICS sectors",  DCM_GOLD,  None),
        ("Largest Sector",      f"{largest_sector_wt:.1%}",         largest_sector,        DCM_BLUE,  None),
        ("Smallest Sector",     f"{smallest_sector_wt:.1%}",        smallest_sector,       DCM_GOLD,  None),
        ("Avg Holdings/Sector", f"{n / n_sectors:.1f}" if n_sectors else "0", "Per active sector", DCM_MUTED, None),
    ]
    for col, (lbl, val, sub, clr, trnd) in zip(cols, kpis):
        col.markdown(metric_card(lbl, val, sub, clr, trend=trnd), unsafe_allow_html=True)

    st.write("")

    # ── Sector filter (controls both comparison table and holdings tables) ─
    # Combine all S&P 500 sectors + any portfolio-only sectors for the filter
    all_sectors = list(SP500_SECTOR_WEIGHTS.keys())
    for s in sector_order:
        if s not in all_sectors:
            all_sectors.append(s)

    selected_sectors = st.multiselect(
        "Filter sectors",
        options=all_sectors,
        default=all_sectors,
        help="Filter which sectors appear in the breakdown and holdings tables below.",
    )

    # ── Sector Breakdown vs S&P 500 ───────────────────────────────────
    section_header(f"{settings.portfolio_short_name} Sector Breakdown")

    # Build portfolio weight per sector
    port_sector_wt: dict[str, float] = {}
    for h in md.holdings:
        port_sector_wt[h.sector] = port_sector_wt.get(h.sector, 0.0) + h.weight

    # Build comparison rows (only for selected sectors that have S&P weights)
    comp_rows = []
    for sector, sp_wt in SP500_SECTOR_WEIGHTS.items():
        if sector not in selected_sectors:
            continue
        p_wt = port_sector_wt.get(sector, 0.0)
        diff = p_wt - sp_wt
        target_dollars = (sp_wt - p_wt) * pv
        target_str = f"-${abs(target_dollars):,.0f}" if target_dollars < 0 else f"${target_dollars:,.0f}"
        comp_rows.append({
            "Sector": sector,
            "S&P 500 (%)": sp_wt * 100,
            "Portfolio (%)": p_wt * 100,
            "Difference (%)": diff * 100,
            "Target Investment ($)": target_str,
        })

    df_comp = pd.DataFrame(comp_rows).sort_values("Difference (%)", ascending=False)
    styled_comp = df_comp.style.set_properties(
        subset=["Target Investment ($)"], **{"text-align": "right"}
    )
    st.dataframe(
        styled_comp,
        hide_index=True,
        use_container_width=True,
        column_config={
            "S&P 500 (%)": st.column_config.NumberColumn(format="%.2f%%"),
            "Portfolio (%)": st.column_config.NumberColumn(format="%.2f%%"),
            "Difference (%)": st.column_config.NumberColumn(format="%+.2f%%"),
        },
    )

    # ── Sector Returns Comparison ─────────────────────────────────────
    section_header("Sector Returns Comparison")

    with st.spinner("Fetching sector ETF data…"):
        etf_prices = _fetch_sector_etf_prices()

    ret_rows = []
    for sector in SP500_SECTOR_WEIGHTS:
        if sector not in selected_sectors:
            continue
        etf_ticker = SECTOR_ETF_MAP.get(sector)
        if not etf_ticker or etf_ticker not in etf_prices.columns:
            continue

        etf_series = etf_prices[etf_ticker].dropna()
        ret_6m = _period_return(etf_series, 6)
        ret_1y = _period_return(etf_series, 12)
        dcm_1y = _sector_portfolio_return(md, sector, 12)
        diff = (dcm_1y - ret_1y) if (dcm_1y is not None and ret_1y is not None) else None
        ret_rows.append({
            "Ticker": etf_ticker,
            "Sector": sector,
            "6-Mo Return (%)": ret_6m * 100 if ret_6m is not None else None,
            "1-Yr Return (%)": ret_1y * 100 if ret_1y is not None else None,
            f"1-Yr {settings.portfolio_short_name} (%)": dcm_1y * 100 if dcm_1y is not None else None,
            "Difference (%)": diff * 100 if diff is not None else None,
        })

    df_ret = pd.DataFrame(ret_rows)
    st.dataframe(
        df_ret,
        hide_index=True,
        use_container_width=True,
        column_config={
            "6-Mo Return (%)": st.column_config.NumberColumn(format="%.1f%%"),
            "1-Yr Return (%)": st.column_config.NumberColumn(format="%.1f%%"),
            f"1-Yr {settings.portfolio_short_name} (%)": st.column_config.NumberColumn(format="%.1f%%"),
            "Difference (%)": st.column_config.NumberColumn(format="%+.1f%%"),
        },
    )

    # ── Individual sector holdings tables ─────────────────────────────
    section_header("Sector Holdings")

    # ── Render each sector table ───────────────────────────────────────
    for idx, sector_name in enumerate(sector_order):
        if sector_name not in selected_sectors:
            continue

        holdings_in_sector = sorted(
            sector_groups[sector_name],
            key=lambda h: -h.market_value,
        )
        sector_mv = sum(h.market_value for h in holdings_in_sector)
        sector_wt = sector_mv / pv if pv > 0 else 0.0
        color = SECTOR_COLORS.get(sector_name, "#78909c")

        n_hold = len(holdings_in_sector)
        hold_word = "holdings" if n_hold != 1 else "holding"

        st.markdown(
            f'<div style="margin-top:16px">'
            f'<span style="display:inline-block;width:12px;height:12px;'
            f'border-radius:2px;background:{color};margin-right:8px;'
            f'vertical-align:middle"></span>'
            f'<b>{sector_name}</b>'
            f'<span style="margin-left:12px;color:var(--ink-dim,#888);font-size:13px">'
            f'{n_hold} {hold_word} · {sector_wt:.1%} of portfolio</span></div>',
            unsafe_allow_html=True,
        )

        sector_rows = []
        for h in holdings_in_sector:
            pct_of_sector = h.market_value / sector_mv if sector_mv > 0 else 0.0
            sector_rows.append({
                "Ticker": h.ticker,
                "Company": h.company_name,
                "Shares": f"{h.shares_held:,.0f}",
                "Price": f"${h.current_price:,.2f}",
                "Market Value": f"${h.market_value:,.0f}",
                "% of Sector": pct_of_sector * 100,
                "% of Portfolio": h.weight * 100,
            })
        df_sector = pd.DataFrame(sector_rows)
        styled_sector = df_sector.style.set_properties(
            subset=["Shares", "Price", "Market Value"], **{"text-align": "right"}
        )
        st.dataframe(
            styled_sector,
            hide_index=True,
            use_container_width=True,
            column_config={
                "% of Sector": st.column_config.NumberColumn(format="%.1f%%"),
                "% of Portfolio": st.column_config.NumberColumn(format="%.1f%%"),
            },
        )

# ════════════════════════════════════════════════════════════════════════════
# PAGE 2 — RISK DASHBOARD
# ════════════════════════════════════════════════════════════════════════════

elif page == "Risk Dashboard":
    st.title("Risk Dashboard")
    st.caption("All risk metrics computed on 2-year return history")

    pv   = md.total_portfolio_value
    var95 = metrics.var_95
    var99 = metrics.var_99
    rfr   = md.risk_free_rate

    # ── Covariance mode indicator ─────────────────────────────────────────
    cov_mode = getattr(settings, "covariance_mode", "ledoit_wolf")
    cov_label = "Ledoit-Wolf Shrinkage" if cov_mode == "ledoit_wolf" else f"EWMA (λ={getattr(settings,'ewma_lambda',0.94)})"
    es_conf = getattr(settings, "es_confidence_level", 0.975)
    var_es_result = getattr(metrics, "var_es", None)

    st.caption(f"Covariance estimator: **{cov_label}** · Primary ES confidence: **{es_conf:.1%}**")

    # ── KPI row ──────────────────────────────────────────────────────────
    cols = st.columns(7)
    es_975_val = fmt_dollar(var_es_result.parametric_cvar) if var_es_result else fmt_dollar(var95.parametric_cvar)
    kpis = [
        ("1-Day VaR (95%)",         fmt_dollar(var95.parametric_var),    "Parametric normal",        DCM_RED),
        ("CF-VaR (95%)",            fmt_dollar(var95.cornish_fisher_var), "Skew/kurt adjusted",       DCM_RED),
        (f"ES ({es_conf:.1%}) ★",   es_975_val,                          "Primary ES — FRTB standard", DCM_RED),
        ("1-Day VaR (99%)",         fmt_dollar(var99.parametric_var),    "Parametric",               "#8b0000"),
        ("Sharpe Ratio",            f"{metrics.sharpe:.2f}",             f"Rfr={rfr:.2%}",           DCM_BLUE),
        ("Sortino Ratio",           f"{metrics.sortino:.2f}",            "Downside deviation",        DCM_BLUE),
        ("Max Drawdown",            fmt_pct(metrics.max_drawdown),       f"{metrics.max_dd_duration}d duration", DCM_RED),
    ]
    for col, (lbl, val, sub, clr) in zip(cols, kpis):
        col.markdown(metric_card(lbl, val, sub, clr), unsafe_allow_html=True)

    st.write("")

    # ── VaR comparison ────────────────────────────────────────────────────
    left, right = st.columns(2)

    with left:
        section_header("VaR Comparison (1-Day, 95%)")
        methods = ["Parametric", "Historical"]
        values  = [var95.parametric_var, abs(var95.historical_var)]
        colors  = [DCM_BLUE, DCM_GOLD]
        fig_var = go.Figure(go.Bar(
            x=methods, y=values,
            marker_color=colors,
            text=[fmt_dollar(v) for v in values],
            textposition="outside",
            hovertemplate="<b>%{x}</b><br>VaR: $%{y:,.0f}<extra></extra>",
        ))
        fig_var.add_hline(y=var95.parametric_cvar, line_dash="dash", line_color=DCM_RED,
                          annotation_text=f"ES (95%) = {fmt_dollar(var95.parametric_cvar)}",
                          annotation_position="top left")
        fig_var.update_layout(
            yaxis_title="Dollar Loss at Risk", height=280,
            margin=dict(t=30, b=40, l=60, r=20),
    
            showlegend=False,
        )
        st.plotly_chart(fig_var, use_container_width=True, config=PLOTLY_CONFIG)

    with right:
        section_header("Return Ratios vs Benchmark")
        port_ret = md.portfolio_returns(log_returns=True)
        bench_ret = md.benchmark_returns(log_returns=True)
        shared_idx = port_ret.index.intersection(bench_ret.index)
        pr = port_ret.loc[shared_idx]
        br = bench_ret.loc[shared_idx]

        metrics_df = pd.DataFrame({
            "Metric": ["Ann. Return", "Ann. Volatility", "Sharpe Ratio",
                       "Sortino Ratio", "Beta", "Calmar Ratio"],
            "Portfolio": [
                fmt_pct(metrics.annualized_return), fmt_pct(metrics.annualized_vol),
                f"{metrics.sharpe:.2f}", f"{metrics.sortino:.2f}",
                f"{metrics.beta:.2f}", f"{metrics.calmar:.2f}",
            ],
        })
        st.dataframe(metrics_df, hide_index=True, use_container_width=True, height=235)

    # ── Rolling volatility ─────────────────────────────────────────────────
    section_header("Rolling 21-Day Annualized Volatility")
    port_ret = md.portfolio_returns(log_returns=True)
    roll_vol = port_ret.rolling(21).std() * np.sqrt(252)

    fig_vol = go.Figure()
    fig_vol.add_trace(go.Scatter(
        x=roll_vol.index, y=roll_vol.values * 100,
        line=dict(color=DCM_BLUE, width=2),
        fill="tozeroy", fillcolor="rgba(41,98,255,0.06)",
        name="Rolling Vol",
        hovertemplate="%{x|%b %d, %Y}<br>Vol: %{y:.1f}%<extra></extra>",
    ))
    fig_vol.add_hline(y=metrics.annualized_vol * 100, line_dash="dash",
                      line_color=DCM_GOLD, line_width=1.5,
                      annotation_text=f"Avg {metrics.annualized_vol:.1%}",
                      annotation_position="top right")
    fig_vol.update_layout(
        xaxis_title="", yaxis_title="Annualized Volatility (%)",
        height=240, margin=dict(t=10, b=40, l=60, r=20),

        showlegend=False,
    )
    st.plotly_chart(fig_vol, use_container_width=True, config=PLOTLY_CONFIG)

    # ── Drawdown ──────────────────────────────────────────────────────────
    section_header("Portfolio Drawdown")
    cum_ret  = (1 + port_ret.apply(lambda x: np.exp(x) - 1)).cumprod()
    roll_max = cum_ret.cummax()
    drawdown = (cum_ret - roll_max) / roll_max

    fig_dd = go.Figure()
    fig_dd.add_trace(go.Scatter(
        x=drawdown.index, y=drawdown.values * 100,
        fill="tozeroy", fillcolor="rgba(198,40,40,0.10)",
        line=dict(color=DCM_RED, width=1.5),
        hovertemplate="%{x|%b %d, %Y}<br>DD: %{y:.1f}%<extra></extra>",
    ))
    fig_dd.add_hline(y=metrics.max_drawdown * 100, line_dash="dash", line_color="#c62828",
                     line_width=1.5,
                     annotation_text=f"Max DD {metrics.max_drawdown:.1%}",
                     annotation_position="bottom right")
    fig_dd.update_layout(
        xaxis_title="", yaxis_title="Drawdown (%)",
        height=220, margin=dict(t=10, b=40, l=60, r=20),

        showlegend=False,
    )
    st.plotly_chart(fig_dd, use_container_width=True, config=PLOTLY_CONFIG)

    # ── Risk Summary table ────────────────────────────────────────────────
    section_header("Full Risk Metrics Summary")
    _es_975_row = (
        (f"1-Day ES ({es_conf:.1%}) ★ [Primary]", fmt_dollar(var_es_result.parametric_cvar), "FRTB standard — Expected Shortfall")
        if var_es_result else
        ("1-Day ES (97.5%) ★ [Primary]", "N/A", "ES at primary confidence level")
    )
    summary = [
        ("1-Day VaR 95% (Parametric)",    fmt_dollar(var95.parametric_var),        "Normal distribution assumption"),
        ("1-Day CF-VaR 95%",              fmt_dollar(var95.cornish_fisher_var),     "Cornish-Fisher skew/kurtosis adjusted"),
        ("1-Day VaR 95% (Historical)",    fmt_dollar(abs(var95.historical_var)),    "Empirical 5th percentile"),
        ("1-Day VaR 99% (Parametric)",    fmt_dollar(var99.parametric_var),        "Normal distribution assumption"),
        ("1-Day CF-VaR 99%",              fmt_dollar(var99.cornish_fisher_var),     "Cornish-Fisher skew/kurtosis adjusted"),
        ("1-Day ES 95% (Exp. Shortfall)", fmt_dollar(var95.parametric_cvar),       "Expected Shortfall — parametric"),
        _es_975_row,
        ("1-Day ES 99% (Exp. Shortfall)", fmt_dollar(var99.parametric_cvar),       "Expected Shortfall — parametric"),
        ("Annualized Volatility",       fmt_pct(metrics.annualized_vol),   "Daily std × √252"),
        ("Annualized Return",           fmt_pct(metrics.annualized_return),"From log returns"),
        ("Sharpe Ratio",                f"{metrics.sharpe:.3f}",     f"Risk-free: {rfr:.2%}"),
        ("Sortino Ratio",               f"{metrics.sortino:.3f}",    "Downside deviation denominator"),
        ("Calmar Ratio",                f"{metrics.calmar:.3f}",     "Return / Max Drawdown"),
        ("Max Drawdown",                fmt_pct(metrics.max_drawdown),     f"{metrics.max_dd_duration} days duration"),
        ("Portfolio Beta",              f"{metrics.beta:.3f}",             "vs SPY"),
        ("Jensen's Alpha",              fmt_pct(metrics.alpha),    "CAPM residual"),
        ("Diversification Ratio",       f"{metrics.diversification_ratio:.3f}", "Wtd avg vol / port vol"),
        ("HHI",                         f"{metrics.hhi:.4f}",              "1/n ideal for equal weights"),
        ("Eff. Num. Bets",              f"{metrics.eff_num_bets:.1f}",     "1/HHI"),
        ("PCA Factors (90% var)",       str(metrics.n_pca_factors_90pct),    "Principal component analysis"),
    ]
    df_sum = pd.DataFrame(summary, columns=["Metric", "Value", "Notes"])
    st.dataframe(df_sum, hide_index=True, use_container_width=True, height=460)

# ════════════════════════════════════════════════════════════════════════════
# PAGE 3 — MONTE CARLO
# ════════════════════════════════════════════════════════════════════════════

elif page == "Monte Carlo":
    st.title("Monte Carlo Simulation")
    st.caption("10,000 Cholesky-correlated paths × 252 trading days (1 year forward)")

    pv = sim.initial_value

    # ── KPIs ──────────────────────────────────────────────────────────────
    mc_es_conf = getattr(settings, "es_confidence_level", 0.975)
    mc_shock   = getattr(settings, "mc_shock_distribution", "normal")
    cols = st.columns(5)
    median_ret  = (np.median(sim.terminal_values) - pv) / pv
    mean_ret    = (np.mean(sim.terminal_values) - pv) / pv
    var_95_pct  = sim.var_95 / pv
    cvar_es_pct = getattr(sim, "cvar_es", sim.cvar_95) / pv
    p_loss      = 1.0 - sim.prob_positive
    kpis = [
        ("Median 1Y Return",          fmt_pct(median_ret),    f"Median: {fmt_dollar(np.median(sim.terminal_values))}",  DCM_BLUE,  "up" if median_ret > 0 else "down"),
        ("Mean 1Y Return",            fmt_pct(mean_ret),      f"Mean: {fmt_dollar(np.mean(sim.terminal_values))}",      DCM_GREEN if mean_ret > 0 else DCM_RED, "up" if mean_ret > 0 else "down"),
        ("1Y VaR (95%)",              fmt_pct(var_95_pct),    "5th percentile of terminal returns",                     DCM_RED,   None),
        (f"1Y ES ({mc_es_conf:.1%})", fmt_pct(cvar_es_pct),   f"Expected Shortfall · {mc_shock} shocks",               DCM_RED,   None),
        ("P(Loss > 0%)",              fmt_pct(p_loss),        "Probability of any loss",                                DCM_GOLD,  None),
    ]
    for col, (lbl, val, sub, clr, trnd) in zip(cols, kpis):
        col.markdown(metric_card(lbl, val, sub, clr, trend=trnd), unsafe_allow_html=True)

    st.write("")

    # ── Fan chart ────────────────────────────────────────────────────────
    section_header("Simulation Fan Chart — 252-Day Forward Paths")

    days  = np.arange(sim.simulation_days + 1)
    pp    = sim.percentile_paths   # DataFrame columns: p05/p25/p50/p75/p95
    dates = pd.bdate_range(pd.Timestamp.today(), periods=sim.simulation_days + 1)[:len(days)]

    fig_fan = go.Figure()
    # 5th–95th band
    # prepend starting value (day 0 = initial_value) to each path
    def _fan_series(col_name):
        """Prepend day-0 value to a percentile column and return as list."""
        vals = [pv] + list(pp[col_name].values)
        return vals[:len(dates)]

    fig_fan.add_trace(go.Scatter(
        x=list(dates)+list(dates[::-1]),
        y=_fan_series("p05")+list(reversed(_fan_series("p95"))),
        fill="toself", fillcolor="rgba(41,98,255,0.05)",
        line=dict(color="rgba(0,0,0,0)"), name="P5–P95", showlegend=True,
        hoverinfo="skip",
    ))
    # 25th–75th band
    fig_fan.add_trace(go.Scatter(
        x=list(dates)+list(dates[::-1]),
        y=_fan_series("p25")+list(reversed(_fan_series("p75"))),
        fill="toself", fillcolor="rgba(41,98,255,0.12)",
        line=dict(color="rgba(0,0,0,0)"), name="P25–P75", showlegend=True,
        hoverinfo="skip",
    ))
    # Median
    fig_fan.add_trace(go.Scatter(
        x=dates, y=_fan_series("p50"),
        line=dict(color=DCM_BLUE, width=2.5), name="Median",
        hovertemplate="Day %{pointNumber}<br>Value: $%{y:,.0f}<extra></extra>",
    ))
    # P05 (VaR line)
    fig_fan.add_trace(go.Scatter(
        x=dates, y=_fan_series("p05"),
        line=dict(color=DCM_RED, width=1.5, dash="dot"), name="P5 (VaR)",
        hovertemplate="Day %{pointNumber}<br>VaR boundary: $%{y:,.0f}<extra></extra>",
    ))
    # Starting value
    fig_fan.add_hline(y=pv, line_dash="dash", line_color=DCM_GRAY, line_width=1,
                      annotation_text=f"Today: {fmt_dollar(pv)}")
    fig_fan.update_layout(
        xaxis_title="Date", yaxis_title="Portfolio Value ($)",
        height=380, margin=dict(t=20, b=50, l=70, r=20),

        legend=dict(orientation="h", y=-0.15),
        hovermode="x unified",
    )
    st.plotly_chart(fig_fan, use_container_width=True, config=PLOTLY_CONFIG)

    # ── Terminal value histogram ────────────────────────────────────────────
    section_header("Terminal Value Distribution (Day 252)")
    tv = sim.terminal_values
    fig_hist = go.Figure()
    fig_hist.add_trace(go.Histogram(
        x=tv, nbinsx=80,
        marker_color=DCM_BLUE, opacity=0.75,
        hovertemplate="Range: $%{x:,.0f}<br>Count: %{y}<extra></extra>",
        name="Simulations",
    ))
    p5_val = np.percentile(tv, 5)
    fig_hist.add_vline(x=pv, line_dash="dash", line_color=DCM_GRAY,
                       annotation_text=f"Today {fmt_dollar(pv)}", annotation_position="top right")
    fig_hist.add_vline(x=p5_val, line_dash="dot", line_color=DCM_RED,
                       annotation_text=f"VaR(95%) {fmt_dollar(p5_val)}", annotation_position="top right")
    fig_hist.add_vline(x=np.median(tv), line_dash="solid", line_color=DCM_GREEN, line_width=2,
                       annotation_text=f"Median {fmt_dollar(np.median(tv))}", annotation_position="top left")
    fig_hist.update_layout(
        xaxis_title="Terminal Portfolio Value ($)", yaxis_title="Number of Paths",
        height=280, margin=dict(t=30, b=50, l=60, r=20),

        showlegend=False,
    )
    st.plotly_chart(fig_hist, use_container_width=True, config=PLOTLY_CONFIG)

    # ── Probability stats ─────────────────────────────────────────────────
    section_header("Return Probability Analysis")
    pct_vals = [-0.30, -0.20, -0.10, 0.00, 0.10, 0.20, 0.30]
    rets = (tv - pv) / pv
    data_prob = []
    for threshold in pct_vals:
        prob = float(np.mean(rets < threshold))
        data_prob.append({"Threshold": f"{threshold:+.0%}", "P(Return < Threshold)": f"{prob:.2%}", "Count": f"{int(prob*len(rets)):,}"})
    st.dataframe(pd.DataFrame(data_prob), hide_index=True, use_container_width=False)

    # ── Multi-horizon ES table ─────────────────────────────────────────────
    section_header("Multi-Horizon Expected Shortfall")
    st.caption("ES from historical overlapping windows and MC (using sqrt-T proxy). Not suitable as a standalone estimate — always compare methods.")
    if mh_es is not None and not mh_es.empty:
        # Format as a readable table
        mh_display = mh_es.copy()
        mh_display.index = [f"{h}-Day" for h in mh_display.index]
        # Format dollar values
        def _fmt_es_val(v):
            try:
                return f"-${abs(float(v)):,.0f}" if float(v) < 0 else f"${float(v):,.0f}"
            except Exception:
                return str(v)
        mh_fmt = mh_display.map(_fmt_es_val)
        # Flatten MultiIndex columns if present
        if isinstance(mh_fmt.columns, pd.MultiIndex):
            mh_fmt.columns = [f"{src} {conf:.1%}" for src, conf in mh_fmt.columns]
        mh_fmt.index.name = "Horizon"
        st.dataframe(mh_fmt, use_container_width=True)
    else:
        st.info("Multi-horizon ES not available (requires successful data load).")

# ════════════════════════════════════════════════════════════════════════════
# PAGE 4 — STRESS TESTS
# ════════════════════════════════════════════════════════════════════════════

elif page == "Stress Tests":
    st.title("Stress Tests")
    st.caption("Historical scenarios & hypothetical shocks applied to current portfolio weights")

    pv = md.total_portfolio_value
    es_conf = getattr(settings, "es_confidence_level", 0.975)
    var99 = metrics.var_99

    def _severity_badge(sev: str, color: str) -> str:
        return (f'<span style="background:{color};color:#fff;padding:2px 8px;'
                f'border-radius:4px;font-size:11px;font-weight:700">{sev}</span>')

    def _build_scenario_rows(scenarios):
        rows = []
        for s in scenarios:
            dollar_loss = s.portfolio_loss_usd
            comp = s.es_comparison or {}
            row = {
                "Scenario":     s.name,
                "Severity":     s.severity,
                "Loss %":       f"{s.portfolio_loss_pct:.1%}",
                "Dollar Loss":  f"-${abs(dollar_loss):,.0f}",
            }
            if comp:
                row["vs ES(95%) ×"] = f"{comp.get('loss_to_es95_ratio', 0.0):.1f}×"
                row["vs ES(97.5%) ×"] = f"{comp.get('loss_to_es975_ratio', 0.0):.1f}×"
                row["vs ES(99%) ×"] = f"{comp.get('loss_to_es99_ratio', 0.0):.1f}×"
            rows.append(row)
        return pd.DataFrame(rows)

    # ── Historical scenarios ──────────────────────────────────────────────
    section_header("Historical Scenarios")
    st.dataframe(_build_scenario_rows(stress.historical), hide_index=True, use_container_width=True)

    # ── Hypothetical scenarios ─────────────────────────────────────────────
    section_header("Hypothetical Scenarios")
    st.dataframe(_build_scenario_rows(stress.hypothetical), hide_index=True, use_container_width=True)

    # ── Combined bar chart ────────────────────────────────────────────────
    section_header("All Scenarios — Portfolio Loss vs Expected Shortfall Benchmarks")
    all_scenarios = stress.historical + stress.hypothetical
    names   = [s.name for s in all_scenarios]
    losses  = [s.portfolio_loss_pct * 100 for s in all_scenarios]
    colors  = [s.severity_color for s in all_scenarios]

    fig_stress = go.Figure(go.Bar(
        x=names, y=losses,
        marker_color=colors,
        text=[f"{p:.1f}%  ({s.severity})" for p, s in zip(losses, all_scenarios)],
        textposition="outside",
        hovertemplate="<b>%{x}</b><br>Loss: %{y:.1f}%<extra></extra>",
    ))
    # Add ES benchmark lines for context
    if var_es_result := getattr(metrics, "var_es", None):
        es_pct = -(var_es_result.parametric_cvar / md.total_portfolio_value) * 100
        fig_stress.add_hline(y=es_pct, line_dash="dash", line_color=DCM_RED, line_width=1.5,
                             annotation_text=f"1-Day ES({es_conf:.0%}) = {es_pct:.2f}%",
                             annotation_position="top right")
    var99_es_pct = -(var99.parametric_cvar / md.total_portfolio_value) * 100
    fig_stress.add_hline(y=var99_es_pct, line_dash="dot", line_color="#8b0000", line_width=1,
                         annotation_text=f"1-Day ES(99%) = {var99_es_pct:.2f}%",
                         annotation_position="bottom right")
    fig_stress.update_layout(
        xaxis_title="", yaxis_title="Portfolio Loss (%)",
        height=400, margin=dict(t=30, b=100, l=60, r=20),

        showlegend=False,
        xaxis_tickangle=-35,
    )
    st.plotly_chart(fig_stress, use_container_width=True, config=PLOTLY_CONFIG)

    # ── Detailed stock impacts ────────────────────────────────────────────
    section_header("Detailed Stock-Level Impact — Select Scenario")
    scenario_names = [s.name for s in all_scenarios]
    selected = st.selectbox("Choose scenario", scenario_names)
    chosen = next(s for s in all_scenarios if s.name == selected)

    stock_impacts = sorted(chosen.stock_impacts, key=lambda si: si.dollar_loss)
    s_tickers = [si.ticker for si in stock_impacts]
    s_losses  = [si.scenario_drawdown * 100 for si in stock_impacts]
    s_colors  = [DCM_RED if v < -20 else DCM_GOLD if v < -10 else DCM_BLUE for v in s_losses]

    fig_detail = go.Figure(go.Bar(
        x=s_tickers, y=s_losses,
        marker_color=s_colors,
        hovertemplate="<b>%{x}</b><br>Impact: %{y:.1f}%<extra></extra>",
    ))
    fig_detail.update_layout(
        title=f"Stock-Level Impact: {selected}",
        xaxis_title="", yaxis_title="P&L Impact (%)",
        height=300, margin=dict(t=50, b=60, l=60, r=20),

        showlegend=False,
    )
    st.plotly_chart(fig_detail, use_container_width=True, config=PLOTLY_CONFIG)

    with st.expander("Scenario Methodology, Assumptions & ES Comparison"):
        st.markdown(f"**{chosen.name}** — Severity: **{chosen.severity}**")
        st.markdown(f"*Methodology:* {chosen.methodology}")
        st.markdown(f"*Assumptions:* {chosen.assumptions}")
        st.markdown(f"*Interpretation:* {chosen.interpretation}")
        if chosen.es_comparison:
            comp = chosen.es_comparison
            st.markdown("**ES Context:**")
            st.markdown(comp["multiples_of_daily_es"])
            cmp_rows = [
                ("Scenario Loss ($)",      f"${comp['scenario_loss_usd']:,.0f}"),
                ("1-Day ES(95%)",          f"${comp['es_95_1d']:,.0f}"),
                ("1-Day ES(97.5%) [primary]", f"${comp['es_975_1d']:,.0f}"),
                ("1-Day ES(99%)",          f"${comp['es_99_1d']:,.0f}"),
                ("21-Day ES(99%) proxy",   f"${comp['es_99_21d']:,.0f}"),
                ("Multiples of 1-day ES(95%)",  f"{comp['loss_to_es95_ratio']:.1f}×"),
                ("Multiples of 1-day ES(97.5%)", f"{comp['loss_to_es975_ratio']:.1f}×"),
                ("Multiples of 1-day ES(99%)",  f"{comp['loss_to_es99_ratio']:.1f}×"),
                ("Multiples of 21-day ES(99%)", f"{comp['loss_to_es99_21d_ratio']:.2f}×"),
            ]
            st.dataframe(pd.DataFrame(cmp_rows, columns=["Metric", "Value"]), hide_index=True)

# ════════════════════════════════════════════════════════════════════════════
# PAGE 5 — STOCK ANALYSIS
# ════════════════════════════════════════════════════════════════════════════

elif page == "Stock Analysis":
    st.title("Individual Stock Analysis")
    st.caption("Per-holding risk breakdown and contribution to portfolio risk")

    # ── Risk metrics table ────────────────────────────────────────────────
    section_header("Stock-Level Risk Metrics")
    # Build weight lookup from holdings (StockRiskMetrics doesn't carry weight)
    weight_map = {h.ticker: h.weight for h in md.holdings}
    rows = []
    for sm in sorted(metrics.stock_metrics, key=lambda s: -weight_map.get(s.ticker, 0.0)):
        rows.append({
            "Ticker":        sm.ticker,
            "Weight":        f"{weight_map.get(sm.ticker, 0.0):.2%}",
            "Ann. Return":   f"{sm.annualized_return:.1%}",
            "Ann. Vol":      f"{sm.annualized_vol:.1%}",
            "Beta":          f"{sm.beta:.2f}",
            "Sharpe":        f"{sm.sharpe:.2f}",
            "Max DD":        f"{sm.max_drawdown:.1%}",
            "Component VaR": f"${sm.component_var_95:,.0f}",
            "VaR Contrib %": f"{sm.component_var_95 / metrics.var_95.parametric_var * 100:.1f}%",
            "Marginal VaR":  f"${sm.marginal_var_95:,.2f}",
        })
    df_stocks = pd.DataFrame(rows)

    # Colour the weight column based on value
    st.dataframe(df_stocks, hide_index=True, use_container_width=True, height=500)

    # ── Component VaR contribution chart ─────────────────────────────────
    section_header("Component VaR Contribution (Euler Decomposition)")
    sm_sorted = sorted(metrics.stock_metrics, key=lambda s: -s.component_var_95)
    top_n = sm_sorted[:20]  # top 20 contributors

    fig_comp = go.Figure(go.Bar(
        x=[s.ticker for s in top_n],
        y=[s.component_var_95 for s in top_n],
        marker_color=[SECTOR_COLORS.get(
            next((h.sector for h in md.holdings if h.ticker == s.ticker), "Other"), "#78909c"
        ) for s in top_n],
        text=[f"${v:,.0f}" for v in [s.component_var_95 for s in top_n]],
        textposition="outside",
        hovertemplate="<b>%{x}</b><br>Component VaR: $%{y:,.0f}<extra></extra>",
    ))
    fig_comp.add_hline(
        y=metrics.var_95.parametric_var / len(metrics.stock_metrics),
        line_dash="dash", line_color=DCM_GRAY, line_width=1,
        annotation_text="Equal contribution",
    )
    fig_comp.update_layout(
        xaxis_title="", yaxis_title="Component VaR ($)",
        height=300, margin=dict(t=30, b=50, l=70, r=20),

        showlegend=False,
    )
    st.plotly_chart(fig_comp, use_container_width=True, config=PLOTLY_CONFIG)

    # ── Scatter: Vol vs Return ─────────────────────────────────────────────
    section_header("Risk-Return Scatter — All Holdings")
    sm_list  = metrics.stock_metrics
    h_lookup = {h.ticker: h for h in md.holdings}

    fig_scatter = go.Figure()
    for sm in sm_list:
        h = h_lookup.get(sm.ticker, None)
        color  = SECTOR_COLORS.get(h.sector if h else "Other", "#78909c")
        w = weight_map.get(sm.ticker, 0.01)
        fig_scatter.add_trace(go.Scatter(
            x=[sm.annualized_vol * 100],
            y=[sm.annualized_return * 100],
            mode="markers+text",
            marker=dict(size=max(8, w * 600), color=color, opacity=0.8,
                        line=dict(color=DCM_SURFACE, width=1)),
            text=[sm.ticker], textposition="top center", textfont_size=9,
            name=sm.ticker, showlegend=False,
            hovertemplate=(f"<b>{sm.ticker}</b><br>"
                           f"Vol: {sm.annualized_vol:.1%}<br>"
                           f"Return: {sm.annualized_return:.1%}<br>"
                           f"Weight: {w:.2%}<extra></extra>"),
        ))
    fig_scatter.add_hline(y=0, line_color=DCM_GRAY, line_width=1)
    fig_scatter.update_layout(
        xaxis_title="Annual Volatility (%)", yaxis_title="Annual Return (%)",
        height=420, margin=dict(t=20, b=50, l=60, r=20),

    )
    st.plotly_chart(fig_scatter, use_container_width=True, config=PLOTLY_CONFIG)

    # ── Correlation heatmap ───────────────────────────────────────────────
    section_header("Return Correlation Matrix (Top 20 by Weight)")
    top20_tickers = [sm.ticker for sm in sorted(metrics.stock_metrics, key=lambda s: -weight_map.get(s.ticker, 0.0))[:20]]
    prices_df = md.prices[[t for t in top20_tickers if t in md.prices.columns]]
    returns_df = prices_df.pct_change().dropna()
    corr_matrix = returns_df.corr()

    fig_hm = go.Figure(go.Heatmap(
        z=corr_matrix.values,
        x=corr_matrix.columns.tolist(),
        y=corr_matrix.index.tolist(),
        colorscale=[
            [0.0, "#c62828"], [0.5, "#ffffff"], [1.0, "#2962ff"]
        ],
        zmid=0, zmin=-1, zmax=1,
        text=[[f"{v:.2f}" for v in row] for row in corr_matrix.values],
        texttemplate="%{text}", textfont_size=9,
        hovertemplate="%{y} × %{x}<br>Correlation: %{z:.3f}<extra></extra>",
    ))
    fig_hm.update_layout(
        height=500, margin=dict(t=10, b=50, l=80, r=20),
    )
    st.plotly_chart(fig_hm, use_container_width=True, config=PLOTLY_CONFIG)

# ════════════════════════════════════════════════════════════════════════════
# PAGE 6 — BACKTESTING
# ════════════════════════════════════════════════════════════════════════════

elif page == "Backtesting":
    st.title("VaR / ES Backtesting")
    st.caption(
        "Rolling 252-day window backtest: Kupiec POF unconditional coverage test "
        "and Christoffersen independence test (exception clustering)."
    )

    if backtest is None:
        st.warning("Backtesting data not available — requires at least 3 years of price history.")
    else:
        bt = backtest
        _bt_conf = bt.confidence  # actual field name from BacktestResult

        # Extract time-series from forecast_df (columns: date, realized_pnl, var_forecast, exception)
        _fdf = bt.forecast_df.set_index("date") if "date" in bt.forecast_df.columns else bt.forecast_df
        _exc_s   = _fdf["exception"].astype(bool)
        _var_s   = _fdf["var_forecast"]     # positive = VaR magnitude
        _loss_s  = _fdf["realized_pnl"]     # negative = loss
        exc_dates = _exc_s[_exc_s].index

        # ES breach rate: fraction of days where loss < -ES (ES ~ var * 1.25 proxy)
        _es_proxy = _var_s * 1.25
        _es_breach_rate = float((_loss_s < -_es_proxy).mean()) if len(_loss_s) > 0 else 0.0

        # ── Summary KPIs ──────────────────────────────────────────────────
        cols = st.columns(5)
        kpis_bt = [
            ("VaR Exceptions",   str(bt.n_exceptions),                   f"out of {bt.n_obs} days",           DCM_RED),
            ("Exception Rate",   fmt_pct(bt.exception_rate),              f"Expected: {fmt_pct(1 - _bt_conf)}", DCM_GOLD),
            ("Kupiec p-value",   f"{bt.kupiec_pvalue:.3f}",               "H0: correct coverage · p≥0.05 = pass", DCM_GREEN if bt.kupiec_pvalue >= 0.05 else DCM_RED),
            ("Christoffersen p", f"{bt.christoffersen_pvalue:.3f}",       "H0: no clustering · p≥0.05 = pass", DCM_GREEN if bt.christoffersen_pvalue >= 0.05 else DCM_RED),
            ("ES Adequacy",      bt.es_adequacy,                          "Avg breach loss vs ES estimate",      DCM_GOLD),
        ]
        for col, (lbl, val, sub, clr) in zip(cols, kpis_bt):
            col.markdown(metric_card(lbl, val, sub, clr), unsafe_allow_html=True)

        st.write("")

        # ── Exception timeline ──────────────────────────────────────────
        section_header("VaR Exception Timeline")
        if len(_exc_s) > 0:
            fig_exc = go.Figure()
            fig_exc.add_trace(go.Scatter(
                x=_exc_s.index, y=[0] * len(_exc_s),
                mode="lines", line=dict(color=DCM_GRAY, width=0.5),
                showlegend=False, hoverinfo="skip",
            ))
            fig_exc.add_trace(go.Scatter(
                x=exc_dates, y=[1] * len(exc_dates),
                mode="markers",
                marker=dict(color=DCM_RED, size=8, symbol="x"),
                name=f"Exception ({len(exc_dates)} total)",
                hovertemplate="%{x|%Y-%m-%d}<br>VaR breach<extra></extra>",
            ))
            fig_exc.update_layout(
                xaxis_title="Date", yaxis=dict(tickvals=[0, 1], ticktext=["Normal", "Breach"]),
                height=200, margin=dict(t=20, b=40, l=80, r=20), showlegend=True,
            )
            st.plotly_chart(fig_exc, use_container_width=True, config=PLOTLY_CONFIG)
        else:
            st.info("Exception timeline not available.")

        # ── Rolling VaR vs actual losses ────────────────────────────────
        section_header("Rolling 1-Day VaR vs Realised Losses")
        if len(_fdf) > 0:
            fig_bt = go.Figure()
            fig_bt.add_trace(go.Scatter(
                x=_loss_s.index, y=_loss_s.values,
                line=dict(color=DCM_BLUE, width=1.5),
                name="Realised Dollar P&L",
                hovertemplate="%{x|%Y-%m-%d}<br>P&L: $%{y:,.0f}<extra></extra>",
            ))
            fig_bt.add_trace(go.Scatter(
                x=_var_s.index, y=-_var_s.values,
                line=dict(color=DCM_RED, width=1.5, dash="dot"),
                name=f"Rolling VaR ({_bt_conf:.0%}) forecast (−)",
                hovertemplate="%{x|%Y-%m-%d}<br>−VaR: $%{y:,.0f}<extra></extra>",
            ))
            for exc_date in exc_dates:
                fig_bt.add_vline(x=exc_date, line_color="rgba(198,40,40,0.25)", line_width=1)
            fig_bt.update_layout(
                xaxis_title="Date", yaxis_title="Dollar P&L / −VaR ($)",
                height=300, margin=dict(t=20, b=50, l=80, r=20),
                legend=dict(orientation="h", y=-0.2), hovermode="x unified",
            )
            st.plotly_chart(fig_bt, use_container_width=True, config=PLOTLY_CONFIG)
        else:
            st.info("Rolling VaR series not available.")

        # ── Test interpretation ──────────────────────────────────────────
        section_header("Statistical Test Results")
        kupiec_pass    = bt.kupiec_pvalue >= 0.05
        christoff_pass = bt.christoffersen_pvalue >= 0.05
        overall_pass   = kupiec_pass and christoff_pass

        result_color = DCM_GREEN if overall_pass else (DCM_GOLD if (kupiec_pass or christoff_pass) else DCM_RED)
        result_text  = (
            "PASS — VaR model performs within statistical expectations." if overall_pass else
            "PARTIAL — one test raises a concern; investigate further." if (kupiec_pass or christoff_pass) else
            "FAIL — VaR model systematically mis-estimates tail risk."
        )

        st.markdown(
            f'<div style="padding:12px;border-left:4px solid {result_color};background:rgba(0,0,0,0.03);margin-bottom:12px">'
            f'<b>Overall Assessment:</b> {result_text}</div>',
            unsafe_allow_html=True,
        )

        test_rows = [
            ("Kupiec POF",           f"χ²={bt.kupiec_statistic:.3f}",   f"p = {bt.kupiec_pvalue:.4f}",         "PASS ✓" if kupiec_pass else "FAIL ✗",    "Tests if observed exception rate matches theoretical 1-α"),
            ("Christoffersen Indep.", f"χ²={bt.christoffersen_stat:.3f}", f"p = {bt.christoffersen_pvalue:.4f}", "PASS ✓" if christoff_pass else "FAIL ✗", "Tests for autocorrelation in exception sequence"),
        ]
        st.dataframe(
            pd.DataFrame(test_rows, columns=["Test", "Statistic", "p-value", "Result", "Description"]),
            hide_index=True, use_container_width=True,
        )

        with st.expander("Interpretation Guide"):
            st.markdown("""
**Kupiec POF (Proportion of Failures):**
At 95% confidence, the model should produce exceptions on approximately 5% of days (1 in 20).
A p-value below 0.05 means the observed rate is statistically inconsistent with 5%.
Too few exceptions → VaR is too conservative (capital inefficient).
Too many → VaR understates risk.

**Christoffersen Independence Test:**
Even if the exception rate is correct on average, exceptions should not cluster together.
Clustering indicates the model fails to capture volatility persistence (GARCH effects) — risk rises sharply
during crises but the model keeps using the full-period average volatility.
p < 0.05 means statistically significant clustering.

**Practical guidance:**
- With < 500 observations, both tests have low power. Treat results as indicative, not conclusive.
- Backtesting VaR only measures one day at a time. Multi-day ES backtesting requires additional methods
  (e.g., Acerbi–Szekely test) not yet implemented here.
- If Christoffersen fails, consider switching to EWMA covariance (`covariance_mode = ewma` in Settings).
""")

# ════════════════════════════════════════════════════════════════════════════
# PAGE 7 — REPORTS & EXPORT
# ════════════════════════════════════════════════════════════════════════════

elif page == "Reports & Export":
    st.title("Reports & Export")
    st.caption("Generate professional PDF-ready HTML reports and write results back to Excel")

    col1, col2 = st.columns(2)

    with col1:
        section_header("HTML Risk Report")
        st.markdown("""
Generate a fully self-contained HTML report including all charts, metrics, stress tests,
and Monte Carlo results. No internet connection required to view the report.
        """)
        if st.button("Generate HTML Report", use_container_width=True, type="primary"):
            with st.spinner("Generating report (may take 10–20 seconds)…"):
                try:
                    report_path = generate_html_report(
                        md, metrics, sim, stress,
                        settings_title=settings.report_title,
                        portfolio_name=settings.portfolio_name,
                        portfolio_short_name=settings.portfolio_short_name,
                    )
                    st.success(f"Report saved: {report_path.name}")
                    with open(report_path, "rb") as f:
                        st.download_button(
                            "Download HTML Report",
                            data=f.read(),
                            file_name=report_path.name,
                            mime="text/html",
                            use_container_width=True,
                        )
                except Exception as e:
                    st.error(f"Report generation failed: {e}")

    with col2:
        section_header("Excel Write-Back")
        st.markdown("""
Write computed risk metrics, stock-level analysis, stress test results,
and Monte Carlo summary back to the portfolio Excel file as additional sheets.
        """)
        if st.button("Write Results to Excel", use_container_width=True, type="primary"):
            with st.spinner("Writing to Excel…"):
                try:
                    xlsx_path = get_portfolio_path()
                    write_results_to_excel(
                        xlsx_path, md, metrics, sim, stress,
                        portfolio_short_name=settings.portfolio_short_name,
                    )
                    st.success(f"Results written to {xlsx_path.name}")
                    with open(xlsx_path, "rb") as f:
                        st.download_button(
                            "Download Excel File",
                            data=f.read(),
                            file_name=xlsx_path.name,
                            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                            use_container_width=True,
                        )
                except Exception as e:
                    st.error(f"Excel write failed: {e}")

    # ── Existing reports ──────────────────────────────────────────────────
    section_header("Previously Generated Reports")
    existing_reports = sorted(EXPORTS_DIR.glob("*_Risk_Report_*.html"), reverse=True)
    audit_logs       = sorted(EXPORTS_DIR.glob("audit_*.txt"), reverse=True)

    if existing_reports:
        st.markdown("**HTML Reports:**")
        for p in existing_reports[:5]:
            size_kb = p.stat().st_size // 1024
            with open(p, "rb") as f:
                st.download_button(
                    f"{p.name}  ({size_kb:,} KB)",
                    data=f.read(),
                    file_name=p.name,
                    mime="text/html",
                    key=f"dl_{p.name}",
                )
    else:
        st.info("No HTML reports generated yet.")

    if audit_logs:
        st.markdown("**Audit Logs:**")
        for p in audit_logs[:5]:
            st.expander(f"{p.name}").text(p.read_text())

    # ── Data quality report ────────────────────────────────────────────────
    section_header("Data Quality Report")
    st.code(md.quality.to_text(), language=None)

    # ── Session info ──────────────────────────────────────────────────────
    section_header("Session Information")
    info_rows = [
        ("Portfolio File",       str(get_portfolio_path().name)),
        ("Holdings Loaded",      str(len(md.holdings))),
        ("Portfolio Value",      fmt_dollar(md.total_portfolio_value)),
        ("Data Source",          f"Live yfinance ({len(md.quality.valid_tickers)} tickers)"),
        ("Price Date Range",     f"{md.quality.date_range[0]} → {md.quality.date_range[1]}"),
        ("Risk-Free Rate",       fmt_pct(md.risk_free_rate)),
        ("Benchmark",            md.benchmark_ticker),
        ("Valid Tickers",        str(len(md.quality.valid_tickers))),
        ("Failed Tickers",       str(len(md.quality.failed_tickers)) + (f" ({', '.join(md.quality.failed_tickers)})" if md.quality.failed_tickers else "")),
        ("Report Output Dir",    str(EXPORTS_DIR)),
    ]
    df_info = pd.DataFrame(info_rows, columns=["Item", "Value"])
    st.dataframe(df_info, hide_index=True, use_container_width=True, height=340)
