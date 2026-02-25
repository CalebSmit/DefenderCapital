"""Phase 1 Audit — Data Infrastructure."""
import sys, time, shutil
from pathlib import Path
sys.path.insert(0, str(Path(__file__).resolve().parent.parent))

from engine.utils import AuditLog, EXPORTS_DIR
from engine.data_loader import load_portfolio
from engine.market_data import fetch_market_data, update_excel_prices

audit = AuditLog("Phase 1 — Data Infrastructure & Portfolio Input")

# ── Test 1: Basic load ──────────────────────────────────────────────────────
try:
    result = load_portfolio()
    n = len(result.holdings)
    if n == 35:
        audit.record("Holdings count", "PASS", f"35 holdings loaded correctly")
    else:
        audit.record("Holdings count", "FAIL", f"Expected 35, got {n}")
except Exception as e:
    audit.record("Holdings count", "FAIL", str(e))

# ── Test 2: Settings ────────────────────────────────────────────────────────
try:
    s = result.settings
    checks = [
        s.benchmark_ticker == "SPY",
        s.risk_free_rate == "auto",
        s.lookback_years == 2,
        s.simulation_paths == 10000,
        abs(s.confidence_level_1 - 0.95) < 0.001,
    ]
    if all(checks):
        audit.record("Settings parsing", "PASS", f"All settings parsed correctly")
    else:
        audit.record("Settings parsing", "FAIL", f"Some settings wrong: {s}")
except Exception as e:
    audit.record("Settings parsing", "FAIL", str(e))

# ── Test 3: Market data fetch (live) ────────────────────────────────────────
try:
    t0 = time.perf_counter()
    md = fetch_market_data(result)
    elapsed_first = time.perf_counter() - t0
    n_valid = len(md.quality.valid_tickers)
    n_failed = len(md.quality.failed_tickers)
    date_range = md.quality.date_range
    missing_pts = md.quality.missing_data_points

    detail = (
        f"Valid: {n_valid}/35 tickers\n"
        f"Failed: {n_failed} ({', '.join(md.quality.failed_tickers) or 'none'})\n"
        f"Date range: {date_range[0]} to {date_range[1]}\n"
        f"Missing data points filled: {missing_pts}\n"
        f"First fetch time: {elapsed_first:.1f}s"
    )
    if n_valid >= 30:
        audit.record("Market data fetch (live)", "PASS", detail)
    else:
        audit.record("Market data fetch (live)", "WARN", detail + "\n⚠️ Many tickers failed")
except Exception as e:
    audit.record("Market data fetch (live)", "FAIL", str(e))
    md = None

# ── Test 4: Second fetch (no caching — always fresh) ──────────────────────
try:
    t1 = time.perf_counter()
    md2 = fetch_market_data(result)
    elapsed_second = time.perf_counter() - t1
    detail = f"First run: {elapsed_first:.2f}s | Second run: {elapsed_second:.2f}s"
    audit.record("Second fetch test", "PASS", detail)
except Exception as e:
    audit.record("Second fetch test", "FAIL", str(e))

# ── Test 5: Auto-populated fields ──────────────────────────────────────────
try:
    if md:
        problems = []
        for h in md.holdings[:10]:
            if h.current_price <= 0:
                problems.append(f"{h.ticker}: price=0")
            if not h.company_name or h.company_name == h.ticker:
                problems.append(f"{h.ticker}: no company name")
            if h.market_value <= 0:
                problems.append(f"{h.ticker}: market_value=0")
            if not (0 < h.weight < 1):
                problems.append(f"{h.ticker}: weight={h.weight:.4f} out of range")

        portfolio_val = md.total_portfolio_value
        spot_checks = ""
        for h in md.holdings[:5]:
            spot_checks += f"  {h.ticker}: price=${h.current_price:.2f}, val=${h.market_value:,.0f}, wt={h.weight:.2%}\n"

        if not problems:
            audit.record("Auto-populated fields", "PASS",
                         f"Portfolio value: ${portfolio_val:,.2f}\nSpot checks:\n{spot_checks}")
        else:
            audit.record("Auto-populated fields", "FAIL",
                         f"Issues: {problems}\n{spot_checks}")
except Exception as e:
    audit.record("Auto-populated fields", "FAIL", str(e))

# ── Test 6: Validation stress test ──────────────────────────────────────────
import openpyxl, copy
from engine.data_loader import load_portfolio
from engine.utils import get_portfolio_path

test_path = EXPORTS_DIR / "test_invalid_holdings.xlsx"
try:
    shutil.copy(get_portfolio_path(), test_path)
    wb = openpyxl.load_workbook(str(test_path))
    ws = wb["Holdings"]
    # Introduce errors
    ws["A4"] = "APPL"       # misspelled ticker
    ws["E5"] = -100          # negative shares
    ws["F6"] = "not a number"  # bad cost basis
    ws["A10"] = ""           # empty ticker row
    ws["E11"] = "abc"        # string shares
    wb.save(str(test_path))

    bad_result = load_portfolio(test_path, raise_on_empty=False)
    n_err    = len(bad_result.errors)
    n_skip   = len(bad_result.skipped_rows)
    n_good   = len(bad_result.holdings)
    summary  = (
        f"Errors caught: {n_err}, Rows skipped: {n_skip}, Valid rows: {n_good}\n"
        f"Errors:\n" + "\n".join(f"  {e}" for e in bad_result.errors[:10])
    )
    if n_err >= 2 and n_good >= 28:
        audit.record("Validation stress test", "PASS", summary)
    else:
        audit.record("Validation stress test", "WARN", summary + "\n  Fewer errors caught than expected")
    test_path.unlink(missing_ok=True)
except Exception as e:
    audit.record("Validation stress test", "FAIL", str(e))

# ── Test 7: Risk-free rate ──────────────────────────────────────────────────
try:
    if md:
        rfr = md.risk_free_rate
        if 0.01 < rfr < 0.15:
            audit.record("Risk-free rate", "PASS", f"Rate = {rfr:.4f} ({rfr*100:.2f}%) — plausible range")
        else:
            audit.record("Risk-free rate", "WARN", f"Rate = {rfr:.4f} — outside expected 1-15% range")
except Exception as e:
    audit.record("Risk-free rate", "FAIL", str(e))

# ── Test 8: Excel price update ──────────────────────────────────────────────
try:
    if md:
        update_excel_prices(md, get_portfolio_path())
        # Verify by re-reading
        check = load_portfolio()
        prices_written = sum(1 for h in check.holdings if h.current_price > 0)
        if prices_written >= 30:
            audit.record("Excel price write-back", "PASS",
                         f"Current prices written for {prices_written}/35 holdings")
        else:
            audit.record("Excel price write-back", "WARN",
                         f"Only {prices_written} prices written — some tickers may have failed")
except Exception as e:
    audit.record("Excel price write-back", "FAIL", str(e))

# ── Test 9: Data quality report ─────────────────────────────────────────────
try:
    if md:
        qr = md.quality.to_text()
        detail = qr[:500]
        audit.record("Data quality report", "PASS", detail)
except Exception as e:
    audit.record("Data quality report", "FAIL", str(e))

# ── Save audit log ──────────────────────────────────────────────────────────
log_path = audit.save("audit_phase1.txt")
print(f"\n{'='*60}")
print(f"Phase 1 Audit Complete")
print(f"{'='*60}")
for e in audit.entries:
    sym = {"PASS":"✅","FAIL":"❌","WARN":"⚠️","INFO":"ℹ️"}.get(e["result"],"?")
    print(f"{sym} {e['result']:4s}  {e['test']}")
print(f"\nAll passed: {audit.all_passed}")
print(f"Audit log: {log_path}")
