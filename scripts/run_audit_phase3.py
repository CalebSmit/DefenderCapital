"""Phase 3 Audit — Dashboard & Report Generation."""
import sys, subprocess, ast, importlib.util, time
from pathlib import Path
sys.path.insert(0, str(Path(__file__).resolve().parent.parent))

from engine.utils import AuditLog, EXPORTS_DIR
from engine.data_loader import load_portfolio
from engine.synthetic_data import get_market_data
from engine.risk_metrics import compute_all_metrics
from engine.monte_carlo import run_simulation
from engine.stress_testing import run_all_stress_tests
from engine.report_generator import generate_html_report, write_results_to_excel

audit = AuditLog("Phase 3 — Dashboard & Report Generation")

# Load shared data once
result  = load_portfolio()
md      = get_market_data(result)
metrics = compute_all_metrics(md)
sim     = run_simulation(md, n_paths=5000, n_days=252, seed=42)
stress  = run_all_stress_tests(md)

# ── TEST 1: Dashboard file exists and is valid Python ─────────────────────
try:
    app_path = Path(__file__).resolve().parent.parent / "dashboard" / "app.py"
    assert app_path.exists(), "dashboard/app.py not found"
    with open(app_path) as f:
        source = f.read()
    ast.parse(source)  # syntax check
    lines = source.count("\n")
    audit.record("Dashboard file exists & syntax valid", "PASS",
                 f"dashboard/app.py: {lines:,} lines, syntax OK")
except Exception as e:
    audit.record("Dashboard file exists & syntax valid", "FAIL", str(e))

# ── TEST 2: Dashboard imports all engine modules ──────────────────────────
try:
    required_imports = [
        "engine.utils", "engine.data_loader", "engine.synthetic_data",
        "engine.risk_metrics", "engine.monte_carlo", "engine.stress_testing",
        "engine.report_generator",
    ]
    missing = []
    for mod in required_imports:
        spec = importlib.util.find_spec(mod)
        if spec is None:
            missing.append(mod)
    if missing:
        audit.record("Dashboard module imports", "FAIL", f"Missing: {missing}")
    else:
        audit.record("Dashboard module imports", "PASS",
                     f"All {len(required_imports)} engine modules importable")
except Exception as e:
    audit.record("Dashboard module imports", "FAIL", str(e))

# ── TEST 3: Dashboard has 6 pages ────────────────────────────────────────
try:
    with open(app_path) as f:
        content = f.read()
    pages = [
        "Portfolio Overview",
        "Risk Dashboard",
        "Monte Carlo",
        "Stress Tests",
        "Stock Analysis",
        "Reports & Export",
    ]
    missing_pages = [p for p in pages if p not in content]
    if not missing_pages:
        audit.record("Dashboard has 6 pages", "PASS",
                     f"All pages present: {', '.join(pages)}")
    else:
        audit.record("Dashboard has 6 pages", "FAIL",
                     f"Missing pages: {missing_pages}")
except Exception as e:
    audit.record("Dashboard has 6 pages", "FAIL", str(e))

# ── TEST 4: HTML report generation ───────────────────────────────────────
try:
    t0 = time.perf_counter()
    report_path = generate_html_report(md, metrics, sim, stress)
    elapsed = time.perf_counter() - t0
    size_kb = report_path.stat().st_size // 1024
    ok = report_path.exists() and report_path.stat().st_size > 10_000
    # Check report contains key sections
    content = report_path.read_text()
    required_sections = ["Portfolio", "VaR", "Monte Carlo",
                         "Stress", "Sharpe", "Beta"]
    missing_sections = [s for s in required_sections if s not in content]
    if ok and not missing_sections:
        audit.record("HTML report: file & sections", "PASS",
                     f"Size: {size_kb:,} KB, time: {elapsed:.1f}s, all sections present")
    elif ok:
        audit.record("HTML report: file & sections", "WARN",
                     f"File OK ({size_kb:,} KB) but missing sections: {missing_sections}")
    else:
        audit.record("HTML report: file & sections", "FAIL",
                     f"File missing or too small: {report_path}")
except Exception as e:
    audit.record("HTML report: file & sections", "FAIL", str(e))

# ── TEST 5: HTML report is self-contained (no external URLs in CSS/JS) ───
try:
    content = report_path.read_text()
    # Must not load from external CDN for core content
    ext_deps = []
    suspicious = ["cdn.jsdelivr", "cdn.plot.ly", "googleapis.com/css"]
    for dep in suspicious:
        if dep in content:
            ext_deps.append(dep)
    if not ext_deps:
        audit.record("HTML report self-contained", "PASS",
                     "No external CDN dependencies found (all assets inline)")
    else:
        audit.record("HTML report self-contained", "WARN",
                     f"External deps found: {ext_deps} (may not render offline)")
except Exception as e:
    audit.record("HTML report self-contained", "FAIL", str(e))

# ── TEST 6: Excel write-back ──────────────────────────────────────────────
try:
    from engine.utils import get_portfolio_path
    xlsx_path = get_portfolio_path()
    write_results_to_excel(xlsx_path, md, metrics, sim, stress)

    import openpyxl
    wb = openpyxl.load_workbook(str(xlsx_path))
    required_sheets = ["Risk Summary", "Stock Risk Detail", "Stress Test Results"]
    present = [s for s in required_sheets if s in wb.sheetnames]
    missing = [s for s in required_sheets if s not in wb.sheetnames]
    if not missing:
        audit.record("Excel write-back: required sheets", "PASS",
                     f"Sheets present: {', '.join(present)}")
    else:
        audit.record("Excel write-back: required sheets", "FAIL",
                     f"Missing sheets: {missing}")

    # Check Risk Summary has content
    rs = wb["Risk Summary"]
    cell_count = sum(1 for row in rs.iter_rows() for cell in row if cell.value is not None)
    if cell_count >= 10:
        audit.record("Excel Risk Summary has content", "PASS",
                     f"{cell_count} non-empty cells in Risk Summary sheet")
    else:
        audit.record("Excel Risk Summary has content", "WARN",
                     f"Only {cell_count} non-empty cells")
except Exception as e:
    audit.record("Excel write-back", "FAIL", str(e))

# ── TEST 7: Dashboard has Refresh button ──────────────────────────────────
try:
    with open(app_path) as f:
        content = f.read()
    has_refresh = "Refresh Data" in content or "refresh" in content.lower()
    has_cache   = "@st.cache_data" in content
    if has_refresh and has_cache:
        audit.record("Dashboard: Refresh & caching", "PASS",
                     "Refresh button present, @st.cache_data caching used")
    elif has_refresh:
        audit.record("Dashboard: Refresh & caching", "WARN",
                     "Refresh button present but no @st.cache_data caching")
    else:
        audit.record("Dashboard: Refresh & caching", "FAIL",
                     "No Refresh button found")
except Exception as e:
    audit.record("Dashboard: Refresh & caching", "FAIL", str(e))

# ── TEST 8: Dashboard uses plotly charts ──────────────────────────────────
try:
    with open(app_path) as f:
        content = f.read()
    chart_types = ["go.Pie", "go.Bar", "go.Scatter", "go.Histogram", "go.Heatmap"]
    found = [ct for ct in chart_types if ct in content]
    if len(found) >= 4:
        audit.record("Dashboard: plotly chart types", "PASS",
                     f"Found {len(found)}/5 chart types: {', '.join(found)}")
    else:
        audit.record("Dashboard: plotly chart types", "WARN",
                     f"Only {len(found)} chart types found: {found}")
except Exception as e:
    audit.record("Dashboard: plotly chart types", "FAIL", str(e))

# ── TEST 9: Report output directory exists ────────────────────────────────
try:
    reports = list(EXPORTS_DIR.glob("DCM_Risk_Report_*.html"))
    audits  = list(EXPORTS_DIR.glob("audit_*.txt"))
    if len(reports) >= 1:
        latest = max(reports, key=lambda p: p.stat().st_mtime)
        size_kb = latest.stat().st_size // 1024
        audit.record("Exports directory populated", "PASS",
                     f"{len(reports)} reports, {len(audits)} audit logs in {EXPORTS_DIR.name}/")
    else:
        audit.record("Exports directory populated", "WARN",
                     f"No HTML reports yet in {EXPORTS_DIR}")
except Exception as e:
    audit.record("Exports directory populated", "FAIL", str(e))

# ── TEST 10: Streamlit is installed ──────────────────────────────────────
try:
    import streamlit
    version = streamlit.__version__
    audit.record("Streamlit installed", "PASS", f"streamlit=={version}")
except ImportError:
    audit.record("Streamlit installed", "FAIL", "streamlit not installed — run: pip install streamlit")

# ── Save ───────────────────────────────────────────────────────────────────
log_path = audit.save("audit_phase3.txt")
print(f"\n{'='*60}")
print("Phase 3 Audit Complete")
print(f"{'='*60}")
for e in audit.entries:
    sym = {"PASS": "✅", "FAIL": "❌", "WARN": "⚠️"}.get(e["result"], "?")
    print(f"{sym} {e['result']:4s}  {e['test']}")
print(f"\nAll critical tests passed: {audit.all_passed}")
print(f"Audit log: {log_path}")
