"""
auth/portfolio_store.py
─────────────────────────────────────────────────────────────────────────────
Save and load portfolio data per user.
Holdings are stored as JSON in dcm_portfolios.holdings.
Settings are stored as JSON in dcm_portfolios.settings.
Each user has exactly one portfolio row (UPSERT pattern).
"""

from __future__ import annotations

import json
from typing import Optional

import openpyxl

from auth.database import (
    execute,
    execute_returning,
    fetchone,
    json_decode,
    json_encode,
    ph,
    _is_postgres,
)


# ── save ───────────────────────────────────────────────────────────────────

def save_portfolio(
    user_id: int,
    portfolio_name: str,
    short_name: str,
    holdings: list[dict],
    settings: dict | None = None,
) -> tuple[bool, str]:
    """
    Upsert the user's portfolio.
    Returns (success, message).
    """
    if not isinstance(holdings, list):
        return False, "Holdings must be a list."
    if not portfolio_name.strip():
        return False, "Portfolio name cannot be empty."

    h_json = json_encode(holdings)
    s_json = json_encode(settings or {})
    name   = portfolio_name.strip()[:100]
    short  = (short_name or "PORT").strip()[:20]

    try:
        if _is_postgres():
            execute(
                """
                INSERT INTO dcm_portfolios
                    (user_id, portfolio_name, short_name, holdings, settings, updated_at)
                VALUES (%s, %s, %s, %s::jsonb, %s::jsonb, NOW())
                ON CONFLICT (user_id) DO UPDATE SET
                    portfolio_name = EXCLUDED.portfolio_name,
                    short_name     = EXCLUDED.short_name,
                    holdings       = EXCLUDED.holdings,
                    settings       = EXCLUDED.settings,
                    updated_at     = NOW()
                """,
                (user_id, name, short, h_json, s_json),
            )
        else:
            # SQLite: manual upsert
            existing = fetchone(
                f"SELECT id FROM dcm_portfolios WHERE user_id = {ph()}", (user_id,)
            )
            if existing:
                execute(
                    f"UPDATE dcm_portfolios "
                    f"SET portfolio_name = {ph()}, short_name = {ph()}, "
                    f"    holdings = {ph()}, settings = {ph()}, "
                    f"    updated_at = datetime('now') "
                    f"WHERE user_id = {ph()}",
                    (name, short, h_json, s_json, user_id),
                )
            else:
                execute(
                    f"INSERT INTO dcm_portfolios "
                    f"(user_id, portfolio_name, short_name, holdings, settings) "
                    f"VALUES ({ph(5)})",
                    (user_id, name, short, h_json, s_json),
                )
        return True, f"Portfolio '{name}' saved."
    except Exception as exc:
        return False, f"Save failed: {exc}"


# ── load ───────────────────────────────────────────────────────────────────

def load_portfolio(user_id: int) -> Optional[dict]:
    """
    Return the user's saved portfolio or None if they have never saved one.

    Return shape:
    {
        "portfolio_name": str,
        "short_name":     str,
        "holdings":       list[dict],   # [{ticker, shares_held, cost_basis, ...}]
        "settings":       dict,
        "updated_at":     str,
    }
    """
    row = fetchone(
        f"SELECT * FROM dcm_portfolios WHERE user_id = {ph()}", (user_id,)
    )
    if not row:
        return None

    return {
        "portfolio_name": row["portfolio_name"],
        "short_name":     row["short_name"] or "PORT",
        "holdings":       json_decode(row["holdings"]) or [],
        "settings":       json_decode(row["settings"]) or {},
        "updated_at":     str(row.get("updated_at", "")),
    }


def has_portfolio(user_id: int) -> bool:
    row = fetchone(
        f"SELECT id FROM dcm_portfolios WHERE user_id = {ph()}", (user_id,)
    )
    return row is not None


# ── Excel ↔ portfolio conversion ──────────────────────────────────────────

def holdings_to_excel_bytes(
    portfolio_name: str,
    short_name: str,
    holdings: list[dict],
    settings: dict | None = None,
) -> bytes:
    """
    Serialise a saved portfolio to Excel bytes (same format the dashboard expects).
    This lets users download their saved portfolio as an Excel file.
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Holdings"

    headers = ["Ticker", "Shares Held", "Cost Basis"]
    ws.append(headers)

    for h in holdings:
        ws.append([
            h.get("ticker", ""),
            h.get("shares_held", 0),
            h.get("cost_basis", 0),
        ])

    # Settings sheet
    ws2 = wb.create_sheet("Settings")
    ws2.append(["Parameter", "Value"])
    ws2.append(["portfolio_name", portfolio_name])
    ws2.append(["short_name",     short_name])

    if settings:
        for k, v in settings.items():
            ws2.append([k, v])

    import io
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()


def excel_path_to_holdings(excel_path: str) -> tuple[list[dict], str, str, dict]:
    """
    Read an Excel file (Holdings sheet) and return
    (holdings_list, portfolio_name, short_name, settings_dict).
    """
    wb = openpyxl.load_workbook(excel_path, data_only=True)
    holdings = []

    if "Holdings" in wb.sheetnames:
        ws = wb["Holdings"]
        rows = list(ws.iter_rows(values_only=True))
        if len(rows) > 1:
            for row in rows[1:]:
                if not row or not row[0]:
                    continue
                ticker = str(row[0]).strip().upper()
                try:
                    shares = float(row[1]) if row[1] is not None else 0.0
                except (ValueError, TypeError):
                    shares = 0.0
                try:
                    cost   = float(row[2]) if row[2] is not None else 0.0
                except (ValueError, TypeError):
                    cost   = 0.0
                if ticker:
                    holdings.append({
                        "ticker":     ticker,
                        "shares_held": shares,
                        "cost_basis":  cost,
                    })

    portfolio_name = "My Portfolio"
    short_name     = "PORT"
    settings: dict = {}

    if "Settings" in wb.sheetnames:
        ws2 = wb["Settings"]
        for row in ws2.iter_rows(values_only=True):
            if not row or row[0] is None:
                continue
            key = str(row[0]).strip()
            val = row[1]
            if key == "portfolio_name" and val:
                portfolio_name = str(val).strip()
            elif key == "short_name" and val:
                short_name = str(val).strip()
            elif val is not None:
                settings[key] = val

    return holdings, portfolio_name, short_name, settings
