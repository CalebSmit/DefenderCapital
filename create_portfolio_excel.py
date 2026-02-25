"""
Script to create the portfolio_holdings.xlsx master data file.
35 realistic holdings for a ~$100,000 student portfolio.
"""
import openpyxl
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, numbers
)
from openpyxl.styles.numbers import FORMAT_PERCENTAGE_00
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

wb = openpyxl.Workbook()

# ── Color palette ──────────────────────────────────────────────────────────────
NAVY   = "1B2A4A"
SILVER = "C0C0C0"
BLUE   = "4A90D9"
WHITE  = "FFFFFF"
LIGHT  = "F0F4FA"
YELLOW = "FFF9C4"
GREEN  = "E8F5E9"
RED_BG = "FFEBEE"

def header_font(bold=True, color=WHITE):
    return Font(name="Arial", bold=bold, size=10, color=color)

def body_font(bold=False, color="000000"):
    return Font(name="Arial", bold=bold, size=10, color=color)

def header_fill(color=NAVY):
    return PatternFill("solid", start_color=color, fgColor=color)

def center():
    return Alignment(horizontal="center", vertical="center", wrap_text=False)

def left():
    return Alignment(horizontal="left", vertical="center")

def right():
    return Alignment(horizontal="right", vertical="center")

thin = Side(style="thin", color="BBBBBB")
thick = Side(style="medium", color=NAVY)

def thin_border():
    return Border(left=thin, right=thin, top=thin, bottom=thin)

# ═══════════════════════════════════════════════════════════════════════════════
# SHEET 1: Holdings
# ═══════════════════════════════════════════════════════════════════════════════
ws_h = wb.active
ws_h.title = "Holdings"

# 35 realistic holdings across sectors for a ~$100K portfolio
# Format: (Ticker, Company Name, Sector, Industry, Shares, Cost Basis)
holdings = [
    # Technology (heavy, ~22% weight)
    ("AAPL",  "Apple Inc.",                   "Technology",           "Consumer Electronics",          28,  158.50),
    ("MSFT",  "Microsoft Corporation",         "Technology",           "Software—Infrastructure",       16,  285.00),
    ("NVDA",  "NVIDIA Corporation",            "Technology",           "Semiconductors",                12,  210.00),
    ("GOOGL", "Alphabet Inc.",                 "Communication Services",   "Internet Content & Info",       14,  108.00),
    ("CRM",   "Salesforce Inc.",               "Technology",           "Software—Application",          10,  195.00),
    # Financials (~15%)
    ("JPM",   "JPMorgan Chase & Co.",          "Financial Services",   "Banks—Diversified",             22,  135.00),
    ("BAC",   "Bank of America Corp.",         "Financial Services",   "Banks—Diversified",             80,   28.00),
    ("GS",    "Goldman Sachs Group Inc.",      "Financial Services",   "Capital Markets",                8,  310.00),
    ("V",     "Visa Inc.",                     "Financial Services",   "Credit Services",               12,  205.00),
    # Healthcare (~14%)
    ("JNJ",   "Johnson & Johnson",             "Healthcare",           "Drug Manufacturers—General",    14,  158.00),
    ("UNH",   "UnitedHealth Group Inc.",       "Healthcare",           "Healthcare Plans",               5,  465.00),
    ("PFE",   "Pfizer Inc.",                   "Healthcare",           "Drug Manufacturers—General",    60,   36.00),
    ("ABBV",  "AbbVie Inc.",                   "Healthcare",           "Drug Manufacturers—General",    15,  140.00),
    # Consumer Staples (~9%)
    ("PG",    "Procter & Gamble Co.",          "Consumer Defensive",     "Household & Personal Products", 18,  135.00),
    ("KO",    "Coca-Cola Co.",                 "Consumer Defensive",     "Beverages—Non-Alcoholic",       35,   58.00),
    ("WMT",   "Walmart Inc.",                  "Consumer Defensive",     "Discount Stores",               14,  148.00),
    # Consumer Discretionary (~8%)
    ("AMZN",  "Amazon.com Inc.",               "Consumer Cyclical","Internet Retail",              12,  112.00),
    ("HD",    "Home Depot Inc.",               "Consumer Cyclical","Home Improvement Retail",       8,  295.00),
    ("NKE",   "Nike Inc.",                     "Consumer Cyclical","Footwear & Accessories",       18,   98.00),
    # Industrials (~8%)
    ("HON",   "Honeywell International Inc.",  "Industrials",          "Conglomerates",                 10,  185.00),
    ("CAT",   "Caterpillar Inc.",              "Industrials",          "Farm & Heavy Construction",      8,  218.00),
    ("UPS",   "United Parcel Service Inc.",    "Industrials",          "Integrated Freight & Logistics", 12, 155.00),
    # Energy (~6%)
    ("XOM",   "Exxon Mobil Corporation",       "Energy",               "Oil & Gas—Integrated",          28,   88.00),
    ("CVX",   "Chevron Corporation",           "Energy",               "Oil & Gas—Integrated",          14,  152.00),
    # Communication Services (~5%)
    ("META",  "Meta Platforms Inc.",           "Communication Services",   "Internet Content & Info",        8,  250.00),
    ("DIS",   "Walt Disney Co.",               "Communication Services",   "Entertainment",                 20,   85.00),
    # Materials (~4%)
    ("LIN",   "Linde PLC",                     "Basic Materials",            "Specialty Chemicals",            6,  345.00),
    ("APD",   "Air Products & Chemicals Inc.", "Basic Materials",            "Specialty Chemicals",            8,  255.00),
    # Real Estate (~3%)
    ("AMT",   "American Tower Corp.",          "Real Estate",          "REIT—Specialty",                 8,  195.00),
    ("PLD",   "Prologis Inc.",                 "Real Estate",          "REIT—Industrial",               12,  115.00),
    # Utilities (~3%)
    ("NEE",   "NextEra Energy Inc.",           "Utilities",            "Utilities—Renewable",           20,   68.00),
    ("DUK",   "Duke Energy Corporation",       "Utilities",            "Utilities—Regulated Electric",  15,   88.00),
    # Smaller / high-conviction positions (~3%)
    ("TSLA",  "Tesla Inc.",                    "Consumer Cyclical","Auto Manufacturers",             8,  195.00),
    ("COST",  "Costco Wholesale Corp.",        "Consumer Defensive",     "Discount Stores",                4,  495.00),
    ("LMT",   "Lockheed Martin Corp.",         "Industrials",          "Aerospace & Defense",            5,  420.00),
]

# Column headers
headers = [
    "Ticker", "Company Name", "Sector", "Industry",
    "Shares Held", "Avg Cost Basis ($)", "Current Price ($)",
    "Market Value ($)", "Weight (%)", "Unrealized P&L ($)", "Unrealized P&L (%)"
]

col_widths = [10, 32, 24, 28, 13, 18, 18, 18, 12, 18, 18]

# Title row
ws_h.merge_cells("A1:K1")
ws_h["A1"] = "DEFENDER CAPITAL MANAGEMENT — PORTFOLIO HOLDINGS"
ws_h["A1"].font = Font(name="Arial", bold=True, size=13, color=WHITE)
ws_h["A1"].fill = header_fill(NAVY)
ws_h["A1"].alignment = center()
ws_h.row_dimensions[1].height = 28

# Header row (row 2)
for col_idx, (hdr, width) in enumerate(zip(headers, col_widths), start=1):
    cell = ws_h.cell(row=2, column=col_idx, value=hdr)
    cell.font = Font(name="Arial", bold=True, size=10, color=WHITE)
    cell.fill = header_fill(NAVY)
    cell.alignment = center()
    cell.border = thin_border()
    ws_h.column_dimensions[get_column_letter(col_idx)].width = width

ws_h.row_dimensions[2].height = 22

# Data rows (rows 3–37)
for row_num, (ticker, name, sector, industry, shares, cost) in enumerate(holdings, start=3):
    row_fill = PatternFill("solid", start_color=LIGHT, fgColor=LIGHT) if row_num % 2 == 0 else PatternFill("solid", start_color=WHITE, fgColor=WHITE)

    # A: Ticker
    c = ws_h.cell(row=row_num, column=1, value=ticker)
    c.font = Font(name="Arial", bold=True, size=10, color=NAVY)
    c.fill = row_fill; c.alignment = center(); c.border = thin_border()

    # B: Company Name — auto-populated by engine, but pre-filled
    c = ws_h.cell(row=row_num, column=2, value=name)
    c.font = body_font(); c.fill = row_fill; c.alignment = left(); c.border = thin_border()

    # C: Sector
    c = ws_h.cell(row=row_num, column=3, value=sector)
    c.font = body_font(); c.fill = row_fill; c.alignment = left(); c.border = thin_border()

    # D: Industry
    c = ws_h.cell(row=row_num, column=4, value=industry)
    c.font = body_font(); c.fill = row_fill; c.alignment = left(); c.border = thin_border()

    # E: Shares Held (blue = user input)
    c = ws_h.cell(row=row_num, column=5, value=shares)
    c.font = Font(name="Arial", size=10, color="0000FF")
    c.fill = row_fill; c.alignment = center(); c.border = thin_border()
    c.number_format = "#,##0"

    # F: Avg Cost Basis (blue = user input)
    c = ws_h.cell(row=row_num, column=6, value=cost)
    c.font = Font(name="Arial", size=10, color="0000FF")
    c.fill = row_fill; c.alignment = right(); c.border = thin_border()
    c.number_format = "$#,##0.00"

    # G: Current Price — populated by engine, leave as 0 placeholder
    c = ws_h.cell(row=row_num, column=7, value=0.0)
    c.font = Font(name="Arial", size=10, color="008000")
    c.fill = row_fill; c.alignment = right(); c.border = thin_border()
    c.number_format = "$#,##0.00"

    # H: Market Value = Shares × Current Price
    c = ws_h.cell(row=row_num, column=8,
                  value=f"=E{row_num}*G{row_num}")
    c.font = body_font(); c.fill = row_fill; c.alignment = right(); c.border = thin_border()
    c.number_format = "$#,##0.00"

    # I: Weight % = Market Value / SUM(Market Values)
    c = ws_h.cell(row=row_num, column=9,
                  value=f"=IF(SUM(H3:H37)=0,0,H{row_num}/SUM(H3:H37))")
    c.font = body_font(); c.fill = row_fill; c.alignment = center(); c.border = thin_border()
    c.number_format = "0.00%"

    # J: Unrealized P&L ($) = (Current Price − Cost Basis) × Shares
    c = ws_h.cell(row=row_num, column=10,
                  value=f"=(G{row_num}-F{row_num})*E{row_num}")
    c.font = body_font(); c.fill = row_fill; c.alignment = right(); c.border = thin_border()
    c.number_format = "$#,##0.00"

    # K: Unrealized P&L (%) = (Current Price − Cost Basis) / Cost Basis
    c = ws_h.cell(row=row_num, column=11,
                  value=f"=IF(F{row_num}=0,0,(G{row_num}-F{row_num})/F{row_num})")
    c.font = body_font(); c.fill = row_fill; c.alignment = center(); c.border = thin_border()
    c.number_format = "0.00%"

# Totals row (row 38)
ws_h.row_dimensions[38].height = 22
labels = ["", "TOTAL PORTFOLIO", "", "", "=SUM(E3:E37)", "", "", "=SUM(H3:H37)", "=SUM(I3:I37)", "=SUM(J3:J37)", ""]
for col_idx, val in enumerate(labels, start=1):
    c = ws_h.cell(row=38, column=col_idx, value=val if val else None)
    c.font = Font(name="Arial", bold=True, size=10, color=WHITE)
    c.fill = header_fill(NAVY)
    c.alignment = center() if col_idx not in (2, 4) else left()
    c.border = thin_border()
    if col_idx == 7:
        c.number_format = "$#,##0.00"
    elif col_idx == 8:
        c.number_format = "$#,##0.00"
    elif col_idx == 9:
        c.number_format = "0.00%"
    elif col_idx == 10:
        c.number_format = "$#,##0.00"
    elif col_idx == 5:
        c.number_format = "#,##0"

# Freeze header rows
ws_h.freeze_panes = "A3"

# ═══════════════════════════════════════════════════════════════════════════════
# SHEET 2: Settings
# ═══════════════════════════════════════════════════════════════════════════════
ws_s = wb.create_sheet("Settings")

settings_title = "DEFENDER CAPITAL MANAGEMENT — MODEL SETTINGS"
ws_s.merge_cells("A1:D1")
ws_s["A1"] = settings_title
ws_s["A1"].font = Font(name="Arial", bold=True, size=13, color=WHITE)
ws_s["A1"].fill = header_fill(NAVY)
ws_s["A1"].alignment = center()
ws_s.row_dimensions[1].height = 28

hdr_row = ["Parameter", "Value", "Default", "Description"]
hdr_widths = [30, 22, 22, 60]
for ci, (h, w) in enumerate(zip(hdr_row, hdr_widths), start=1):
    c = ws_s.cell(row=2, column=ci, value=h)
    c.font = Font(name="Arial", bold=True, size=10, color=WHITE)
    c.fill = header_fill(NAVY)
    c.alignment = center()
    c.border = thin_border()
    ws_s.column_dimensions[get_column_letter(ci)].width = w
ws_s.row_dimensions[2].height = 22

settings = [
    ("portfolio_name",           "Defender Capital Management", "Defender Capital Management", "Full organization or portfolio name. Displayed in sidebar, report headers, and page captions."),
    ("portfolio_short_name",     "DCM",    "DCM",    "Short name or abbreviation. Used in table headers, filenames, and compact labels."),
    ("benchmark_ticker",         "SPY",    "SPY",    "Ticker for the benchmark index. SPY = S&P 500 ETF (recommended). Use QQQ for Nasdaq, IWM for Russell 2000."),
    ("risk_free_rate",           "auto",   "auto",   "Risk-free rate used in Sharpe/Sortino calculations. 'auto' = pulls current 10-year Treasury yield (^TNX). Override with a decimal like 0.043 for 4.3%."),
    ("confidence_level_1",       0.95,     0.95,     "First VaR confidence level. 0.95 means 95% confidence (5% tail)."),
    ("confidence_level_2",       0.99,     0.99,     "Second VaR confidence level. 0.99 means 99% confidence (1% tail)."),
    ("lookback_years",           2,        2,        "Years of historical price data to download. 2 years = ~504 trading days. Increase to 5 for more stability."),
    ("simulation_paths",         10000,    10000,    "Number of Monte Carlo simulation paths. 10,000 is standard. Increase to 50,000 for smoother results (slower run)."),
    ("simulation_days",          252,      252,      "Number of trading days to simulate forward. 252 = 1 year."),
    ("stress_custom_drawdown",   -0.20,    -0.20,    "Custom stress test scenario: uniform portfolio drawdown. -0.20 = down 20%. Change to any negative decimal."),
    ("report_title",             "Defender Capital Management — Portfolio Risk Report", "Defender Capital Management — Portfolio Risk Report", "Title that appears on generated reports."),
    ("color_primary",            "#1B2A4A","#1B2A4A","Primary brand color (dark navy). Used in report headers and dashboard."),
    ("color_secondary",          "#C0C0C0","#C0C0C0","Secondary brand color (silver). Used for borders and accents."),
    ("color_accent",             "#4A90D9","#4A90D9","Accent color (blue). Used for interactive elements and highlights."),
    ("max_position_warning_pct", 0.10,     0.10,     "Portfolio weight above which a position is flagged as concentrated. 0.10 = warn if any holding > 10% of portfolio."),
    ("min_data_points",          100,      100,      "Minimum trading days of data required for a ticker to be included in risk calculations."),
]

for row_num, (param, val, default, desc) in enumerate(settings, start=3):
    rfl = PatternFill("solid", start_color=LIGHT, fgColor=LIGHT) if row_num % 2 == 0 else PatternFill("solid", start_color=WHITE, fgColor=WHITE)

    c = ws_s.cell(row=row_num, column=1, value=param)
    c.font = Font(name="Arial", bold=True, size=10, color=NAVY)
    c.fill = rfl; c.alignment = left(); c.border = thin_border()

    c = ws_s.cell(row=row_num, column=2, value=val)
    c.font = Font(name="Arial", size=10, color="0000FF")
    c.fill = PatternFill("solid", start_color=YELLOW, fgColor=YELLOW)
    c.alignment = center(); c.border = thin_border()
    if isinstance(val, float) and abs(val) < 1 and val != 0.0:
        c.number_format = "0.00%"

    c = ws_s.cell(row=row_num, column=3, value=default)
    c.font = body_font(color="888888"); c.fill = rfl; c.alignment = center(); c.border = thin_border()

    c = ws_s.cell(row=row_num, column=4, value=desc)
    c.font = body_font(); c.fill = rfl; c.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    c.border = thin_border()
    ws_s.row_dimensions[row_num].height = 30

ws_s.freeze_panes = "A3"

# ═══════════════════════════════════════════════════════════════════════════════
# SHEET 3: Instructions
# ═══════════════════════════════════════════════════════════════════════════════
ws_i = wb.create_sheet("Instructions")

ws_i.column_dimensions["A"].width = 100

ws_i.merge_cells("A1:A1")
ws_i["A1"] = "DEFENDER CAPITAL MANAGEMENT — HOW TO USE THE PORTFOLIO RISK SYSTEM"
ws_i["A1"].font = Font(name="Arial", bold=True, size=14, color=WHITE)
ws_i["A1"].fill = header_fill(NAVY)
ws_i["A1"].alignment = center()
ws_i.row_dimensions[1].height = 32

instructions = [
    ("", ""),
    ("QUICK START — 3 STEPS TO UPDATE THE MODEL", "header"),
    ("  Step 1:  Update this file — add/remove holdings, adjust share counts", "step"),
    ("  Step 2:  Double-click 'run_dashboard' script in the scripts/ folder", "step"),
    ("  Step 3:  Click the blue '🔄 Refresh Model' button in the dashboard", "step"),
    ("", ""),
    ("HOW TO ADD A NEW HOLDING", "header"),
    ("  1. Go to the 'Holdings' sheet.", "body"),
    ("  2. Click on any empty row below the last holding (before row 38 Total).", "body"),
    ("  3. Type the ticker symbol in column A (e.g., TSLA). It MUST be exact — use Yahoo Finance to verify.", "body"),
    ("  4. Type the number of shares in column E (Shares Held).", "body"),
    ("  5. Type your average cost per share in column F (Avg Cost Basis).", "body"),
    ("  6. Leave all other columns alone — the engine will auto-fill Company Name, Sector, Industry,", "body"),
    ("     Current Price, Market Value, Weight, and P&L when you click Refresh.", "body"),
    ("  7. Save the file, then click Refresh in the dashboard.", "body"),
    ("", ""),
    ("HOW TO REMOVE A HOLDING", "header"),
    ("  1. Go to the 'Holdings' sheet.", "body"),
    ("  2. Right-click the row number of the holding you want to remove.", "body"),
    ("  3. Select 'Delete Row' (NOT just Clear Contents — the row must be gone entirely).", "body"),
    ("  4. Save the file, then click Refresh.", "body"),
    ("", ""),
    ("HOW TO UPDATE SHARES AFTER A TRADE", "header"),
    ("  1. Go to the 'Holdings' sheet.", "body"),
    ("  2. Find the ticker you traded.", "body"),
    ("  3. Update column E (Shares Held) to the new total shares owned.", "body"),
    ("  4. If your average cost changed, update column F (Avg Cost Basis).", "body"),
    ("  5. Save and click Refresh.", "body"),
    ("", ""),
    ("HOW TO CHANGE THE BENCHMARK OR RISK-FREE RATE", "header"),
    ("  1. Go to the 'Settings' sheet.", "body"),
    ("  2. Find the row for 'benchmark_ticker' and change the value in the yellow column.", "body"),
    ("     Examples: SPY = S&P 500, QQQ = Nasdaq-100, IWM = Russell 2000", "body"),
    ("  3. To override the risk-free rate, change 'risk_free_rate' from 'auto' to a decimal.", "body"),
    ("     Example: 0.043 = 4.3%. Leave as 'auto' to use the current 10-year Treasury yield.", "body"),
    ("  4. Save and click Refresh.", "body"),
    ("", ""),
    ("⚠️  THINGS YOU MUST NOT CHANGE ⚠️", "warning"),
    ("  ✗  Do NOT rename any sheet (Holdings, Settings, Instructions).", "warning_body"),
    ("  ✗  Do NOT change any column headers in the Holdings sheet.", "warning_body"),
    ("  ✗  Do NOT modify the formula cells (columns G through K are managed by the engine).", "warning_body"),
    ("  ✗  Do NOT insert rows above row 3 in the Holdings sheet.", "warning_body"),
    ("  ✗  Do NOT delete the total row (row 38).", "warning_body"),
    ("  ✓  You CAN add rows anywhere between row 3 and row 37.", "body"),
    ("  ✓  You CAN change values in columns A, B, E, and F.", "body"),
    ("  ✓  You CAN change any yellow cell in the Settings sheet.", "body"),
    ("", ""),
    ("COMMON MISTAKES AND FIXES", "header"),
    ("  Problem: Red error message 'Ticker APPL not recognized'", "body"),
    ("  Fix:     Check for typos. AAPL not APPL. Use Yahoo Finance to verify the exact ticker.", "body"),
    ("", ""),
    ("  Problem: 'Shares Held must be a positive number'", "body"),
    ("  Fix:     Make sure the Shares column contains a number, not text. Delete and retype.", "body"),
    ("", ""),
    ("  Problem: Dashboard shows '$0.00' for all prices", "body"),
    ("  Fix:     Click the Refresh button. Prices are not live — they update on Refresh.", "body"),
    ("", ""),
    ("  Problem: 'yfinance error for ticker XYZ'", "body"),
    ("  Fix:     The stock may be delisted or the ticker changed. Check Yahoo Finance.", "body"),
    ("", ""),
    ("  Problem: Dashboard won't open", "body"),
    ("  Fix:     Make sure you double-clicked the 'run_dashboard' script (NOT the Python files).", "body"),
    ("           If it still won't open, run setup again: double-click 'setup'.", "body"),
    ("", ""),
    ("COLUMN REFERENCE GUIDE (Holdings Sheet)", "header"),
    ("  A — Ticker:          YOU fill this in. Must be an exact US stock ticker.", "body"),
    ("  B — Company Name:    Auto-filled by the engine on Refresh.", "body"),
    ("  C — Sector:          Auto-filled by the engine on Refresh.", "body"),
    ("  D — Industry:        Auto-filled by the engine on Refresh.", "body"),
    ("  E — Shares Held:     YOU fill this in. Whole numbers only.", "body"),
    ("  F — Avg Cost Basis:  YOU fill this in. Price per share you paid on average.", "body"),
    ("  G — Current Price:   Updated by the engine on Refresh.", "body"),
    ("  H — Market Value:    Formula: Shares × Current Price. Updates automatically.", "body"),
    ("  I — Weight (%):      Formula: Market Value ÷ Total Portfolio. Updates automatically.", "body"),
    ("  J — Unrealized P&L ($): Formula: (Current Price − Cost Basis) × Shares.", "body"),
    ("  K — Unrealized P&L (%): Formula: (Current Price − Cost Basis) ÷ Cost Basis.", "body"),
]

for row_num, (text, style) in enumerate(instructions, start=2):
    c = ws_i.cell(row=row_num, column=1, value=text)
    ws_i.row_dimensions[row_num].height = 18
    if style == "header":
        c.font = Font(name="Arial", bold=True, size=11, color=WHITE)
        c.fill = header_fill(NAVY)
        c.alignment = left()
        ws_i.row_dimensions[row_num].height = 22
    elif style == "step":
        c.font = Font(name="Arial", bold=True, size=10, color=NAVY)
        c.fill = PatternFill("solid", start_color=LIGHT, fgColor=LIGHT)
        c.alignment = left()
    elif style == "warning":
        c.font = Font(name="Arial", bold=True, size=10, color="CC0000")
        c.fill = PatternFill("solid", start_color="FFEBEE", fgColor="FFEBEE")
        c.alignment = left()
        ws_i.row_dimensions[row_num].height = 22
    elif style == "warning_body":
        c.font = Font(name="Arial", size=10, color="CC0000")
        c.fill = PatternFill("solid", start_color="FFEBEE", fgColor="FFEBEE")
        c.alignment = left()
    else:
        c.font = body_font()
        c.fill = PatternFill("solid", start_color=WHITE, fgColor=WHITE)
        c.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)

# ═══════════════════════════════════════════════════════════════════════════════
# Save
# ═══════════════════════════════════════════════════════════════════════════════
OUTPUT = "/sessions/optimistic-magical-mccarthy/mnt/Risk Model/DefenderCapital/data/portfolio_holdings.xlsx"
wb.save(OUTPUT)
print(f"✅ Created: {OUTPUT}")
print(f"   Holdings: {len(holdings)} positions")
print(f"   Sheets: {wb.sheetnames}")
