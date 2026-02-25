"""
engine/market_data.py — Market data fetcher with caching and auto-population.

Downloads adjusted close prices from yfinance, caches results locally,
auto-populates company metadata, and fetches the live risk-free rate.
"""
from __future__ import annotations

import pickle
import time
from datetime import datetime, timedelta
from pathlib import Path
from typing import Optional

import numpy as np
import pandas as pd
import yfinance as yf

from engine.data_loader import Holding, LoadResult, PortfolioSettings
from engine.utils import (
    get_logger, get_cache_path, retry, timer,
    CACHE_DIR, EXPORTS_DIR,
)

log = get_logger("defender.market_data")


# ── Static ticker metadata fallback ──────────────────────────────────────────
# When yfinance .info is rate-limited (403) on shared cloud infrastructure
# (e.g. Streamlit Cloud), we fall back to this table for sector/industry/name.
# Format: ticker → (shortName, sector, industry)
_STATIC_META: dict[str, tuple[str, str, str]] = {
    # ── Technology ────────────────────────────────────────────────────────────
    "AAPL":  ("Apple Inc.",                         "Technology",             "Consumer Electronics"),
    "MSFT":  ("Microsoft Corporation",              "Technology",             "Software - Infrastructure"),
    "GOOG":  ("Alphabet Inc.",                      "Communication Services", "Internet Content & Information"),
    "GOOGL": ("Alphabet Inc.",                      "Communication Services", "Internet Content & Information"),
    "AMZN":  ("Amazon.com, Inc.",                   "Consumer Cyclical",      "Internet Retail"),
    "META":  ("Meta Platforms, Inc.",               "Communication Services", "Internet Content & Information"),
    "NVDA":  ("NVIDIA Corporation",                 "Technology",             "Semiconductors"),
    "AMD":   ("Advanced Micro Devices, Inc.",       "Technology",             "Semiconductors"),
    "TSM":   ("Taiwan Semiconductor Mfg.",          "Technology",             "Semiconductors"),
    "QCOM":  ("QUALCOMM Incorporated",             "Technology",             "Semiconductors"),
    "CSCO":  ("Cisco Systems, Inc.",                "Technology",             "Communication Equipment"),
    "ADBE":  ("Adobe Inc.",                         "Technology",             "Software - Infrastructure"),
    "PYPL":  ("PayPal Holdings, Inc.",              "Financial Services",     "Credit Services"),
    "INTC":  ("Intel Corporation",                  "Technology",             "Semiconductors"),
    "ORCL":  ("Oracle Corporation",                 "Technology",             "Software - Infrastructure"),
    "CRM":   ("Salesforce, Inc.",                   "Technology",             "Software - Application"),
    "AVGO":  ("Broadcom Inc.",                      "Technology",             "Semiconductors"),
    "TXN":   ("Texas Instruments Incorporated",     "Technology",             "Semiconductors"),
    "NOW":   ("ServiceNow, Inc.",                   "Technology",             "Software - Application"),
    "IBM":   ("International Business Machines",    "Technology",             "Information Technology Services"),
    "SHOP":  ("Shopify Inc.",                       "Technology",             "Software - Application"),
    "SQ":    ("Block, Inc.",                        "Technology",             "Software - Infrastructure"),
    "SNOW":  ("Snowflake Inc.",                     "Technology",             "Software - Application"),
    "PLTR":  ("Palantir Technologies Inc.",         "Technology",             "Software - Infrastructure"),
    "NET":   ("Cloudflare, Inc.",                   "Technology",             "Software - Infrastructure"),
    "MU":    ("Micron Technology, Inc.",            "Technology",             "Semiconductors"),
    "AMAT":  ("Applied Materials, Inc.",            "Technology",             "Semiconductor Equipment"),
    "LRCX":  ("Lam Research Corporation",           "Technology",             "Semiconductor Equipment"),
    "KLAC":  ("KLA Corporation",                    "Technology",             "Semiconductor Equipment"),
    "MRVL":  ("Marvell Technology, Inc.",           "Technology",             "Semiconductors"),
    "PANW":  ("Palo Alto Networks, Inc.",           "Technology",             "Software - Infrastructure"),
    "CRWD":  ("CrowdStrike Holdings, Inc.",         "Technology",             "Software - Infrastructure"),
    "ZS":    ("Zscaler, Inc.",                      "Technology",             "Software - Infrastructure"),

    # ── Financial Services ────────────────────────────────────────────────────
    "BLK":   ("BlackRock, Inc.",                    "Financial Services",     "Asset Management"),
    "MA":    ("Mastercard Incorporated",            "Financial Services",     "Credit Services"),
    "V":     ("Visa Inc.",                          "Financial Services",     "Credit Services"),
    "WFC":   ("Wells Fargo & Company",              "Financial Services",     "Banks - Diversified"),
    "USB":   ("U.S. Bancorp",                       "Financial Services",     "Banks - Regional"),
    "JPM":   ("JPMorgan Chase & Co.",               "Financial Services",     "Banks - Diversified"),
    "BAC":   ("Bank of America Corporation",        "Financial Services",     "Banks - Diversified"),
    "GS":    ("The Goldman Sachs Group, Inc.",      "Financial Services",     "Capital Markets"),
    "MS":    ("Morgan Stanley",                     "Financial Services",     "Capital Markets"),
    "C":     ("Citigroup Inc.",                     "Financial Services",     "Banks - Diversified"),
    "SCHW":  ("The Charles Schwab Corporation",     "Financial Services",     "Capital Markets"),
    "AXP":   ("American Express Company",           "Financial Services",     "Credit Services"),
    "SPGI":  ("S&P Global Inc.",                    "Financial Services",     "Financial Data & Stock Exchanges"),
    "CME":   ("CME Group Inc.",                     "Financial Services",     "Financial Data & Stock Exchanges"),
    "ICE":   ("Intercontinental Exchange, Inc.",     "Financial Services",     "Financial Data & Stock Exchanges"),
    "BRK-B": ("Berkshire Hathaway Inc.",            "Financial Services",     "Insurance - Diversified"),
    "CBRE":  ("CBRE Group, Inc.",                   "Real Estate",           "Real Estate Services"),

    # ── Healthcare ────────────────────────────────────────────────────────────
    "PFE":   ("Pfizer Inc.",                        "Healthcare",            "Drug Manufacturers - General"),
    "SYK":   ("Stryker Corporation",                "Healthcare",            "Medical Devices"),
    "ELV":   ("Elevance Health, Inc.",              "Healthcare",            "Healthcare Plans"),
    "JNJ":   ("Johnson & Johnson",                  "Healthcare",            "Drug Manufacturers - General"),
    "UNH":   ("UnitedHealth Group Incorporated",    "Healthcare",            "Healthcare Plans"),
    "LLY":   ("Eli Lilly and Company",              "Healthcare",            "Drug Manufacturers - General"),
    "ABBV":  ("AbbVie Inc.",                        "Healthcare",            "Drug Manufacturers - General"),
    "MRK":   ("Merck & Co., Inc.",                  "Healthcare",            "Drug Manufacturers - General"),
    "TMO":   ("Thermo Fisher Scientific Inc.",      "Healthcare",            "Diagnostics & Research"),
    "ABT":   ("Abbott Laboratories",                "Healthcare",            "Medical Devices"),
    "DHR":   ("Danaher Corporation",                "Healthcare",            "Diagnostics & Research"),
    "BMY":   ("Bristol-Myers Squibb Company",       "Healthcare",            "Drug Manufacturers - General"),
    "AMGN":  ("Amgen Inc.",                         "Healthcare",            "Drug Manufacturers - General"),
    "GILD":  ("Gilead Sciences, Inc.",              "Healthcare",            "Drug Manufacturers - General"),
    "ISRG":  ("Intuitive Surgical, Inc.",           "Healthcare",            "Medical Instruments"),
    "VRTX":  ("Vertex Pharmaceuticals Incorporated","Healthcare",            "Biotechnology"),
    "MDT":   ("Medtronic plc",                      "Healthcare",            "Medical Devices"),

    # ── Consumer Cyclical ─────────────────────────────────────────────────────
    "DIS":   ("The Walt Disney Company",            "Communication Services", "Entertainment"),
    "NKE":   ("NIKE, Inc.",                         "Consumer Cyclical",      "Footwear & Accessories"),
    "CROX":  ("Crocs, Inc.",                        "Consumer Cyclical",      "Footwear & Accessories"),
    "TSLA":  ("Tesla, Inc.",                        "Consumer Cyclical",      "Auto Manufacturers"),
    "HD":    ("The Home Depot, Inc.",               "Consumer Cyclical",      "Home Improvement Retail"),
    "LOW":   ("Lowe's Companies, Inc.",             "Consumer Cyclical",      "Home Improvement Retail"),
    "MCD":   ("McDonald's Corporation",             "Consumer Cyclical",      "Restaurants"),
    "SBUX":  ("Starbucks Corporation",              "Consumer Cyclical",      "Restaurants"),
    "TJX":   ("The TJX Companies, Inc.",            "Consumer Cyclical",      "Apparel Retail"),
    "BKNG":  ("Booking Holdings Inc.",              "Consumer Cyclical",      "Travel Services"),
    "ABNB":  ("Airbnb, Inc.",                       "Consumer Cyclical",      "Travel Services"),
    "LULU":  ("Lululemon Athletica Inc.",           "Consumer Cyclical",      "Apparel Retail"),
    "GM":    ("General Motors Company",             "Consumer Cyclical",      "Auto Manufacturers"),
    "F":     ("Ford Motor Company",                 "Consumer Cyclical",      "Auto Manufacturers"),
    "CMG":   ("Chipotle Mexican Grill, Inc.",       "Consumer Cyclical",      "Restaurants"),

    # ── Consumer Defensive ────────────────────────────────────────────────────
    "DG":    ("Dollar General Corporation",         "Consumer Defensive",     "Discount Stores"),
    "PG":    ("The Procter & Gamble Company",       "Consumer Defensive",     "Household & Personal Products"),
    "KO":    ("The Coca-Cola Company",              "Consumer Defensive",     "Beverages - Non-Alcoholic"),
    "PEP":   ("PepsiCo, Inc.",                      "Consumer Defensive",     "Beverages - Non-Alcoholic"),
    "WMT":   ("Walmart Inc.",                       "Consumer Defensive",     "Discount Stores"),
    "COST":  ("Costco Wholesale Corporation",       "Consumer Defensive",     "Discount Stores"),
    "PM":    ("Philip Morris International Inc.",    "Consumer Defensive",     "Tobacco"),
    "MO":    ("Altria Group, Inc.",                  "Consumer Defensive",     "Tobacco"),
    "CL":    ("Colgate-Palmolive Company",          "Consumer Defensive",     "Household & Personal Products"),
    "MDLZ":  ("Mondelez International, Inc.",       "Consumer Defensive",     "Confectioners"),

    # ── Industrials ───────────────────────────────────────────────────────────
    "RTX":   ("RTX Corporation",                    "Industrials",            "Aerospace & Defense"),
    "DE":    ("Deere & Company",                    "Industrials",            "Farm & Heavy Construction Machinery"),
    "CARR":  ("Carrier Global Corporation",         "Industrials",            "Building Products & Equipment"),
    "BA":    ("The Boeing Company",                 "Industrials",            "Aerospace & Defense"),
    "CAT":   ("Caterpillar Inc.",                   "Industrials",            "Farm & Heavy Construction Machinery"),
    "HON":   ("Honeywell International Inc.",       "Industrials",            "Conglomerates"),
    "UPS":   ("United Parcel Service, Inc.",        "Industrials",            "Integrated Freight & Logistics"),
    "GE":    ("GE Aerospace",                       "Industrials",            "Aerospace & Defense"),
    "LMT":   ("Lockheed Martin Corporation",        "Industrials",            "Aerospace & Defense"),
    "UNP":   ("Union Pacific Corporation",          "Industrials",            "Railroads"),
    "WM":    ("Waste Management, Inc.",             "Industrials",            "Waste Management"),
    "ETN":   ("Eaton Corporation plc",              "Industrials",            "Electrical Equipment"),
    "MMM":   ("3M Company",                         "Industrials",            "Conglomerates"),

    # ── Energy ────────────────────────────────────────────────────────────────
    "CVX":   ("Chevron Corporation",                "Energy",                 "Oil & Gas Integrated"),
    "XOM":   ("Exxon Mobil Corporation",            "Energy",                 "Oil & Gas Integrated"),
    "COP":   ("ConocoPhillips",                     "Energy",                 "Oil & Gas E&P"),
    "SLB":   ("Schlumberger Limited",               "Energy",                 "Oil & Gas Equipment & Services"),
    "EOG":   ("EOG Resources, Inc.",                "Energy",                 "Oil & Gas E&P"),
    "MPC":   ("Marathon Petroleum Corporation",     "Energy",                 "Oil & Gas Refining & Marketing"),
    "PSX":   ("Phillips 66",                        "Energy",                 "Oil & Gas Refining & Marketing"),
    "OXY":   ("Occidental Petroleum Corporation",   "Energy",                 "Oil & Gas E&P"),
    "SMR":   ("NuScale Power Corporation",          "Energy",                 "Specialty Industrial Machinery"),

    # ── Communication Services ────────────────────────────────────────────────
    "T":     ("AT&T Inc.",                          "Communication Services", "Telecom Services"),
    "VZ":    ("Verizon Communications Inc.",         "Communication Services", "Telecom Services"),
    "TMUS":  ("T-Mobile US, Inc.",                  "Communication Services", "Telecom Services"),
    "NFLX":  ("Netflix, Inc.",                      "Communication Services", "Entertainment"),
    "CMCSA": ("Comcast Corporation",                "Communication Services", "Telecom Services"),
    "ATVI":  ("Activision Blizzard, Inc.",          "Communication Services", "Electronic Gaming & Multimedia"),

    # ── Utilities ─────────────────────────────────────────────────────────────
    "EXC":   ("Exelon Corporation",                 "Utilities",              "Utilities - Regulated Electric"),
    "NEE":   ("NextEra Energy, Inc.",               "Utilities",              "Utilities - Regulated Electric"),
    "DUK":   ("Duke Energy Corporation",            "Utilities",              "Utilities - Regulated Electric"),
    "SO":    ("The Southern Company",               "Utilities",              "Utilities - Regulated Electric"),
    "D":     ("Dominion Energy, Inc.",              "Utilities",              "Utilities - Regulated Electric"),
    "AEP":   ("American Electric Power Company",    "Utilities",              "Utilities - Regulated Electric"),

    # ── Real Estate ───────────────────────────────────────────────────────────
    "DLR":   ("Digital Realty Trust, Inc.",          "Real Estate",            "REIT - Specialty"),
    "AMT":   ("American Tower Corporation",         "Real Estate",            "REIT - Specialty"),
    "PLD":   ("Prologis, Inc.",                     "Real Estate",            "REIT - Industrial"),
    "CCI":   ("Crown Castle Inc.",                  "Real Estate",            "REIT - Specialty"),
    "EQIX":  ("Equinix, Inc.",                      "Real Estate",            "REIT - Specialty"),
    "SPG":   ("Simon Property Group, Inc.",         "Real Estate",            "REIT - Retail"),
    "O":     ("Realty Income Corporation",          "Real Estate",            "REIT - Retail"),

    # ── Basic Materials ───────────────────────────────────────────────────────
    "RIO":   ("Rio Tinto Group",                    "Basic Materials",        "Other Industrial Metals & Mining"),
    "LIN":   ("Linde plc",                          "Basic Materials",        "Specialty Chemicals"),
    "APD":   ("Air Products and Chemicals, Inc.",   "Basic Materials",        "Specialty Chemicals"),
    "NEM":   ("Newmont Corporation",                "Basic Materials",        "Gold"),
    "FCX":   ("Freeport-McMoRan Inc.",              "Basic Materials",        "Copper"),
    "SHW":   ("The Sherwin-Williams Company",       "Basic Materials",        "Specialty Chemicals"),
    "ECL":   ("Ecolab Inc.",                        "Basic Materials",        "Specialty Chemicals"),
    "DD":    ("DuPont de Nemours, Inc.",            "Basic Materials",        "Specialty Chemicals"),
    "BHP":   ("BHP Group Limited",                  "Basic Materials",        "Other Industrial Metals & Mining"),
    "VALE":  ("Vale S.A.",                          "Basic Materials",        "Other Industrial Metals & Mining"),

    # ── Other / Specialty ─────────────────────────────────────────────────────
    "LTH":   ("Life Time Group Holdings, Inc.",     "Consumer Cyclical",      "Leisure"),

    # ── Popular ETFs (for benchmark / reference) ──────────────────────────────
    "SPY":   ("SPDR S&P 500 ETF Trust",             "ETF",                    "Large Blend"),
    "QQQ":   ("Invesco QQQ Trust",                  "ETF",                    "Large Growth"),
    "IWM":   ("iShares Russell 2000 ETF",           "ETF",                    "Small Blend"),
    "VTI":   ("Vanguard Total Stock Market ETF",    "ETF",                    "Large Blend"),
    "DIA":   ("SPDR Dow Jones Industrial Avg ETF",  "ETF",                    "Large Value"),
}


def _static_meta_lookup(ticker: str) -> dict:
    """Return a metadata dict from the static fallback table, or empty dict."""
    entry = _STATIC_META.get(ticker.upper())
    if entry:
        return {"shortName": entry[0], "sector": entry[1], "industry": entry[2]}
    return {}


# ── Cache helpers ──────────────────────────────────────────────────────────────
def _cache_valid(path: Path, expiry_hours: int) -> bool:
    """Return True if a cache file exists and was written within expiry_hours."""
    if not path.exists():
        return False
    age = datetime.now() - datetime.fromtimestamp(path.stat().st_mtime)
    return age < timedelta(hours=expiry_hours)


def _load_cache(path: Path):
    """Load a pickled object from disk."""
    with open(path, "rb") as f:
        return pickle.load(f)


def _save_cache(path: Path, obj) -> None:
    """Pickle an object to disk."""
    path.parent.mkdir(parents=True, exist_ok=True)
    with open(path, "wb") as f:
        pickle.dump(obj, f)


# ── Single-ticker metadata fetch ───────────────────────────────────────────────
def _fetch_info(ticker: str) -> dict:
    """
    Fetch ticker metadata using yfinance fast_info (avoids 403 on .info endpoint).
    Returns a dict with keys: shortName, sector, industry, regularMarketPrice.
    """
    t = yf.Ticker(ticker)
    result: dict = {}

    # fast_info is reliable and avoids rate-limit issues
    try:
        fi = t.fast_info
        price = None
        for attr in ("last_price", "regular_market_price", "previous_close"):
            try:
                v = getattr(fi, attr, None)
                if v and float(v) > 0:
                    price = float(v)
                    break
            except Exception:
                pass
        if price:
            result["regularMarketPrice"] = price

        # Try to get quote_type data for name
        try:
            result["shortName"] = str(fi.quote_type or ticker)
        except Exception:
            result["shortName"] = ticker
    except Exception as e:
        log.debug(f"{ticker}: fast_info failed — {e}")

    # Try to get sector/industry from ticker.info (may 403 on some environments)
    # Use a very short timeout and don't retry — it's nice-to-have
    try:
        info = t.info
        if info and isinstance(info, dict):
            result.update({k: v for k, v in info.items()
                           if k in ("shortName", "longName", "sector", "industry",
                                    "regularMarketPrice", "currentPrice")})
    except Exception:
        pass  # 403 or timeout — use fast_info result only

    if not result.get("regularMarketPrice"):
        raise ValueError(f"No price data available for '{ticker}'")

    return result


def _extract_price(info: dict) -> float:
    """Pull the most reliable current price from an info dict."""
    for key in ("currentPrice", "regularMarketPrice", "navPrice", "ask", "bid"):
        val = info.get(key)
        if val and float(val) > 0:
            return float(val)
    return 0.0


# ── Price history fetch ────────────────────────────────────────────────────────
@retry(max_attempts=3, base_delay=2.0)
def _fetch_history(ticker: str, start: str, end: str) -> pd.Series:
    """
    Download adjusted close price history for a single ticker.

    Returns a pd.Series with datetime index and float values.
    """
    data = yf.download(
        ticker,
        start=start,
        end=end,
        auto_adjust=True,
        progress=False,
        actions=False,
    )
    if data.empty:
        raise ValueError(f"yfinance returned no data for '{ticker}' ({start} to {end})")

    # Handle MultiIndex columns (yfinance >= 0.2.x)
    if isinstance(data.columns, pd.MultiIndex):
        data.columns = data.columns.get_level_values(0)

    if "Close" not in data.columns:
        raise ValueError(f"'Close' column not found for '{ticker}'")

    series = data["Close"].dropna()
    series.name = ticker
    return series


# ── Risk-free rate ─────────────────────────────────────────────────────────────
@retry(max_attempts=3, base_delay=1.0)
def fetch_risk_free_rate() -> float:
    """
    Fetch the current 10-year US Treasury yield from Yahoo Finance (^TNX).

    The ^TNX ticker reports yield in percentage points, so we divide by 100.
    Falls back to 4.5% if the fetch fails.

    Returns
    -------
    float
        Annual risk-free rate as a decimal, e.g. 0.045 for 4.5%.
    """
    cache_path = get_cache_path("risk_free_rate")
    if _cache_valid(cache_path, expiry_hours=6):
        rate = _load_cache(cache_path)
        log.debug(f"Risk-free rate from cache: {rate:.4f}")
        return rate

    try:
        t = yf.Ticker("^TNX")
        info = t.info
        rate = None
        for key in ("regularMarketPrice", "currentPrice", "previousClose"):
            val = info.get(key)
            if val and val > 0:
                rate = val / 100.0  # ^TNX is in % (e.g., 4.5 → 0.045)
                break
        if rate is None or not (0.001 < rate < 0.30):
            raise ValueError(f"Implausible rate value: {rate}")
        log.info(f"Risk-free rate (^TNX): {rate:.4f} ({rate*100:.2f}%)")
        _save_cache(cache_path, rate)
        return rate
    except Exception as exc:
        log.warning(f"Could not fetch risk-free rate: {exc}. Using fallback 4.5%.")
        return 0.045


# ── S&P 500 sector weights from ETF market caps ──────────────────────────────
_SECTOR_ETF_TICKERS = {
    "XLK": "Technology",
    "XLF": "Financial Services",
    "XLV": "Healthcare",
    "XLY": "Consumer Cyclical",
    "XLC": "Communication Services",
    "XLI": "Industrials",
    "XLP": "Consumer Defensive",
    "XLE": "Energy",
    "XLU": "Utilities",
    "XLRE": "Real Estate",
    "XLB": "Basic Materials",
}


def fetch_sp500_sector_weights(cache_expiry_hours: int = 24) -> Optional[dict[str, float]]:
    """
    Compute approximate S&P 500 sector weights from sector SPDR ETF market caps.

    Fetches market cap for each of the 11 sector ETFs via yfinance fast_info,
    then computes each sector's weight as its share of the total.

    Returns None on failure (caller should fall back to hardcoded values).
    """
    cache_path = get_cache_path("sp500_sector_weights")
    if _cache_valid(cache_path, expiry_hours=cache_expiry_hours):
        weights = _load_cache(cache_path)
        log.debug(f"S&P 500 sector weights from cache ({len(weights)} sectors)")
        return weights

    try:
        caps: dict[str, float] = {}
        for etf, sector in _SECTOR_ETF_TICKERS.items():
            t = yf.Ticker(etf)
            mc = getattr(t, "fast_info", {}).get("market_cap") or t.info.get("marketCap")
            if mc and mc > 0:
                caps[sector] = float(mc)
            else:
                log.warning(f"No market cap for {etf} ({sector}), skipping")

        if len(caps) < 8:
            raise ValueError(f"Only got market caps for {len(caps)}/11 sector ETFs")

        total = sum(caps.values())
        weights = {sector: mc / total for sector, mc in caps.items()}

        log.info(
            f"S&P 500 sector weights from ETF market caps: "
            + ", ".join(f"{s} {w:.1%}" for s, w in sorted(weights.items(), key=lambda x: -x[1]))
        )
        _save_cache(cache_path, weights)
        return weights
    except Exception as exc:
        log.warning(f"Could not fetch S&P 500 sector weights: {exc}")
        return None


# ── Main market data fetcher ───────────────────────────────────────────────────
@timer
def fetch_market_data(
    load_result: LoadResult,
    settings: Optional[PortfolioSettings] = None,
) -> "MarketData":
    """
    Download and cache all required market data for the portfolio.

    1. For each holding: fetch current price, company name, sector, industry.
    2. Fetch historical adjusted close prices for all tickers + benchmark.
    3. Fetch the risk-free rate (live or cached).

    Parameters
    ----------
    load_result : LoadResult
        Output of data_loader.load_portfolio().
    settings : PortfolioSettings, optional
        Configuration. Defaults to load_result.settings.

    Returns
    -------
    MarketData
        Container with price history DataFrame, current prices, metadata.
    """
    cfg = settings or load_result.settings
    holdings = load_result.holdings

    # MED-2 FIX: Portfolio size guard
    MAX_TICKERS_HARD = 200
    MAX_TICKERS_SOFT = 100
    if len(holdings) > MAX_TICKERS_HARD:
        raise ValueError(
            f"Portfolio has {len(holdings)} holdings, which exceeds the maximum of {MAX_TICKERS_HARD}. "
            f"Please reduce to {MAX_TICKERS_HARD} or fewer holdings."
        )
    if len(holdings) > MAX_TICKERS_SOFT:
        log.warning(
            f"Portfolio has {len(holdings)} holdings (>{MAX_TICKERS_SOFT}). "
            f"Computation may be slow. Consider reducing for better performance."
        )

    tickers = [h.ticker for h in holdings]
    all_tickers = tickers + [cfg.benchmark_ticker]

    end_date   = datetime.now().strftime("%Y-%m-%d")
    start_date = (datetime.now() - timedelta(days=int(cfg.lookback_years * 365.25 + 10))).strftime("%Y-%m-%d")

    log.info(
        f"Fetching market data for {len(tickers)} holdings + benchmark ({cfg.benchmark_ticker}), "
        f"{start_date} to {end_date}"
    )

    data_quality_log: list[str] = []
    failed_tickers: list[str] = []

    # ── Historical price series (batch download first — needed for current prices too)
    price_cache_path = get_cache_path("price_history")
    tickers_key      = "_".join(sorted(all_tickers)) + f"_{start_date}"
    tickers_key_path = get_cache_path(f"prices_key")

    use_cache = False
    if _cache_valid(price_cache_path, cfg.cache_expiry_hours):
        try:
            cached_key = _load_cache(tickers_key_path)
            if cached_key == tickers_key:
                use_cache = True
        except Exception:
            pass

    if use_cache:
        prices_df = _load_cache(price_cache_path)
        log.info("Price history loaded from cache.")
    else:
        t0 = time.perf_counter()
        prices_df = _batch_fetch_prices(all_tickers, start_date, end_date, data_quality_log, failed_tickers)
        elapsed = time.perf_counter() - t0
        log.info(f"Price history fetched in {elapsed:.1f}s for {len(prices_df.columns)} tickers.")
        _save_cache(price_cache_path, prices_df)
        _save_cache(tickers_key_path, tickers_key)

    # Use last available price from history as current price
    # This is more reliable than yfinance .info/.fast_info in restricted environments
    latest_prices: dict[str, float] = {}
    for ticker in all_tickers:
        if ticker in prices_df.columns:
            last = prices_df[ticker].dropna()
            if len(last) > 0:
                latest_prices[ticker] = float(last.iloc[-1])

    # ── Individual ticker metadata (optional — nice to have, not critical) ─────
    # Try to get company name / sector from yfinance, but use Excel pre-populated
    # values as fallback. Never block on this.
    def _try_fetch_meta(ticker: str) -> dict:
        cache_path = get_cache_path(f"meta_{ticker}")
        if _cache_valid(cache_path, cfg.cache_expiry_hours * 24):  # meta cached longer
            return _load_cache(cache_path)
        try:
            result = _fetch_info(ticker)
            _save_cache(cache_path, result)
            return result
        except Exception:
            return {}

    # Batch metadata fetch (parallel, best-effort)
    from concurrent.futures import ThreadPoolExecutor, as_completed
    meta_map: dict[str, dict] = {}
    with ThreadPoolExecutor(max_workers=6) as pool:
        futures = {pool.submit(_try_fetch_meta, t): t for t in tickers}
        for fut in as_completed(futures, timeout=30):
            try:
                ticker = futures[fut]
                meta_map[ticker] = fut.result()
            except Exception:
                pass

    # Apply prices and metadata to holdings
    for h in holdings:
        # Current price: prefer live (from history), fall back to stored value
        h.current_price = latest_prices.get(h.ticker, h.current_price)

        # Metadata: prefer yfinance, then static fallback, then Excel values
        meta = meta_map.get(h.ticker, {})
        static = _static_meta_lookup(h.ticker)
        if not h.company_name or h.company_name in ("nan", ""):
            h.company_name = (meta.get("shortName") or meta.get("longName")
                              or static.get("shortName") or h.ticker)
        if not h.sector or h.sector in ("nan", ""):
            h.sector = meta.get("sector") or static.get("sector", "Unknown")
        if not h.industry or h.industry in ("nan", ""):
            h.industry = meta.get("industry") or static.get("industry", "Unknown")

        # MED-5 FIX: Flag non-USD tickers for currency risk warning
        _currency = meta.get("currency") or meta.get("financialCurrency", "USD")
        if _currency and _currency.upper() not in ("USD", "USX", ""):
            data_quality_log.append(
                f"CURRENCY [{h.ticker}]: Reports in {_currency}, not USD. "
                f"Currency risk is NOT modelled — P&L treated as USD."
            )

        # Recalculate derived fields
        h.market_value   = h.shares_held * h.current_price
        h.unrealized_pnl = (h.current_price - h.cost_basis) * h.shares_held
        h.unrealized_pct = (
            (h.current_price - h.cost_basis) / h.cost_basis
            if h.cost_basis > 0 else 0.0
        )

    # Recalculate weights (requires all prices to be set first)
    total_value = sum(h.market_value for h in holdings)
    for h in holdings:
        h.weight = h.market_value / total_value if total_value > 0 else 0.0

    # ── Risk-free rate ─────────────────────────────────────────────────────────
    rfr_override = cfg.risk_free_rate_value
    rfr_warning  = ""
    if rfr_override is not None:
        risk_free_rate = rfr_override
        log.info(f"Risk-free rate (manual override): {risk_free_rate:.4f}")
    else:
        risk_free_rate = fetch_risk_free_rate()
        # HIGH-5 FIX: Check if we're using the hardcoded fallback
        cache_path = get_cache_path("risk_free_rate")
        if not _cache_valid(cache_path, expiry_hours=6):
            rfr_warning = (
                f"Risk-free rate fallback used (4.5%). Live ^TNX fetch failed. "
                f"Sharpe ratio and other RFR-dependent metrics may be slightly off."
            )
            log.warning(rfr_warning)

    # MED-5 FIX: Check if any non-USD tickers were detected in the quality log
    _currency_warnings = [l for l in data_quality_log if l.startswith("CURRENCY")]
    _currency_warning_str = ""
    if _currency_warnings:
        _unique_tickers = [w.split("[")[1].split("]")[0] for w in _currency_warnings]
        _currency_warning_str = (
            f"Non-USD tickers detected: {', '.join(_unique_tickers)}. "
            f"Currency risk is NOT modelled — all P&L treated as USD. "
            f"Foreign exchange fluctuations will affect actual returns."
        )
        log.warning(_currency_warning_str)

    # ── Data quality report ────────────────────────────────────────────────────
    missing_cols = [t for t in tickers if t not in prices_df.columns]
    total_nas    = int(prices_df[tickers].isna().sum().sum())

    quality = DataQualityReport(
        valid_tickers=[t for t in tickers if t not in failed_tickers],
        failed_tickers=failed_tickers,
        date_range=(str(prices_df.index.min().date()), str(prices_df.index.max().date())),
        total_rows=len(prices_df),
        missing_data_points=total_nas,
        missing_tickers=missing_cols,
        log_lines=data_quality_log,
        rfr_warning=rfr_warning,
        currency_warning=_currency_warning_str,
    )
    log.info(str(quality))

    return MarketData(
        prices=prices_df,
        holdings=holdings,
        benchmark_ticker=cfg.benchmark_ticker,
        risk_free_rate=risk_free_rate,
        quality=quality,
    )


def _batch_fetch_prices(
    tickers: list[str],
    start: str,
    end: str,
    dq_log: list[str],
    failed: list[str],
) -> pd.DataFrame:
    """
    Download price history for all tickers using yfinance batch download.
    Handles failures gracefully and falls back to individual fetches.
    Returns a DataFrame with one column per ticker (forward-filled).
    """
    import yfinance as yf

    log.info(f"Batch downloading {len(tickers)} tickers via yfinance…")

    # Try batch download first (much faster)
    try:
        raw = yf.download(
            tickers,
            start=start,
            end=end,
            auto_adjust=True,
            progress=False,
            actions=False,
            group_by="ticker",
            threads=True,
        )

        # Handle MultiIndex columns
        if isinstance(raw.columns, pd.MultiIndex):
            # Extract "Close" for each ticker
            frames = {}
            for ticker in tickers:
                try:
                    if ticker in raw.columns.get_level_values(0):
                        s = raw[ticker]["Close"].dropna()
                    elif ticker in raw.columns.get_level_values(1):
                        s = raw.xs(ticker, level=1, axis=1)["Close"].dropna()
                    else:
                        # Single-ticker download returns flat columns
                        s = raw["Close"].dropna() if "Close" in raw.columns else pd.Series(dtype=float)
                    if len(s) > 10:
                        frames[ticker] = s
                        dq_log.append(f"OK   {ticker}: {len(s)} data points")
                    else:
                        log.warning(f"{ticker}: insufficient data ({len(s)} points)")
                        failed.append(ticker)
                        dq_log.append(f"FAIL {ticker}: insufficient data")
                except Exception as e:
                    log.warning(f"{ticker}: extraction failed — {e}")
                    failed.append(ticker)
                    dq_log.append(f"FAIL {ticker}: {e}")
        elif "Close" in raw.columns:
            # Single ticker download
            ticker = tickers[0]
            s = raw["Close"].dropna()
            frames = {ticker: s}
            dq_log.append(f"OK   {ticker}: {len(s)} data points")
        else:
            frames = {}

    except Exception as batch_exc:
        log.warning(f"Batch download failed ({batch_exc}), falling back to individual fetches")
        frames = {}

    # Fall back: fetch individually any tickers that are still missing
    missing = [t for t in tickers if t not in frames and t not in failed]
    if missing:
        log.info(f"Fetching {len(missing)} tickers individually…")
        for ticker in missing:
            try:
                s = _fetch_history(ticker, start, end)
                frames[ticker] = s
                dq_log.append(f"OK   {ticker}: {len(s)} data points (individual)")
            except Exception as exc:
                log.error(f"{ticker}: individual fetch failed — {exc}")
                if ticker not in failed:
                    failed.append(ticker)
                dq_log.append(f"FAIL {ticker}: {exc}")

    if not frames:
        raise ValueError("All ticker data fetches failed. Check internet connection.")

    df = pd.DataFrame(frames)
    df.index = pd.to_datetime(df.index)
    df.sort_index(inplace=True)

    # CRIT-2 FIX: Check for excessive consecutive missing values per ticker
    # More than 5 consecutive missing trading days indicates a data failure
    # (not just weekends/holidays) that would be silently masked by ffill.
    MAX_CONSEC_MISSING = 5
    for ticker_col in df.columns:
        series = df[ticker_col]
        consec = 0
        max_consec = 0
        for is_missing in series.isna():
            if is_missing:
                consec += 1
                max_consec = max(max_consec, consec)
            else:
                consec = 0
        if max_consec > MAX_CONSEC_MISSING:
            warn_msg = (
                f"DATA QUALITY WARNING [{ticker_col}]: {max_consec} consecutive "
                f"missing price days detected (threshold: {MAX_CONSEC_MISSING}). "
                f"Forward-fill will mask this gap — verify yfinance data source."
            )
            dq_log.append(warn_msg)
            log.warning(warn_msg)

    # Forward-fill then back-fill to handle weekends/holidays/trading halts
    n_before = int(df.isna().sum().sum())
    filled = df.ffill().bfill()
    if n_before:
        log.debug(f"Forward/back-filled {n_before} missing price points.")

    return filled


# ── MarketData container ───────────────────────────────────────────────────────
class DataQualityReport:
    """Summary of data quality after a market data fetch."""

    def __init__(
        self,
        valid_tickers:    list[str],
        failed_tickers:   list[str],
        date_range:       tuple[str, str],
        total_rows:       int,
        missing_data_points: int,
        missing_tickers:  list[str],
        log_lines:        list[str],
        rfr_warning:      str = "",
        currency_warning: str = "",
    ):
        self.valid_tickers       = valid_tickers
        self.failed_tickers      = failed_tickers
        self.date_range          = date_range
        self.total_rows          = total_rows
        self.missing_data_points = missing_data_points
        self.missing_tickers     = missing_tickers
        self.log_lines           = log_lines
        self.rfr_warning         = rfr_warning  # HIGH-5: non-empty = stale/fallback rate used
        self.currency_warning    = currency_warning  # MED-5: non-empty if non-USD tickers detected

    def __str__(self) -> str:
        return (
            f"DataQuality: {len(self.valid_tickers)} valid tickers, "
            f"{len(self.failed_tickers)} failed, "
            f"range={self.date_range[0]}..{self.date_range[1]}, "
            f"rows={self.total_rows}, missing_points={self.missing_data_points}"
        )

    def to_text(self) -> str:
        lines = [
            "Data Quality Report",
            "=" * 50,
            f"Valid tickers ({len(self.valid_tickers)}): {', '.join(self.valid_tickers)}",
            f"Failed tickers ({len(self.failed_tickers)}): {', '.join(self.failed_tickers) or 'None'}",
            f"Date range: {self.date_range[0]} → {self.date_range[1]}",
            f"Total rows: {self.total_rows}",
            f"Missing data points: {self.missing_data_points}",
            f"Missing ticker columns: {', '.join(self.missing_tickers) or 'None'}",
            "",
            "Ticker-level log:",
        ]
        lines.extend(f"  {line}" for line in self.log_lines)
        return "\n".join(lines)


class MarketData:
    """
    Container for all fetched market data, ready for the risk engine.

    Attributes
    ----------
    prices : pd.DataFrame
        Adjusted close prices, columns = tickers, index = trading dates.
    holdings : list[Holding]
        Holdings with current_price, market_value, weight, P&L populated.
    benchmark_ticker : str
        Ticker of the benchmark index.
    risk_free_rate : float
        Annual risk-free rate (decimal).
    quality : DataQualityReport
        Summary of any data issues encountered.
    """

    def __init__(
        self,
        prices:           pd.DataFrame,
        holdings:         list[Holding],
        benchmark_ticker: str,
        risk_free_rate:   float,
        quality:          DataQualityReport,
    ):
        self.prices           = prices
        self.holdings         = holdings
        self.benchmark_ticker = benchmark_ticker
        self.risk_free_rate   = risk_free_rate
        self.quality          = quality

    @property
    def portfolio_tickers(self) -> list[str]:
        """Return list of portfolio tickers that have valid price data."""
        return [
            h.ticker for h in self.holdings
            if h.ticker in self.prices.columns
        ]

    @property
    def weights(self) -> pd.Series:
        """
        Weight Series indexed by ticker, using only tickers with price data.
        Automatically renormalises if some tickers are missing from price data.
        """
        w = {h.ticker: h.weight for h in self.holdings if h.ticker in self.prices.columns}
        s = pd.Series(w)
        total = s.sum()
        return s / total if total > 0 else s

    @property
    def benchmark_prices(self) -> pd.Series:
        """Benchmark close price series."""
        return self.prices[self.benchmark_ticker]

    @property
    def total_portfolio_value(self) -> float:
        """Sum of all holding market values."""
        return sum(h.market_value for h in self.holdings)

    def portfolio_returns(self, log_returns: bool = False) -> pd.Series:
        """
        Compute daily portfolio returns using current weights.

        Parameters
        ----------
        log_returns : bool
            If True, use log returns; else simple returns.

        Returns
        -------
        pd.Series
            Daily portfolio return series.
        """
        tickers = self.portfolio_tickers
        prices  = self.prices[tickers].dropna(how="all")

        if log_returns:
            rets = np.log(prices / prices.shift(1)).dropna()
        else:
            rets = prices.pct_change().dropna()

        w = self.weights.reindex(tickers).fillna(0)
        w = w / w.sum()  # renormalise

        port_returns = (rets * w).sum(axis=1)
        port_returns.name = "Portfolio"
        return port_returns

    def benchmark_returns(self, log_returns: bool = False) -> pd.Series:
        """Compute daily benchmark returns."""
        bp = self.prices[self.benchmark_ticker].dropna()
        if log_returns:
            r = np.log(bp / bp.shift(1)).dropna()
        else:
            r = bp.pct_change().dropna()
        r.name = self.benchmark_ticker
        return r


# ── Write prices back to Excel ─────────────────────────────────────────────────
def update_excel_prices(market_data: MarketData, excel_path: Path) -> None:
    """
    Write current prices back to the Holdings sheet.

    Updates column G (Current Price) for each row. Uses openpyxl to preserve
    all existing formatting and formulas.
    """
    import openpyxl

    log.info(f"Writing prices back to: {excel_path}")
    wb = openpyxl.load_workbook(str(excel_path))
    ws = wb["Holdings"]

    # Build ticker → row mapping (data starts at row 3)
    ticker_to_row: dict[str, int] = {}
    for row in range(3, ws.max_row + 1):
        t = ws.cell(row=row, column=1).value
        if t and str(t).strip().upper() not in ("", "TICKER", "TOTAL PORTFOLIO"):
            ticker_to_row[str(t).strip().upper()] = row

    updated = 0
    for h in market_data.holdings:
        row = ticker_to_row.get(h.ticker)
        if row and h.current_price > 0:
            ws.cell(row=row, column=2).value = h.company_name  or ws.cell(row=row, column=2).value
            ws.cell(row=row, column=3).value = h.sector        or ws.cell(row=row, column=3).value
            ws.cell(row=row, column=4).value = h.industry      or ws.cell(row=row, column=4).value
            ws.cell(row=row, column=7).value = h.current_price
            updated += 1

    wb.save(str(excel_path))
    log.info(f"Updated prices for {updated} holdings in Excel.")


# ── Quick self-test ─────────────────────────────────────────────────────────────
if __name__ == "__main__":
    import sys
    sys.path.insert(0, str(Path(__file__).resolve().parent.parent))

    from engine.data_loader import load_portfolio

    result = load_portfolio()
    md = fetch_market_data(result)

    print(f"\n✅ MarketData ready")
    print(f"   Portfolio value: ${md.total_portfolio_value:,.2f}")
    print(f"   Risk-free rate:  {md.risk_free_rate:.4f}")
    print(f"   Price data:      {md.prices.shape[0]} days × {md.prices.shape[1]} tickers")
    print(f"\n{md.quality}")
