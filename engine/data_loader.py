"""
engine/data_loader.py — Excel reader, validator, and settings loader.

Reads portfolio_holdings.xlsx, validates every row rigorously, and returns
clean Python dataclasses. Handles every common user error gracefully — no
tracebacks are ever exposed to end-users.
"""
from __future__ import annotations

import re
from dataclasses import dataclass, field
from pathlib import Path
from typing import Optional

import pandas as pd

from engine.utils import get_logger, get_portfolio_path, AuditLog

log = get_logger("defender.data_loader")

# ── Known US tickers for fuzzy-match suggestions ──────────────────────────────
# A representative sample; full validation is done by yfinance in market_data.py
_COMMON_TICKERS = {
    "AAPL", "MSFT", "AMZN", "GOOGL", "GOOG", "META", "TSLA", "NVDA", "JPM",
    "V", "UNH", "JNJ", "WMT", "XOM", "CVX", "PG", "HD", "BAC", "MA", "PFE",
    "ABBV", "KO", "LLY", "MRK", "AVGO", "PEP", "TMO", "COST", "CSCO", "MCD",
    "ACN", "NKE", "TXN", "DHR", "NEE", "HON", "ADBE", "LIN", "QCOM", "SBUX",
    "CRM", "GS", "IBM", "CAT", "BA", "RTX", "INTC", "GE", "MMM", "CVS", "CI",
    "UPS", "FDX", "MS", "BLK", "AXP", "SPGI", "ISRG", "GILD", "AMGN", "VRTX",
    "REGN", "MDLZ", "TMUS", "DIS", "CMCSA", "T", "VZ", "NFLX", "PYPL", "SQ",
    "APD", "LMT", "NOC", "GD", "ROP", "AMT", "PLD", "EQIX", "CCI", "SPG",
    "DUK", "SO", "AEP", "D", "EXC", "SRE", "WEC", "XEL", "ES", "ETR",
}


@dataclass
class Holding:
    """A single portfolio position, validated and ready for analysis."""
    ticker:          str
    company_name:    str
    sector:          str
    industry:        str
    shares_held:     int
    cost_basis:      float
    current_price:   float = 0.0
    market_value:    float = 0.0
    weight:          float = 0.0
    unrealized_pnl:  float = 0.0
    unrealized_pct:  float = 0.0
    row_number:      int   = 0      # Excel row for error messages

    @property
    def cost_value(self) -> float:
        return self.shares_held * self.cost_basis


@dataclass
class PortfolioSettings:
    """All configurable parameters read from the Settings sheet."""
    portfolio_name:           str   = "Defender Capital Management"
    portfolio_short_name:     str   = "DCM"
    benchmark_ticker:         str   = "SPY"
    risk_free_rate:           str   = "auto"   # "auto" or a decimal string
    confidence_level_1:       float = 0.95
    confidence_level_2:       float = 0.99
    lookback_years:           int   = 2
    simulation_paths:         int   = 10_000
    simulation_days:          int   = 252
    stress_custom_drawdown:   float = -0.20
    report_title:             str   = "Defender Capital Management — Portfolio Risk Report"
    color_primary:            str   = "#1B2A4A"
    color_secondary:          str   = "#C0C0C0"
    color_accent:             str   = "#4A90D9"
    max_position_warning_pct: float = 0.10
    min_data_points:          int   = 100
    es_confidence_level:      float = 0.975   # primary ES confidence (FRTB standard)
    covariance_mode:          str   = "ledoit_wolf"  # "ledoit_wolf" or "ewma"
    ewma_lambda:              float = 0.94    # EWMA decay factor
    mc_shock_distribution:    str   = "normal"  # "normal" or "student_t"
    mc_df:                    int   = 7       # Student-t degrees of freedom

    @property
    def risk_free_rate_value(self) -> Optional[float]:
        """Return the numeric risk-free rate if manually set, else None (fetch live)."""
        if str(self.risk_free_rate).strip().lower() == "auto":
            return None
        try:
            return float(self.risk_free_rate)
        except (ValueError, TypeError):
            return None

    @property
    def confidence_levels(self) -> list[float]:
        return [self.confidence_level_1, self.confidence_level_2]


@dataclass
class LoadResult:
    """Output of load_portfolio()."""
    holdings:       list[Holding]
    settings:       PortfolioSettings
    warnings:       list[str] = field(default_factory=list)
    errors:         list[str] = field(default_factory=list)
    skipped_rows:   list[int] = field(default_factory=list)
    # MED-1 FIX: Type-coercion failure details (shown as data quality warnings)
    type_warnings:  list[str] = field(default_factory=list)

    @property
    def is_valid(self) -> bool:
        return len(self.holdings) > 0 and len(self.errors) == 0


# ── Fuzzy ticker suggestion ────────────────────────────────────────────────────
def _suggest_ticker(bad_ticker: str) -> Optional[str]:
    """
    Return the closest known ticker using simple edit-distance heuristic.
    Returns None if no close match found.
    """
    bad = bad_ticker.upper()
    best_match = None
    best_score = 999
    for known in _COMMON_TICKERS:
        # Simple: number of character positions that differ
        if abs(len(known) - len(bad)) > 2:
            continue
        padded_bad   = bad.ljust(max(len(bad), len(known)))
        padded_known = known.ljust(max(len(bad), len(known)))
        dist = sum(a != b for a, b in zip(padded_bad, padded_known))
        if dist < best_score:
            best_score = dist
            best_match = known
    return best_match if best_score <= 2 else None


# ── Main loader ────────────────────────────────────────────────────────────────
def load_portfolio(
    excel_path: Optional[Path] = None,
    raise_on_empty: bool = True,
) -> LoadResult:
    """
    Read and validate the portfolio holdings Excel file.

    Parameters
    ----------
    excel_path : Path, optional
        Path to portfolio_holdings.xlsx. Defaults to the canonical location.
    raise_on_empty : bool
        If True, raise ValueError when no valid holdings are found.

    Returns
    -------
    LoadResult
        Validated holdings, settings, and any warnings/errors encountered.

    Raises
    ------
    FileNotFoundError
        If the Excel file doesn't exist.
    ValueError
        If the file structure is unrecognisable or raise_on_empty and empty.
    """
    path = excel_path or get_portfolio_path()
    log.info(f"Loading portfolio from: {path}")

    if not path.exists():
        raise FileNotFoundError(
            f"Portfolio file not found: {path}\n"
            "Please ensure portfolio_holdings.xlsx is in the data/ folder."
        )

    result = LoadResult(holdings=[], settings=PortfolioSettings())

    # ── Read Excel ─────────────────────────────────────────────────────────────
    try:
        xls = pd.ExcelFile(str(path))
    except Exception as exc:
        raise ValueError(
            f"Cannot open portfolio file: {exc}\n"
            "The file may be corrupted or open in Excel. Close Excel and try again."
        ) from exc

    sheet_names = [s.lower() for s in xls.sheet_names]
    if "holdings" not in sheet_names:
        raise ValueError(
            "Sheet 'Holdings' not found in the Excel file.\n"
            "Do not rename sheets. Expected: Holdings, Settings, Instructions."
        )

    # ── Parse Settings ─────────────────────────────────────────────────────────
    if "settings" in sheet_names:
        result.settings = _parse_settings(xls)
    else:
        result.warnings.append("Settings sheet not found — using all defaults.")
        log.warning("Settings sheet missing; using defaults.")

    # ── Parse Holdings ─────────────────────────────────────────────────────────
    try:
        df = xls.parse("Holdings", header=1)
    except Exception as exc:
        raise ValueError(f"Cannot read Holdings sheet: {exc}") from exc

    # Normalise column names: lowercase, strip spaces
    df.columns = [str(c).strip().lower().replace(" ", "_") for c in df.columns]
    log.debug(f"Holdings columns: {list(df.columns)}")

    # Expected column substrings
    col_map = _map_columns(df.columns.tolist())
    if col_map is None:
        raise ValueError(
            "Holdings sheet column structure not recognised.\n"
            "Do not rename column headers."
        )

    for excel_row, raw_row in df.iterrows():
        row_num = int(excel_row) + 3  # offset: 1 header + 1 title row + 0-indexed

        # Skip rows that are clearly empty or the totals row
        if pd.isna(raw_row.get(col_map["ticker"], None)):
            continue
        ticker_raw = str(raw_row[col_map["ticker"]]).strip()
        if not ticker_raw or ticker_raw.lower() in ("ticker", "total portfolio", "nan"):
            continue

        holding, warn, err = _validate_row(ticker_raw, raw_row, col_map, row_num)
        if warn:
            result.warnings.extend(warn)
        if err:
            result.errors.extend(err)
            result.skipped_rows.append(row_num)
            log.warning(f"Row {row_num} skipped: {'; '.join(err)}")
            continue

        result.holdings.append(holding)  # type: ignore[arg-type]

    log.info(
        f"Loaded {len(result.holdings)} valid holdings, "
        f"{len(result.skipped_rows)} skipped, "
        f"{len(result.warnings)} warnings."
    )

    if raise_on_empty and not result.holdings:
        raise ValueError(
            "No valid holdings found in the portfolio file.\n"
            "Please add at least one valid holding and click Save."
        )

    return result


# ── Column mapping ─────────────────────────────────────────────────────────────
def _map_columns(cols: list[str]) -> Optional[dict]:
    """
    Map logical names to actual DataFrame column names using substring matching.
    Returns None if required columns are not found.
    """
    mapping: dict[str, str] = {}

    def find(keyword: str) -> Optional[str]:
        for c in cols:
            if keyword in c:
                return c
        return None

    mapping["ticker"]     = find("ticker")
    mapping["name"]       = find("company") or find("name")
    mapping["sector"]     = find("sector")
    mapping["industry"]   = find("industry")
    mapping["shares"]     = find("shares")
    mapping["cost"]       = find("cost")
    mapping["price"]      = find("current_price") or find("current price") or find("price")
    mapping["mktval"]     = find("market_value") or find("market value")
    mapping["weight"]     = find("weight")
    mapping["pnl_dollar"] = find("unrealized_p") or find("unrealized p")
    mapping["pnl_pct"]    = find("unrealized_p&l_(%)") or find("unrealized_p&l_(%)")

    if not mapping["ticker"] or not mapping["shares"] or not mapping["cost"]:
        log.error(f"Missing required columns. Found: {cols}")
        return None

    return mapping


# ── Row validator ──────────────────────────────────────────────────────────────
def _validate_row(
    ticker_raw: str,
    row: pd.Series,
    col_map: dict,
    row_num: int,
) -> tuple[Optional[Holding], list[str], list[str]]:
    """
    Validate a single data row and return (Holding or None, warnings, errors).
    """
    warnings: list[str] = []
    errors:   list[str] = []

    # 1. Ticker
    ticker = re.sub(r"[^A-Za-z0-9.\-]", "", ticker_raw).upper()
    if not ticker:
        errors.append(f"Row {row_num}: Empty ticker — row skipped.")
        return None, warnings, errors
    if len(ticker) > 6:
        suggestion = _suggest_ticker(ticker)
        hint = f"  Did you mean '{suggestion}'?" if suggestion else ""
        errors.append(
            f"Row {row_num}: Ticker '{ticker_raw}' looks invalid (> 6 characters).{hint}"
        )
        return None, warnings, errors
    if ticker != ticker_raw.strip().upper():
        warnings.append(
            f"Row {row_num}: Ticker normalised '{ticker_raw}' → '{ticker}'."
        )
    # Fuzzy match warning for common mis-spellings
    suggestion = _suggest_ticker(ticker)
    if suggestion and suggestion != ticker:
        warnings.append(
            f"Row {row_num}: Ticker '{ticker}' unrecognised in common list. "
            f"Did you mean '{suggestion}'? (Will be validated by data fetcher.)"
        )

    # 2. Shares Held
    shares_raw = row.get(col_map["shares"])
    shares = _parse_positive_int(shares_raw)
    if shares is None:
        # MED-1 FIX: Clearer error with type hint
        _type_hint = ""
        if isinstance(shares_raw, str) and any(c.isalpha() for c in str(shares_raw)):
            _type_hint = (
                f" (found text '{shares_raw}' — enter a number like 100, not '100 shares')"
            )
        errors.append(
            f"Row {row_num} ({ticker}): Shares Held must be a positive whole number, "
            f"got '{shares_raw}'{_type_hint}. Row skipped."
        )
        return None, warnings, errors

    # 3. Cost Basis
    cost_raw = row.get(col_map["cost"])
    cost = _parse_positive_float(cost_raw)
    if cost is None:
        errors.append(
            f"Row {row_num} ({ticker}): Cost Basis must be a positive number, "
            f"got '{cost_raw}'. Row skipped."
        )
        return None, warnings, errors

    # 4. Optional pre-populated fields (fallback to empty string / 0)
    name     = _safe_str(row.get(col_map.get("name",     ""), ""))
    sector   = _safe_str(row.get(col_map.get("sector",   ""), ""))
    industry = _safe_str(row.get(col_map.get("industry", ""), ""))
    price    = _parse_nonneg_float(row.get(col_map.get("price",  ""), 0.0)) or 0.0

    holding = Holding(
        ticker=ticker,
        company_name=name,
        sector=sector,
        industry=industry,
        shares_held=shares,
        cost_basis=cost,
        current_price=price,
        row_number=row_num,
    )
    return holding, warnings, errors


# ── Settings parser ────────────────────────────────────────────────────────────
def _parse_settings(xls: pd.ExcelFile) -> PortfolioSettings:
    """Read the Settings sheet and return a PortfolioSettings dataclass."""
    try:
        df = xls.parse("Settings", header=1)
    except Exception as exc:
        log.warning(f"Could not read Settings sheet: {exc}. Using defaults.")
        return PortfolioSettings()

    settings = PortfolioSettings()
    # Find Parameter column and Value column
    param_col = None
    value_col = None
    for c in df.columns:
        cs = str(c).strip().lower()
        if "parameter" in cs:
            param_col = c
        if "value" in cs and "default" not in cs:
            value_col = c

    if param_col is None or value_col is None:
        log.warning("Settings sheet structure unexpected. Using defaults.")
        return settings

    for _, row in df.iterrows():
        param = _safe_str(row.get(param_col, "")).lower().strip()
        val   = row.get(value_col)
        if not param or pd.isna(val):
            continue

        try:
            if param == "benchmark_ticker":
                settings.benchmark_ticker = str(val).strip().upper()
            elif param == "risk_free_rate":
                settings.risk_free_rate = str(val).strip()
            elif param == "confidence_level_1":
                settings.confidence_level_1 = float(val)
            elif param == "confidence_level_2":
                settings.confidence_level_2 = float(val)
            elif param == "lookback_years":
                settings.lookback_years = int(float(val))
            elif param == "simulation_paths":
                settings.simulation_paths = int(float(val))
            elif param == "simulation_days":
                settings.simulation_days = int(float(val))
            elif param == "cache_expiry_hours":
                pass  # cache_expiry_hours is no longer used; ignore for backwards compat
            elif param == "stress_custom_drawdown" or param == "stress_test_custom_drawdown":
                settings.stress_custom_drawdown = float(val)
            elif param == "report_title":
                settings.report_title = str(val).strip()
            elif param == "color_primary":
                settings.color_primary = str(val).strip()
            elif param == "color_secondary":
                settings.color_secondary = str(val).strip()
            elif param == "color_accent":
                settings.color_accent = str(val).strip()
            elif param == "max_position_warning_pct":
                settings.max_position_warning_pct = float(val)
            elif param == "min_data_points":
                settings.min_data_points = int(float(val))
            elif param == "portfolio_name":
                settings.portfolio_name = str(val).strip()
            elif param == "portfolio_short_name":
                settings.portfolio_short_name = str(val).strip()
            elif param == "es_confidence_level":
                settings.es_confidence_level = float(val)
            elif param == "covariance_mode":
                settings.covariance_mode = str(val).strip().lower()
            elif param == "ewma_lambda":
                # MED-4 FIX: Validate EWMA lambda is in acceptable range
                lam = float(val)
                if not (0.85 <= lam <= 0.99):
                    log.warning(
                        f"Settings: ewma_lambda={lam:.3f} is outside recommended range "
                        f"[0.85, 0.99]. Using 0.94 (RiskMetrics standard)."
                    )
                    lam = 0.94
                settings.ewma_lambda = lam
            elif param == "mc_shock_distribution":
                settings.mc_shock_distribution = str(val).strip().lower()
            elif param == "mc_df":
                settings.mc_df = int(float(val))
        except (ValueError, TypeError) as exc:
            log.warning(f"Settings: could not parse '{param}' = '{val}': {exc}")

    log.info(
        f"Settings loaded: benchmark={settings.benchmark_ticker}, "
        f"lookback={settings.lookback_years}y, paths={settings.simulation_paths}"
    )
    return settings


def save_settings(settings: PortfolioSettings, excel_path: Optional[Path] = None) -> None:
    """Write PortfolioSettings back to the Settings sheet of the Excel file."""
    import openpyxl

    path = excel_path or get_portfolio_path()
    try:
        wb = openpyxl.load_workbook(str(path))
    except PermissionError:
        raise PermissionError(
            f"Cannot write to {path.name} — close the file in Excel first."
        )

    if "Settings" not in wb.sheetnames:
        log.warning("Settings sheet not found; cannot save settings.")
        wb.close()
        return

    ws = wb["Settings"]

    field_map = {
        "portfolio_name":           settings.portfolio_name,
        "portfolio_short_name":     settings.portfolio_short_name,
        "benchmark_ticker":         settings.benchmark_ticker,
        "risk_free_rate":           settings.risk_free_rate,
        "confidence_level_1":       settings.confidence_level_1,
        "confidence_level_2":       settings.confidence_level_2,
        "lookback_years":           settings.lookback_years,
        "simulation_paths":         settings.simulation_paths,
        "simulation_days":          settings.simulation_days,
        "stress_custom_drawdown":   settings.stress_custom_drawdown,
        "report_title":             settings.report_title,
        "color_primary":            settings.color_primary,
        "color_secondary":          settings.color_secondary,
        "color_accent":             settings.color_accent,
        "max_position_warning_pct": settings.max_position_warning_pct,
        "min_data_points":          settings.min_data_points,
        "es_confidence_level":      settings.es_confidence_level,
        "covariance_mode":          settings.covariance_mode,
        "ewma_lambda":              settings.ewma_lambda,
        "mc_shock_distribution":    settings.mc_shock_distribution,
        "mc_df":                    settings.mc_df,
    }

    # Find parameter and value columns (header row = row 2)
    param_col = None
    value_col = None
    for col_idx in range(1, ws.max_column + 1):
        header = str(ws.cell(row=2, column=col_idx).value or "").strip().lower()
        if "parameter" in header:
            param_col = col_idx
        if "value" in header and "default" not in header:
            value_col = col_idx

    if param_col is None or value_col is None:
        log.warning("Settings sheet structure unexpected; cannot save.")
        wb.close()
        return

    # Update existing rows
    updated: set[str] = set()
    for row_idx in range(3, ws.max_row + 1):
        param = str(ws.cell(row=row_idx, column=param_col).value or "").strip().lower()
        if param in field_map:
            ws.cell(row=row_idx, column=value_col).value = field_map[param]
            updated.add(param)

    # Append new rows for settings not yet in the file
    for param, val in field_map.items():
        if param not in updated:
            next_row = ws.max_row + 1
            ws.cell(row=next_row, column=param_col).value = param
            ws.cell(row=next_row, column=value_col).value = val

    wb.save(str(path))
    log.info(f"Settings saved to {path}")


# ── Type helpers ───────────────────────────────────────────────────────────────
def _parse_positive_int(value) -> Optional[int]:
    """Return positive int or None."""
    if pd.isna(value):
        return None
    try:
        v = int(float(str(value).replace(",", "")))
        return v if v > 0 else None
    except (ValueError, TypeError):
        return None


def _parse_positive_float(value) -> Optional[float]:
    """Return positive float or None."""
    if pd.isna(value):
        return None
    try:
        v = float(str(value).replace(",", "").replace("$", ""))
        return v if v > 0 else None
    except (ValueError, TypeError):
        return None


def _parse_nonneg_float(value, default: float = 0.0) -> float:
    """Return non-negative float, or default."""
    if pd.isna(value):
        return default
    try:
        v = float(str(value).replace(",", "").replace("$", ""))
        return v if v >= 0 else default
    except (ValueError, TypeError):
        return default


def _safe_str(value, default: str = "") -> str:
    """Convert to string, returning default for NaN/None."""
    if pd.isna(value):
        return default
    s = str(value).strip()
    return s if s.lower() != "nan" else default


# ── Quick self-test ─────────────────────────────────────────────────────────────
if __name__ == "__main__":
    result = load_portfolio()
    print(f"\n✅ Loaded {len(result.holdings)} holdings")
    for w in result.warnings:
        print(f"  ⚠️  {w}")
    for e in result.errors:
        print(f"  ❌  {e}")
    print(f"\nSettings: {result.settings}")
    print("\nFirst 5 holdings:")
    for h in result.holdings[:5]:
        print(f"  {h.ticker:6s}  {h.shares_held:5d} shares @ ${h.cost_basis:.2f}")
