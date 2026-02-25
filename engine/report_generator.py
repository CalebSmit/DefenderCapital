"""
engine/report_generator.py — HTML report builder and Excel result writer.

Generates:
  1. A standalone HTML report with embedded charts and professional styling.
  2. Updated Excel sheets: Risk Summary, Stock Risk Detail, Correlation Matrix,
     Stress Test Results, Monte Carlo Summary, Last Updated.
"""
from __future__ import annotations

import base64
import io
from datetime import datetime
from pathlib import Path
from typing import Optional

import matplotlib
matplotlib.use("Agg")
import matplotlib.pyplot as plt
import matplotlib.ticker as mticker
import numpy as np
import pandas as pd
from matplotlib.figure import Figure

from engine.market_data import MarketData
from engine.monte_carlo import SimulationResult
from engine.risk_metrics import PortfolioRiskMetrics, build_drawdown_series, build_rolling_vol_df
from engine.stress_testing import StressTestResults
from engine.utils import (
    get_logger, timestamp_str, EXPORTS_DIR, fmt_currency, fmt_pct, fmt_number,
)

log = get_logger("defender.report")

# ── Brand colours ──────────────────────────────────────────────────────────────
NAVY   = "#1B2A4A"
SILVER = "#C0C0C0"
BLUE   = "#4A90D9"
WHITE  = "#FFFFFF"
GREEN  = "#27AE60"
RED    = "#E74C3C"
GOLD   = "#F39C12"

# ── S&P 500 sector data (shared with dashboard) ──────────────────────────────
# Fallback weights if live ETF market-cap fetch fails. Keys use yfinance sector names.
_SP500_SECTOR_WEIGHTS_FALLBACK = {
    "Technology": 0.3147,
    "Financial Services": 0.1280,
    "Communication Services": 0.1114,
    "Consumer Cyclical": 0.1049,
    "Healthcare": 0.0959,
    "Industrials": 0.0882,
    "Consumer Defensive": 0.0611,
    "Energy": 0.0332,
    "Utilities": 0.0235,
    "Basic Materials": 0.0206,
    "Real Estate": 0.0189,
}


def _init_sector_weights() -> dict[str, float]:
    """Fetch live S&P 500 sector weights from ETF market caps, fall back to hardcoded."""
    try:
        from engine.market_data import fetch_sp500_sector_weights
        live = fetch_sp500_sector_weights()
        if live:
            return live
    except Exception:
        pass
    return _SP500_SECTOR_WEIGHTS_FALLBACK


SP500_SECTOR_WEIGHTS = _init_sector_weights()

# Sector SPDR ETF tickers — one per GICS sector
SECTOR_ETF_MAP = {
    "Technology": "XLK",
    "Financial Services": "XLF",
    "Healthcare": "XLV",
    "Consumer Cyclical": "XLY",
    "Communication Services": "XLC",
    "Industrials": "XLI",
    "Consumer Defensive": "XLP",
    "Energy": "XLE",
    "Utilities": "XLU",
    "Real Estate": "XLRE",
    "Basic Materials": "XLB",
}


# ═══════════════════════════════════════════════════════════════════════════════
# Chart generation helpers
# ═══════════════════════════════════════════════════════════════════════════════

def _fig_to_base64(fig: Figure) -> str:
    """Encode a matplotlib Figure as a base64 PNG string for HTML embedding."""
    buf = io.BytesIO()
    fig.savefig(buf, format="png", bbox_inches="tight", dpi=120, facecolor=WHITE)
    plt.close(fig)
    buf.seek(0)
    return base64.b64encode(buf.read()).decode("utf-8")


def _brand_axes(ax, title: str = "", xlabel: str = "", ylabel: str = "") -> None:
    """Apply consistent brand styling to a matplotlib axes."""
    ax.set_facecolor(WHITE)
    ax.spines["top"].set_visible(False)
    ax.spines["right"].set_visible(False)
    ax.spines["left"].set_color(SILVER)
    ax.spines["bottom"].set_color(SILVER)
    ax.tick_params(colors="#555555", labelsize=9)
    if title:
        ax.set_title(title, fontsize=12, fontweight="bold", color=NAVY, pad=10)
    if xlabel:
        ax.set_xlabel(xlabel, fontsize=9, color="#555555")
    if ylabel:
        ax.set_ylabel(ylabel, fontsize=9, color="#555555")


def _chart_sector_allocation(metrics: PortfolioRiskMetrics, md: MarketData) -> str:
    """Sector allocation donut chart."""
    sector_weights: dict[str, float] = {}
    for h in md.holdings:
        sector_weights[h.sector] = sector_weights.get(h.sector, 0.0) + h.weight

    labels = list(sector_weights.keys())
    sizes  = list(sector_weights.values())
    colors = [
        "#1B2A4A", "#4A90D9", "#27AE60", "#E74C3C", "#F39C12",
        "#9B59B6", "#1ABC9C", "#E67E22", "#2ECC71", "#3498DB",
        "#95A5A6", "#D35400",
    ][:len(labels)]

    fig, ax = plt.subplots(figsize=(7, 5), facecolor=WHITE)
    wedges, texts, autotexts = ax.pie(
        sizes, labels=None, colors=colors,
        autopct=lambda p: f"{p:.1f}%" if p > 3 else "",
        startangle=90, pctdistance=0.85,
        wedgeprops={"linewidth": 2, "edgecolor": WHITE},
    )
    for at in autotexts:
        at.set_fontsize(8)
        at.set_color(WHITE)
    centre_circle = plt.Circle((0, 0), 0.65, fc=WHITE)
    ax.add_artist(centre_circle)
    ax.legend(
        wedges, [f"{l} ({s:.1%})" for l, s in zip(labels, sizes)],
        loc="center left", bbox_to_anchor=(1.0, 0.5), fontsize=8,
        framealpha=0,
    )
    ax.set_title("Sector Allocation", fontsize=12, fontweight="bold", color=NAVY)
    fig.tight_layout()
    return _fig_to_base64(fig)


def _chart_var_comparison(metrics: PortfolioRiskMetrics, sim: SimulationResult) -> str:
    """VaR comparison bar chart — includes Cornish-Fisher adjusted VaR."""
    labels     = ["Parametric\nVaR 95%", "Cornish-Fisher\nVaR 95%", "Historical\nVaR 95%", "MC\nVaR 95%",
                  "Parametric\nVaR 99%", "Cornish-Fisher\nVaR 99%", "Historical\nVaR 99%", "MC\nVaR 99%"]
    values_raw = [
        abs(metrics.var_95.parametric_var),
        abs(metrics.var_95.cornish_fisher_var),
        abs(metrics.var_95.historical_var),
        abs(sim.var_95),
        abs(metrics.var_99.parametric_var),
        abs(metrics.var_99.cornish_fisher_var),
        abs(metrics.var_99.historical_var),
        abs(sim.var_99),
    ]

    fig, ax = plt.subplots(figsize=(10, 4), facecolor=WHITE)
    bar_colors = [BLUE, GREEN, NAVY, GOLD, "#E74C3C", "#27AE60", "#C0392B", "#922B21"]
    bars = ax.bar(labels, values_raw, color=bar_colors, width=0.6, edgecolor=WHITE, linewidth=1.5)
    for bar, val in zip(bars, values_raw):
        ax.text(bar.get_x() + bar.get_width() / 2, bar.get_height() + 200,
                f"${val:,.0f}", ha="center", va="bottom", fontsize=7, color=NAVY)
    _brand_axes(ax, "Value at Risk Comparison (Daily, $)", ylabel="Dollar Loss ($)")
    ax.yaxis.set_major_formatter(mticker.FuncFormatter(lambda x, _: f"${x:,.0f}"))
    fig.tight_layout()
    return _fig_to_base64(fig)


def _chart_monte_carlo_fan(sim: SimulationResult) -> str:
    """Monte Carlo fan chart."""
    fig, ax = plt.subplots(figsize=(9, 5), facecolor=WHITE)
    days = np.arange(1, sim.simulation_days + 1)
    pp   = sim.percentile_paths

    ax.fill_between(days, pp["p05"], pp["p95"], alpha=0.15, color=BLUE, label="5th–95th percentile")
    ax.fill_between(days, pp["p25"], pp["p75"], alpha=0.30, color=BLUE, label="25th–75th percentile")
    ax.plot(days, pp["p50"], color=NAVY, linewidth=2, label="Median")
    ax.axhline(sim.initial_value, color=RED, linestyle="--", linewidth=1, alpha=0.7, label="Starting Value")

    _brand_axes(ax, "Monte Carlo Simulation — Portfolio Value Fan Chart",
                xlabel="Trading Days", ylabel="Portfolio Value ($)")
    ax.yaxis.set_major_formatter(mticker.FuncFormatter(lambda x, _: f"${x:,.0f}"))
    ax.legend(fontsize=9, framealpha=0.8)
    fig.tight_layout()
    return _fig_to_base64(fig)


def _chart_drawdown(md: MarketData) -> str:
    """Drawdown chart: portfolio vs benchmark."""
    try:
        dd_df = build_drawdown_series(md)
        fig, ax = plt.subplots(figsize=(9, 4), facecolor=WHITE)
        ax.fill_between(dd_df.index, dd_df["Portfolio"] * 100, 0, alpha=0.4, color=NAVY, label="Portfolio")
        ax.plot(dd_df.index, dd_df.iloc[:, 1] * 100, color=BLUE, linewidth=1, alpha=0.7,
                label=md.benchmark_ticker)
        ax.plot(dd_df.index, dd_df["Portfolio"] * 100, color=NAVY, linewidth=1.5)
        _brand_axes(ax, "Portfolio Drawdown Over Time", ylabel="Drawdown (%)")
        ax.yaxis.set_major_formatter(mticker.FuncFormatter(lambda x, _: f"{x:.0f}%"))
        ax.legend(fontsize=9)
        fig.tight_layout()
        return _fig_to_base64(fig)
    except Exception as e:
        log.warning(f"Drawdown chart failed: {e}")
        return ""


def _chart_rolling_vol(md: MarketData) -> str:
    """Rolling 30-day volatility chart."""
    try:
        rv = build_rolling_vol_df(md, window=30).dropna()
        if rv.empty:
            return ""
        fig, ax = plt.subplots(figsize=(9, 4), facecolor=WHITE)
        ax.plot(rv.index, rv.iloc[:, 0] * 100, color=NAVY, linewidth=1.5, label="Portfolio")
        ax.plot(rv.index, rv.iloc[:, 1] * 100, color=BLUE, linewidth=1, alpha=0.7, label=md.benchmark_ticker)
        _brand_axes(ax, "Rolling 30-Day Annualised Volatility", ylabel="Volatility (% p.a.)")
        ax.yaxis.set_major_formatter(mticker.FuncFormatter(lambda x, _: f"{x:.0f}%"))
        ax.legend(fontsize=9)
        fig.tight_layout()
        return _fig_to_base64(fig)
    except Exception as e:
        log.warning(f"Rolling vol chart failed: {e}")
        return ""


def _chart_risk_contribution(metrics: PortfolioRiskMetrics) -> str:
    """Risk contribution horizontal bar chart — top 15 positions."""
    sm = sorted(metrics.stock_metrics, key=lambda x: abs(x.component_var_95), reverse=True)[:15]
    labels  = [f"{s.ticker}" for s in sm]
    values  = [abs(s.component_var_95) for s in sm]
    colors  = [RED if v > np.mean(values) else BLUE for v in values]

    fig, ax = plt.subplots(figsize=(8, 6), facecolor=WHITE)
    bars = ax.barh(labels[::-1], values[::-1], color=colors[::-1], edgecolor=WHITE, linewidth=0.5)
    for bar, val in zip(bars, values[::-1]):
        ax.text(val + 50, bar.get_y() + bar.get_height() / 2,
                f"${val:,.0f}", va="center", fontsize=8, color=NAVY)
    _brand_axes(ax, "Risk Contribution by Holding (Component VaR 95%)", xlabel="Component VaR ($)")
    ax.xaxis.set_major_formatter(mticker.FuncFormatter(lambda x, _: f"${x:,.0f}"))
    fig.tight_layout()
    return _fig_to_base64(fig)


def _chart_stress_scenarios(stress: StressTestResults) -> str:
    """Stress test comparison bar chart."""
    scenarios = [s for s in stress.all_scenarios]
    names  = [s.name[:30] for s in scenarios]
    losses = [abs(s.portfolio_loss_pct) * 100 for s in scenarios]

    fig, ax = plt.subplots(figsize=(10, 5), facecolor=WHITE)
    colors = [RED if l > 20 else GOLD if l > 10 else BLUE for l in losses]
    bars = ax.bar(names, losses, color=colors, edgecolor=WHITE, linewidth=1)
    for bar, val in zip(bars, losses):
        ax.text(bar.get_x() + bar.get_width() / 2, bar.get_height() + 0.3,
                f"{val:.1f}%", ha="center", fontsize=8, color=NAVY)
    _brand_axes(ax, "Estimated Portfolio Loss by Scenario (%)", ylabel="Estimated Loss (%)")
    plt.xticks(rotation=30, ha="right", fontsize=8)
    fig.tight_layout()
    return _fig_to_base64(fig)


# ═══════════════════════════════════════════════════════════════════════════════
# Sector analysis helpers
# ═══════════════════════════════════════════════════════════════════════════════

def _period_return(prices: pd.Series, months: int) -> float | None:
    """Simple return over the last N months of a price series."""
    if prices is None or len(prices) < 2:
        return None
    end_price = float(prices.iloc[-1])
    cutoff = prices.index[-1] - pd.DateOffset(months=months)
    past = prices.loc[:cutoff]
    if len(past) == 0:
        return None
    start_price = float(past.iloc[-1])
    if start_price <= 0:
        return None
    return (end_price / start_price) - 1.0


def _sector_portfolio_return(
    prices: pd.DataFrame, holdings: list, sector: str, months: int,
) -> float | None:
    """Weight-weighted return for portfolio holdings in a given sector."""
    in_sector = [h for h in holdings if h.sector == sector and h.ticker in prices.columns]
    if not in_sector:
        return None
    total_mv = sum(h.market_value for h in in_sector)
    if total_mv <= 0:
        return None
    weighted = 0.0
    for h in in_sector:
        ret = _period_return(prices[h.ticker].dropna(), months)
        if ret is None:
            continue
        weighted += (h.market_value / total_mv) * ret
    return weighted


def _fetch_sector_etf_prices() -> pd.DataFrame:
    """Fetch 13 months of price history for all sector ETFs."""
    import yfinance as yf
    etf_tickers = list(SECTOR_ETF_MAP.values())
    end = pd.Timestamp.now().strftime("%Y-%m-%d")
    start = (pd.Timestamp.now() - pd.DateOffset(months=13)).strftime("%Y-%m-%d")
    try:
        raw = yf.download(etf_tickers, start=start, end=end, auto_adjust=True,
                          progress=False, actions=False, group_by="ticker", threads=True)
        frames = {}
        if isinstance(raw.columns, pd.MultiIndex):
            for t in etf_tickers:
                try:
                    s = raw[t]["Close"].dropna()
                    if len(s) > 5:
                        frames[t] = s
                except Exception:
                    pass
        elif "Close" in raw.columns and len(etf_tickers) == 1:
            frames[etf_tickers[0]] = raw["Close"].dropna()
        return pd.DataFrame(frames)
    except Exception as exc:
        log.warning(f"Could not fetch sector ETF prices: {exc}")
        return pd.DataFrame()


# ═══════════════════════════════════════════════════════════════════════════════
# HTML Report Generator
# ═══════════════════════════════════════════════════════════════════════════════

def generate_html_report(
    md:        MarketData,
    metrics:   PortfolioRiskMetrics,
    sim:       SimulationResult,
    stress:    StressTestResults,
    settings_title: str = "Portfolio Risk Report",
    portfolio_name: str = "Portfolio",
    portfolio_short_name: str = "Portfolio",
) -> Path:
    """
    Generate a standalone HTML risk report.

    The report is a single file with all CSS inline and all charts
    base64-encoded — it can be opened in any browser, emailed, or printed.

    Returns
    -------
    Path
        Path to the saved HTML file.
    """
    log.info("Generating HTML report…")
    now = datetime.now().strftime("%B %d, %Y at %I:%M %p")

    # Generate charts
    chart_sector    = _chart_sector_allocation(metrics, md)
    chart_var       = _chart_var_comparison(metrics, sim)
    chart_mc        = _chart_monte_carlo_fan(sim)
    chart_dd        = _chart_drawdown(md)
    chart_rv        = _chart_rolling_vol(md)
    chart_risk_ctb  = _chart_risk_contribution(metrics)
    chart_stress    = _chart_stress_scenarios(stress)

    def img_tag(b64: str, alt: str = "") -> str:
        if not b64:
            return f'<p style="color:#888">Chart unavailable</p>'
        return f'<img src="data:image/png;base64,{b64}" alt="{alt}" style="max-width:100%;border-radius:6px;">'

    # ── Executive summary ──────────────────────────────────────────────────────
    sharpe_interp = (
        "poor — below risk-free rate" if metrics.sharpe < 0 else
        "mediocre — below 0.5" if metrics.sharpe < 0.5 else
        "acceptable (0.5–1.0)" if metrics.sharpe < 1.0 else
        "good (1.0–2.0)" if metrics.sharpe < 2.0 else "excellent (> 2.0)"
    )
    top_sector = max(
        {h.sector: 0.0 for h in md.holdings}.items(),
        key=lambda kv: sum(h.weight for h in md.holdings if h.sector == kv[0])
    )
    top_sector_wt = sum(h.weight for h in md.holdings if h.sector == top_sector[0])

    exec_summary = f"""
    As of {now}, the {portfolio_name} portfolio is valued at
    <strong>{fmt_currency(md.total_portfolio_value)}</strong> across {len(md.holdings)} equity positions.
    <br><br>
    The 95% daily Value at Risk (VaR) is <strong>{fmt_currency(abs(metrics.var_95.parametric_var))}</strong>,
    meaning there is a 5% probability of losing more than this amount on any given trading day.
    In the worst 5% of days, the expected loss (CVaR) is
    <strong>{fmt_currency(abs(metrics.var_95.parametric_cvar))}</strong>.
    <br><br>
    The portfolio's annualised volatility is <strong>{metrics.annualized_vol:.1%}</strong>
    with a Sharpe ratio of <strong>{metrics.sharpe:.2f}</strong>, which is {sharpe_interp}.
    Portfolio beta versus {md.benchmark_ticker} is <strong>{metrics.beta:.2f}</strong>.
    The maximum drawdown over the lookback period was <strong>{metrics.max_drawdown:.1%}</strong>.
    <br><br>
    The largest sector concentration is <strong>{top_sector[0]}</strong>
    ({top_sector_wt:.1%} of portfolio). The effective number of independent positions is
    <strong>{metrics.eff_num_bets:.1f}</strong> (vs. {len(md.holdings)} actual holdings),
    reflecting the impact of correlations on true diversification.
    <br><br>
    In a severe stress scenario (2008 GFC), the portfolio is estimated to lose approximately
    <strong>{fmt_pct(stress.historical[0].portfolio_loss_pct)}</strong>
    ({fmt_currency(abs(stress.historical[0].portfolio_loss_usd))}).
    """

    # ── Holdings table ─────────────────────────────────────────────────────────
    holdings_rows = ""
    for h in sorted(md.holdings, key=lambda x: -x.market_value):
        pnl_color = GREEN if h.unrealized_pnl >= 0 else RED
        holdings_rows += f"""
        <tr>
            <td style="font-weight:600;color:{NAVY}">{h.ticker}</td>
            <td>{h.company_name}</td>
            <td>{h.sector}</td>
            <td style="text-align:right">{h.shares_held:,}</td>
            <td style="text-align:right">{fmt_currency(h.cost_basis)}</td>
            <td style="text-align:right">{fmt_currency(h.current_price)}</td>
            <td style="text-align:right;font-weight:600">{fmt_currency(h.market_value)}</td>
            <td style="text-align:right">{h.weight:.2%}</td>
            <td style="text-align:right;color:{pnl_color}">{fmt_currency(h.unrealized_pnl)}</td>
            <td style="text-align:right;color:{pnl_color}">{h.unrealized_pct:.1%}</td>
        </tr>"""

    # ── Stock risk table ───────────────────────────────────────────────────────
    risk_rows = ""
    for sm in sorted(metrics.stock_metrics, key=lambda x: -abs(x.component_var_95)):
        b_color = RED if abs(sm.beta) > 1.5 else (GREEN if abs(sm.beta) < 0.8 else NAVY)
        risk_rows += f"""
        <tr>
            <td style="font-weight:600;color:{NAVY}">{sm.ticker}</td>
            <td>{sm.sector}</td>
            <td style="text-align:right">{sm.annualized_vol:.1%}</td>
            <td style="text-align:right;color:{b_color}">{sm.beta:.2f}</td>
            <td style="text-align:right">{sm.sharpe:.2f}</td>
            <td style="text-align:right">{sm.sortino:.2f}</td>
            <td style="text-align:right;color:{RED}">{sm.max_drawdown:.1%}</td>
            <td style="text-align:right">{sm.annualized_return:.1%}</td>
            <td style="text-align:right">{sm.risk_contribution_pct:.1f}%</td>
        </tr>"""

    # ── Stress test tables ─────────────────────────────────────────────────────
    stress_rows = ""
    for s in stress.all_scenarios:
        color = RED if abs(s.portfolio_loss_pct) > 0.20 else GOLD if abs(s.portfolio_loss_pct) > 0.10 else NAVY
        stress_rows += f"""
        <tr>
            <td style="font-weight:600">{s.name}</td>
            <td>{s.period}</td>
            <td style="text-align:right;color:{color};font-weight:600">{s.portfolio_loss_pct:.1%}</td>
            <td style="text-align:right;color:{color};font-weight:600">{fmt_currency(s.portfolio_loss_usd)}</td>
            <td style="text-align:right">{s.benchmark_loss_pct:.1%}</td>
        </tr>"""

    # ── Sector analysis ───────────────────────────────────────────────────────
    pv = md.total_portfolio_value
    port_sector_wt: dict[str, float] = {}
    for h in md.holdings:
        port_sector_wt[h.sector] = port_sector_wt.get(h.sector, 0.0) + h.weight

    # Weight comparison rows
    wt_rows_data = []
    for sector, sp_wt in SP500_SECTOR_WEIGHTS.items():
        p_wt = port_sector_wt.get(sector, 0.0)
        diff = p_wt - sp_wt
        target = (sp_wt - p_wt) * pv
        wt_rows_data.append((sector, sp_wt, p_wt, diff, target))
    ow = sorted([r for r in wt_rows_data if r[3] >= 0], key=lambda r: -r[3])
    uw = sorted([r for r in wt_rows_data if r[3] < 0], key=lambda r: r[3])

    sector_wt_rows = ""
    for sector, sp_wt, p_wt, diff, target in ow + uw:
        d_color = GREEN if diff >= 0 else RED
        d_str = f"+{diff:.2%}" if diff >= 0 else f"{diff:.2%}"
        t_str = f"${target:,.2f}" if target >= 0 else f"(${abs(target):,.2f})"
        sector_wt_rows += (
            f'<tr><td style="font-weight:600">{sector}</td>'
            f'<td style="text-align:right">{sp_wt:.2%}</td>'
            f'<td style="text-align:right">{p_wt:.2%}</td>'
            f'<td style="text-align:right;color:{d_color};font-weight:600">{d_str}</td>'
            f'<td style="text-align:right;color:{d_color}">{t_str}</td></tr>'
        )
    tot_sp = sum(r[1] for r in wt_rows_data)
    tot_pt = sum(r[2] for r in wt_rows_data)
    sector_wt_rows += (
        f'<tr style="font-weight:700;border-top:2px solid {NAVY}">'
        f'<td>Totals</td><td style="text-align:right">{tot_sp:.2%}</td>'
        f'<td style="text-align:right">{tot_pt:.2%}</td>'
        f'<td style="text-align:right"></td><td style="text-align:right"></td></tr>'
    )

    # Sector returns comparison
    etf_prices = _fetch_sector_etf_prices()
    sector_ret_rows = ""
    for sector, etf_ticker in SECTOR_ETF_MAP.items():
        if etf_ticker not in etf_prices.columns:
            continue
        etf_s = etf_prices[etf_ticker].dropna()
        r6 = _period_return(etf_s, 6)
        r12 = _period_return(etf_s, 12)
        dcm12 = _sector_portfolio_return(md.prices, md.holdings, sector, 12)
        diff_r = (dcm12 - r12) if (dcm12 is not None and r12 is not None) else None

        def _fc(v):
            if v is None:
                return '<span style="color:#999">N/A</span>'
            c = GREEN if v >= 0 else RED
            return f'<span style="color:{c}">{v:.1%}</span>'

        sector_ret_rows += (
            f'<tr><td style="font-weight:600">{etf_ticker}</td><td>{sector}</td>'
            f'<td style="text-align:right">{_fc(r6)}</td>'
            f'<td style="text-align:right">{_fc(r12)}</td>'
            f'<td style="text-align:right">{_fc(dcm12)}</td>'
            f'<td style="text-align:right">{_fc(diff_r)}</td></tr>'
        )

    # Per-sector holdings tables
    from collections import defaultdict
    sector_groups: dict[str, list] = defaultdict(list)
    for h in md.holdings:
        sector_groups[h.sector].append(h)
    sector_order = sorted(sector_groups.keys(),
                          key=lambda s: sum(h.market_value for h in sector_groups[s]),
                          reverse=True)

    sector_holdings_html = ""
    for sector in sector_order:
        holds = sorted(sector_groups[sector], key=lambda h: -h.market_value)
        s_mv = sum(h.market_value for h in holds)
        s_wt = s_mv / pv if pv > 0 else 0.0
        rows_h = ""
        for h in holds:
            ps = h.market_value / s_mv if s_mv > 0 else 0.0
            rows_h += (
                f'<tr><td style="font-weight:600;color:{NAVY}">{h.ticker}</td>'
                f'<td>{h.company_name}</td><td style="text-align:right">{h.shares_held:,}</td>'
                f'<td style="text-align:right">{fmt_currency(h.current_price)}</td>'
                f'<td style="text-align:right;font-weight:600">{fmt_currency(h.market_value)}</td>'
                f'<td style="text-align:right">{ps:.1%}</td>'
                f'<td style="text-align:right">{h.weight:.1%}</td></tr>'
            )
        rows_h += (
            f'<tr style="font-weight:700;border-top:2px solid {NAVY}">'
            f'<td>{sector} Total</td><td></td><td></td><td></td>'
            f'<td style="text-align:right">{fmt_currency(s_mv)}</td>'
            f'<td style="text-align:right">100.0%</td>'
            f'<td style="text-align:right">{s_wt:.1%}</td></tr>'
        )
        sector_holdings_html += f"""
        <h3>{sector} — {len(holds)} holding{'s' if len(holds) != 1 else ''}</h3>
        <table><thead><tr>
            <th>Ticker</th><th>Company</th><th style="text-align:right">Shares</th>
            <th style="text-align:right">Price</th><th style="text-align:right">Market Value</th>
            <th style="text-align:right">% of Sector</th><th style="text-align:right">% of Portfolio</th>
        </tr></thead><tbody>{rows_h}</tbody></table>
        """

    # ── Full HTML ──────────────────────────────────────────────────────────────
    html = f"""<!DOCTYPE html>
<html lang="en">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width, initial-scale=1">
<title>{settings_title}</title>
<style>
  * {{ box-sizing: border-box; margin: 0; padding: 0; }}
  body {{ font-family: 'Arial', sans-serif; background: #f5f7fa; color: #333; font-size: 13px; }}
  .header {{ background: {NAVY}; color: white; padding: 32px 48px; }}
  .header h1 {{ font-size: 24px; font-weight: 700; letter-spacing: 0.5px; }}
  .header p  {{ font-size: 13px; opacity: 0.8; margin-top: 6px; }}
  .container {{ max-width: 1100px; margin: 0 auto; padding: 32px 24px; }}
  .section {{ background: white; border-radius: 10px; padding: 28px 32px; margin-bottom: 28px;
              box-shadow: 0 2px 8px rgba(0,0,0,0.06); }}
  h2 {{ color: {NAVY}; font-size: 16px; font-weight: 700; border-bottom: 2px solid {BLUE};
        padding-bottom: 10px; margin-bottom: 20px; }}
  h3 {{ color: {NAVY}; font-size: 13px; font-weight: 700; margin: 16px 0 8px; }}
  .metric-grid {{ display: grid; grid-template-columns: repeat(auto-fill, minmax(180px, 1fr)); gap: 16px; }}
  .metric-card {{ background: #f0f4fa; border-left: 4px solid {BLUE}; border-radius: 6px;
                  padding: 14px 16px; }}
  .metric-card.bad  {{ border-left-color: {RED}; }}
  .metric-card.good {{ border-left-color: {GREEN}; }}
  .metric-label {{ font-size: 11px; color: #888; text-transform: uppercase; letter-spacing: 0.5px; }}
  .metric-value {{ font-size: 20px; font-weight: 700; color: {NAVY}; margin-top: 4px; }}
  .metric-sub   {{ font-size: 10px; color: #999; margin-top: 3px; }}
  table {{ width: 100%; border-collapse: collapse; font-size: 12px; }}
  th    {{ background: {NAVY}; color: white; padding: 9px 10px; text-align: left; font-weight: 600; }}
  td    {{ padding: 8px 10px; border-bottom: 1px solid #eee; }}
  tr:nth-child(even) td {{ background: #f8f9fc; }}
  tr:hover td {{ background: #e8f0fe; }}
  .exec-box {{ background: #f0f4fa; border-left: 4px solid {NAVY}; padding: 20px 24px;
               border-radius: 6px; line-height: 1.8; font-size: 13px; }}
  .chart-row {{ display: grid; grid-template-columns: 1fr 1fr; gap: 24px; }}
  .chart-box {{ text-align: center; }}
  .chart-caption {{ font-size: 11px; color: #888; margin-top: 8px; font-style: italic; }}
  .tag {{ display: inline-block; padding: 2px 8px; border-radius: 12px; font-size: 10px; font-weight: 600; }}
  .tag-pass {{ background: #e8f5e9; color: {GREEN}; }}
  .tag-warn {{ background: #fff9c4; color: #c67c00; }}
  .tag-fail {{ background: #ffebee; color: {RED}; }}
  @media print {{ body {{ background: white; }} .section {{ box-shadow: none; }} }}
</style>
</head>
<body>
<div class="header">
  <h1>{settings_title}</h1>
  <p>Generated: {now} &nbsp;|&nbsp; Portfolio Value: {fmt_currency(md.total_portfolio_value)} &nbsp;|&nbsp; {len(md.holdings)} Holdings</p>
</div>
<div class="container">

<!-- EXECUTIVE SUMMARY -->
<div class="section">
  <h2>Executive Summary</h2>
  <div class="exec-box">{exec_summary}</div>
</div>

<!-- KEY METRICS -->
<div class="section">
  <h2>Key Risk Metrics</h2>
  <div class="metric-grid">
    <div class="metric-card {'good' if metrics.sharpe > 1.0 else 'bad' if metrics.sharpe < 0.5 else ''}">
      <div class="metric-label">Sharpe Ratio</div>
      <div class="metric-value">{metrics.sharpe:.2f}</div>
      <div class="metric-sub">Annualised, vs {md.risk_free_rate:.2%} RFR</div>
    </div>
    <div class="metric-card {'bad' if abs(metrics.var_95.parametric_var) > md.total_portfolio_value * 0.025 else ''}">
      <div class="metric-label">VaR (95%, Daily)</div>
      <div class="metric-value">{fmt_currency(abs(metrics.var_95.parametric_var))}</div>
      <div class="metric-sub">Parametric Normal</div>
    </div>
    <div class="metric-card bad">
      <div class="metric-label">CF-VaR (95%, Daily)</div>
      <div class="metric-value">{fmt_currency(abs(metrics.var_95.cornish_fisher_var))}</div>
      <div class="metric-sub">Skew/kurtosis adjusted</div>
    </div>
    <div class="metric-card bad">
      <div class="metric-label">CVaR (95%, Daily)</div>
      <div class="metric-value">{fmt_currency(abs(metrics.var_95.parametric_cvar))}</div>
      <div class="metric-sub">Expected Shortfall</div>
    </div>
    <div class="metric-card">
      <div class="metric-label">Annualised Vol</div>
      <div class="metric-value">{metrics.annualized_vol:.1%}</div>
      <div class="metric-sub">Daily σ × √252</div>
    </div>
    <div class="metric-card">
      <div class="metric-label">Beta (vs {md.benchmark_ticker})</div>
      <div class="metric-value">{metrics.beta:.2f}</div>
      <div class="metric-sub">Market sensitivity</div>
    </div>
    <div class="metric-card bad">
      <div class="metric-label">Max Drawdown</div>
      <div class="metric-value">{metrics.max_drawdown:.1%}</div>
      <div class="metric-sub">{metrics.max_dd_duration} trading days</div>
    </div>
    <div class="metric-card">
      <div class="metric-label">Sortino Ratio</div>
      <div class="metric-value">{metrics.sortino:.2f}</div>
      <div class="metric-sub">Downside-adjusted</div>
    </div>
    <div class="metric-card">
      <div class="metric-label">Eff. Number of Bets</div>
      <div class="metric-value">{metrics.eff_num_bets:.1f}</div>
      <div class="metric-sub">1 / HHI concentration</div>
    </div>
    <div class="metric-card">
      <div class="metric-label">Diversification Ratio</div>
      <div class="metric-value">{metrics.diversification_ratio:.2f}x</div>
      <div class="metric-sub">Wtd avg vol / port vol</div>
    </div>
    <div class="metric-card">
      <div class="metric-label">Avg Pairwise Corr</div>
      <div class="metric-value">{metrics.avg_pairwise_corr:.2f}</div>
      <div class="metric-sub">Higher = less diversified</div>
    </div>
    <div class="metric-card">
      <div class="metric-label">Portfolio Return</div>
      <div class="metric-value {'good' if metrics.annualized_return > 0 else 'bad'}">{metrics.annualized_return:.1%}</div>
      <div class="metric-sub">Geometric annualised</div>
    </div>
    <div class="metric-card">
      <div class="metric-label">PCA Factors (90%)</div>
      <div class="metric-value">{metrics.n_pca_factors_90pct}</div>
      <div class="metric-sub">Independent risk factors</div>
    </div>
  </div>
</div>

<!-- CHARTS ROW 1 -->
<div class="section">
  <h2>Portfolio Visualisations</h2>
  <div class="chart-row">
    <div class="chart-box">
      {img_tag(chart_sector, "Sector Allocation")}
      <div class="chart-caption">Current sector weights. Hover values show percentage of total portfolio.</div>
    </div>
    <div class="chart-box">
      {img_tag(chart_var, "VaR Comparison")}
      <div class="chart-caption">Three VaR methodologies at 95% and 99% confidence. All are daily estimates.</div>
    </div>
  </div>
  <div style="margin-top:24px">
    {img_tag(chart_dd, "Drawdown Chart")}
    <div class="chart-caption">Portfolio drawdown shows percentage below the rolling peak. Deeper troughs = larger drawdowns.</div>
  </div>
  <div style="margin-top:24px">
    {img_tag(chart_rv, "Rolling Volatility")}
    <div class="chart-caption">30-day rolling annualised volatility. Spikes indicate periods of heightened risk.</div>
  </div>
</div>

<!-- MONTE CARLO -->
<div class="section">
  <h2>Monte Carlo Simulation ({sim.n_paths:,} paths × {sim.simulation_days} days)</h2>
  {img_tag(chart_mc, "Monte Carlo Fan Chart")}
  <div class="chart-caption">Each shaded band represents a percentile range of simulated outcomes. Median is the central line.</div>
  <div class="metric-grid" style="margin-top:24px">
    <div class="metric-card good">
      <div class="metric-label">Median Outcome</div>
      <div class="metric-value">{fmt_currency(sim.median_terminal)}</div>
    </div>
    <div class="metric-card good">
      <div class="metric-label">Best Case (95th)</div>
      <div class="metric-value">{fmt_currency(sim.p95_terminal)}</div>
    </div>
    <div class="metric-card bad">
      <div class="metric-label">Worst Case (5th)</div>
      <div class="metric-value">{fmt_currency(sim.p05_terminal)}</div>
    </div>
    <div class="metric-card">
      <div class="metric-label">P(finish positive)</div>
      <div class="metric-value">{sim.prob_positive:.1%}</div>
    </div>
    <div class="metric-card bad">
      <div class="metric-label">P(loss &gt; 10%)</div>
      <div class="metric-value">{sim.prob_loss_10:.1%}</div>
    </div>
    <div class="metric-card bad">
      <div class="metric-label">P(loss &gt; 20%)</div>
      <div class="metric-value">{sim.prob_loss_20:.1%}</div>
    </div>
  </div>
</div>

<!-- STRESS TESTS -->
<div class="section">
  <h2>Stress Test Scenarios</h2>
  {img_tag(chart_stress, "Stress Test Comparison")}
  <div class="chart-caption">Estimated portfolio loss under each scenario. Red = severe (&gt;20%), orange = moderate (&gt;10%).</div>
  <table style="margin-top:20px">
    <thead><tr>
      <th>Scenario</th><th>Period</th>
      <th style="text-align:right">Portfolio Loss (%)</th>
      <th style="text-align:right">Portfolio Loss ($)</th>
      <th style="text-align:right">Benchmark Loss (%)</th>
    </tr></thead>
    <tbody>{stress_rows}</tbody>
  </table>
  <p style="font-size:11px;color:#888;margin-top:12px">
    <em>Historical scenarios use sector-level drawdowns applied to current weights. Actual losses may differ.
    See engine/stress_testing.py for full methodology.</em>
  </p>
</div>

<!-- RISK CONTRIBUTION -->
<div class="section">
  <h2>Risk Contribution Analysis</h2>
  {img_tag(chart_risk_ctb, "Risk Contribution")}
  <div class="chart-caption">Component VaR shows each holding's dollar contribution to total portfolio VaR (95%). Components sum to total VaR by construction.</div>
  <table style="margin-top:20px">
    <thead><tr>
      <th>Ticker</th><th>Sector</th>
      <th style="text-align:right">Ann. Vol</th>
      <th style="text-align:right">Beta</th>
      <th style="text-align:right">Sharpe</th>
      <th style="text-align:right">Sortino</th>
      <th style="text-align:right">Max DD</th>
      <th style="text-align:right">Ann. Return</th>
      <th style="text-align:right">Risk Contrib %</th>
    </tr></thead>
    <tbody>{risk_rows}</tbody>
  </table>
</div>

<!-- HOLDINGS TABLE -->
<div class="section">
  <h2>Portfolio Holdings</h2>
  <table>
    <thead><tr>
      <th>Ticker</th><th>Company</th><th>Sector</th>
      <th style="text-align:right">Shares</th>
      <th style="text-align:right">Cost Basis</th>
      <th style="text-align:right">Current Price</th>
      <th style="text-align:right">Market Value</th>
      <th style="text-align:right">Weight</th>
      <th style="text-align:right">P&amp;L ($)</th>
      <th style="text-align:right">P&amp;L (%)</th>
    </tr></thead>
    <tbody>{holdings_rows}</tbody>
  </table>
</div>

<!-- SECTOR ANALYSIS -->
<div class="section">
  <h2>Sector Analysis — {portfolio_short_name} vs S&amp;P 500</h2>
  <h3>Sector Weight Comparison</h3>
  <table>
    <thead><tr>
      <th>Sector</th>
      <th style="text-align:right">% of S&amp;P 500</th>
      <th style="text-align:right">% of {portfolio_short_name} Portfolio</th>
      <th style="text-align:right">Difference (%)</th>
      <th style="text-align:right">Target Investment ($)</th>
    </tr></thead>
    <tbody>{sector_wt_rows}</tbody>
  </table>
  <p style="font-size:11px;color:#888;margin-top:8px">
    <em>S&amp;P 500 weights are approximate (Feb 2026). Target Investment shows the dollar amount
    needed to match index weight — positive = buy, parenthesised = trim.</em>
  </p>

  <h3 style="margin-top:28px">Sector Returns Comparison</h3>
  <table>
    <thead><tr>
      <th>ETF</th><th>Sector</th>
      <th style="text-align:right">6-Month Return</th>
      <th style="text-align:right">1-Year Return</th>
      <th style="text-align:right">1-Year {portfolio_short_name}</th>
      <th style="text-align:right">Difference</th>
    </tr></thead>
    <tbody>{sector_ret_rows}</tbody>
  </table>
  <p style="font-size:11px;color:#888;margin-top:8px">
    <em>Returns are simple price returns over trailing periods. {portfolio_short_name} returns are weight-weighted
    across holdings in each sector.</em>
  </p>

  <h3 style="margin-top:28px">Holdings by Sector</h3>
  {sector_holdings_html}
</div>

<!-- FOOTER -->
<div style="text-align:center;color:#aaa;font-size:11px;padding:24px 0;">
  Generated by {portfolio_name} Risk System &nbsp;|&nbsp; {now}<br>
  <em>This report is for internal educational purposes only. Not investment advice.</em>
</div>

</div><!-- /container -->
</body>
</html>"""

    # Save file
    filename = f"{portfolio_short_name}_Risk_Report_{timestamp_str()}.html"
    out_path = EXPORTS_DIR / filename
    out_path.write_text(html, encoding="utf-8")
    log.info(f"HTML report saved → {out_path}")
    return out_path


# ═══════════════════════════════════════════════════════════════════════════════
# Excel result writer
# ═══════════════════════════════════════════════════════════════════════════════

def write_results_to_excel(
    excel_path:  Path,
    md:          MarketData,
    metrics:     PortfolioRiskMetrics,
    sim:         SimulationResult,
    stress:      StressTestResults,
    portfolio_short_name: str = "Portfolio",
) -> None:
    """
    Write all risk results back into portfolio_holdings.xlsx as new sheets.
    Overwrites existing results sheets; preserves Holdings, Settings, Instructions.
    """
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    NAVY_CODE  = "1B2A4A"
    BLUE_CODE  = "4A90D9"
    WHITE_CODE = "FFFFFF"
    LIGHT_CODE = "F0F4FA"

    def hdr(ws, row: int, col: int, value, bold: bool = True):
        c = ws.cell(row=row, column=col, value=value)
        c.font = Font(name="Arial", bold=bold, size=10, color=WHITE_CODE)
        c.fill = PatternFill("solid", start_color=NAVY_CODE, fgColor=NAVY_CODE)
        c.alignment = Alignment(horizontal="center", vertical="center")
        return c

    def val(ws, row: int, col: int, value, fmt: str = "", bold: bool = False, color: str = "000000"):
        c = ws.cell(row=row, column=col, value=value)
        c.font = Font(name="Arial", size=10, bold=bold, color=color)
        if fmt:
            c.number_format = fmt
        return c

    log.info(f"Writing results to Excel: {excel_path}")
    wb = openpyxl.load_workbook(str(excel_path))

    # Remove old result sheets if they exist
    for sheet_name in ["Risk Summary", "Stock Risk Detail", "Correlation Matrix",
                        "Stress Test Results", "Sector Analysis",
                        "Monte Carlo Summary", "Last Updated"]:
        if sheet_name in wb.sheetnames:
            del wb[sheet_name]

    # ── Risk Summary ───────────────────────────────────────────────────────────
    ws = wb.create_sheet("Risk Summary")
    rows = [
        ("PORTFOLIO OVERVIEW", ""),
        ("Portfolio Value",         f"${md.total_portfolio_value:,.2f}"),
        ("Number of Holdings",      len(md.holdings)),
        ("Benchmark",               md.benchmark_ticker),
        ("Risk-Free Rate",          f"{md.risk_free_rate:.4f} ({md.risk_free_rate:.2%})"),
        ("", ""),
        ("RETURN & VOLATILITY", ""),
        ("Annualised Return",        f"{metrics.annualized_return:.2%}"),
        ("Annualised Volatility",    f"{metrics.annualized_vol:.2%}"),
        ("Portfolio Beta",           f"{metrics.beta:.3f}"),
        ("Jensen's Alpha",           f"{metrics.alpha:.2%}"),
        ("", ""),
        ("RISK-ADJUSTED RATIOS", ""),
        ("Sharpe Ratio",             f"{metrics.sharpe:.3f}"),
        ("Sortino Ratio",            f"{metrics.sortino:.3f}"),
        ("Calmar Ratio",             f"{metrics.calmar:.3f}"),
        ("", ""),
        ("DRAWDOWN", ""),
        ("Maximum Drawdown",         f"{metrics.max_drawdown:.2%}"),
        ("Max Drawdown Duration",    f"{metrics.max_dd_duration} trading days"),
        ("", ""),
        ("VALUE AT RISK (Daily)", ""),
        ("Parametric VaR 95%",       f"${abs(metrics.var_95.parametric_var):,.2f}"),
        ("Cornish-Fisher VaR 95%",   f"${abs(metrics.var_95.cornish_fisher_var):,.2f}"),
        ("Historical VaR 95%",       f"${abs(metrics.var_95.historical_var):,.2f}"),
        ("Monte Carlo VaR 95%",      f"${abs(sim.var_95):,.2f}"),
        ("Parametric CVaR 95%",      f"${abs(metrics.var_95.parametric_cvar):,.2f}"),
        ("Historical CVaR 95%",      f"${abs(metrics.var_95.historical_cvar):,.2f}"),
        ("Parametric VaR 99%",       f"${abs(metrics.var_99.parametric_var):,.2f}"),
        ("Cornish-Fisher VaR 99%",   f"${abs(metrics.var_99.cornish_fisher_var):,.2f}"),
        ("Historical VaR 99%",       f"${abs(metrics.var_99.historical_var):,.2f}"),
        ("", ""),
        ("CONCENTRATION", ""),
        ("HHI Index",                f"{metrics.hhi:.4f}"),
        ("Effective # of Bets",      f"{metrics.eff_num_bets:.1f}"),
        ("Diversification Ratio",    f"{metrics.diversification_ratio:.3f}"),
        ("Avg Pairwise Correlation", f"{metrics.avg_pairwise_corr:.3f}"),
        ("PCA Factors (90% expl.)",  metrics.n_pca_factors_90pct),
        ("Portfolio Skewness",       f"{metrics.skewness:.3f}"),
        ("Portfolio Kurtosis",       f"{metrics.kurtosis:.3f}"),
    ]
    ws.column_dimensions["A"].width = 32
    ws.column_dimensions["B"].width = 24
    for r, (k, v) in enumerate(rows, start=1):
        if v == "":
            ws.row_dimensions[r].height = 8
            continue
        if not k.replace(" ", "").isupper() or k == "":
            c_a = ws.cell(row=r, column=1, value=k)
            c_a.font = Font(name="Arial", size=10)
            c_b = ws.cell(row=r, column=2, value=v)
            c_b.font = Font(name="Arial", size=10, bold=True, color=NAVY_CODE)
            c_b.alignment = Alignment(horizontal="right")
        else:
            c = ws.cell(row=r, column=1, value=k)
            c.font = Font(name="Arial", bold=True, size=10, color=WHITE_CODE)
            c.fill = PatternFill("solid", start_color=NAVY_CODE, fgColor=NAVY_CODE)
            ws.merge_cells(f"A{r}:B{r}")

    # ── Stock Risk Detail ──────────────────────────────────────────────────────
    ws2 = wb.create_sheet("Stock Risk Detail")
    headers2 = ["Ticker", "Company", "Sector", "Ann Return", "Ann Vol",
                "Beta", "Alpha", "Sharpe", "Sortino", "Max DD",
                "Calmar", "Skew", "Kurt", "Roll Vol 30d", "Roll Vol 90d",
                "Comp VaR 95%", "Risk Contrib %"]
    col_widths2 = [10, 28, 22, 12, 12, 10, 12, 10, 10, 12, 10, 10, 10, 14, 14, 14, 14]
    for ci, (h, w) in enumerate(zip(headers2, col_widths2), start=1):
        c = ws2.cell(row=1, column=ci, value=h)
        c.font = Font(name="Arial", bold=True, size=10, color=WHITE_CODE)
        c.fill = PatternFill("solid", start_color=NAVY_CODE, fgColor=NAVY_CODE)
        c.alignment = Alignment(horizontal="center")
        ws2.column_dimensions[get_column_letter(ci)].width = w

    for ri, sm in enumerate(sorted(metrics.stock_metrics, key=lambda x: -abs(x.component_var_95)), start=2):
        row_fill = PatternFill("solid", start_color=LIGHT_CODE, fgColor=LIGHT_CODE) if ri % 2 == 0 else None
        data = [
            sm.ticker, sm.company_name, sm.sector,
            sm.annualized_return, sm.annualized_vol,
            sm.beta, sm.alpha, sm.sharpe, sm.sortino, sm.max_drawdown,
            sm.calmar, sm.skewness, sm.kurtosis, sm.roll_vol_30d, sm.roll_vol_90d,
            sm.component_var_95, sm.risk_contribution_pct,
        ]
        pct_cols = {4, 5, 7, 10, 14, 15}
        dollar_cols = {16}
        for ci, d in enumerate(data, start=1):
            c = ws2.cell(row=ri, column=ci, value=d)
            c.font = Font(name="Arial", size=10)
            if row_fill:
                c.fill = row_fill
            if ci in pct_cols:
                c.number_format = "0.00%"
            elif ci in dollar_cols:
                c.number_format = "$#,##0.00"

    # ── Correlation Matrix ─────────────────────────────────────────────────────
    if metrics.correlation_matrix is not None:
        ws3 = wb.create_sheet("Correlation Matrix")
        tickers_list = list(metrics.correlation_matrix.columns)
        # Header row
        for ci, t in enumerate([""] + tickers_list, start=1):
            c = ws3.cell(row=1, column=ci, value=t)
            c.font = Font(name="Arial", bold=True, size=9, color=WHITE_CODE)
            c.fill = PatternFill("solid", start_color=NAVY_CODE, fgColor=NAVY_CODE)
            c.alignment = Alignment(horizontal="center")
            ws3.column_dimensions[get_column_letter(ci)].width = 8
        # Data rows
        from openpyxl.styles import PatternFill as PF
        for ri, row_ticker in enumerate(tickers_list, start=2):
            ws3.cell(row=ri, column=1, value=row_ticker).font = Font(name="Arial", bold=True, size=9)
            for ci, col_ticker in enumerate(tickers_list, start=2):
                corr_val = metrics.correlation_matrix.loc[row_ticker, col_ticker]
                c = ws3.cell(row=ri, column=ci, value=round(float(corr_val), 3))
                c.number_format = "0.000"
                c.font = Font(name="Arial", size=8)
                # Color: blue = high correlation, white = low
                r_int = int(max(0, min(255, (1 - abs(corr_val)) * 255)))
                b_int = int(max(0, min(255, abs(corr_val) * 200)))
                hex_color = f"{r_int:02X}DDFF" if corr_val > 0 else f"FF{r_int:02X}{r_int:02X}"
                try:
                    c.fill = PF("solid", start_color=hex_color, fgColor=hex_color)
                except Exception:
                    pass

    # ── Stress Test Results ────────────────────────────────────────────────────
    ws4 = wb.create_sheet("Stress Test Results")
    stress_hdrs = ["Scenario", "Period", "Portfolio Loss (%)", "Portfolio Loss ($)",
                   "Benchmark Loss (%)", "Methodology", "Key Assumptions"]
    for ci, h in enumerate(stress_hdrs, start=1):
        c = ws4.cell(row=1, column=ci, value=h)
        c.font = Font(name="Arial", bold=True, size=10, color=WHITE_CODE)
        c.fill = PatternFill("solid", start_color=NAVY_CODE, fgColor=NAVY_CODE)
        ws4.column_dimensions[get_column_letter(ci)].width = [30,20,18,18,18,50,60][ci-1]
    for ri, s in enumerate(stress.all_scenarios, start=2):
        row_fill = PatternFill("solid", start_color=LIGHT_CODE, fgColor=LIGHT_CODE) if ri % 2 == 0 else None
        data4 = [s.name, s.period, s.portfolio_loss_pct, s.portfolio_loss_usd,
                 s.benchmark_loss_pct, s.methodology, s.assumptions]
        for ci, d in enumerate(data4, start=1):
            c = ws4.cell(row=ri, column=ci, value=d)
            c.font = Font(name="Arial", size=10)
            if row_fill:
                c.fill = row_fill
            if ci in (3, 5):
                c.number_format = "0.00%"
            elif ci == 4:
                c.number_format = "$#,##0.00"
            c.alignment = Alignment(wrap_text=(ci in (6, 7)), vertical="top")
        ws4.row_dimensions[ri].height = 40

    # ── Sector Analysis ────────────────────────────────────────────────────────
    ws_sa = wb.create_sheet("Sector Analysis")

    # --- Section A: Sector Weight Comparison ---
    sa_title = ws_sa.cell(row=1, column=1, value="SECTOR WEIGHT COMPARISON")
    sa_title.font = Font(name="Arial", bold=True, size=10, color=WHITE_CODE)
    sa_title.fill = PatternFill("solid", start_color=NAVY_CODE, fgColor=NAVY_CODE)
    ws_sa.merge_cells("A1:E1")
    for ci in range(2, 6):
        c = ws_sa.cell(row=1, column=ci)
        c.fill = PatternFill("solid", start_color=NAVY_CODE, fgColor=NAVY_CODE)

    sa_wt_hdrs = ["Sector", "% of S&P 500", f"% of {portfolio_short_name} Portfolio", "Difference (%)", "Target Investment ($)"]
    sa_wt_widths = [24, 16, 18, 16, 22]
    for ci, (h, w) in enumerate(zip(sa_wt_hdrs, sa_wt_widths), start=1):
        hdr(ws_sa, 2, ci, h)
        ws_sa.column_dimensions[get_column_letter(ci)].width = w

    pv = md.total_portfolio_value
    port_sw: dict[str, float] = {}
    for h in md.holdings:
        port_sw[h.sector] = port_sw.get(h.sector, 0.0) + h.weight

    sa_row = 3
    for sector, sp_wt in SP500_SECTOR_WEIGHTS.items():
        p_wt = port_sw.get(sector, 0.0)
        diff = p_wt - sp_wt
        target = (sp_wt - p_wt) * pv
        row_fill = PatternFill("solid", start_color=LIGHT_CODE, fgColor=LIGHT_CODE) if sa_row % 2 == 0 else None
        d_color = "27AE60" if diff >= 0 else "E74C3C"

        val(ws_sa, sa_row, 1, sector, bold=True)
        val(ws_sa, sa_row, 2, sp_wt, fmt="0.00%")
        val(ws_sa, sa_row, 3, p_wt, fmt="0.00%")
        val(ws_sa, sa_row, 4, diff, fmt="+0.00%;-0.00%", color=d_color)
        val(ws_sa, sa_row, 5, target, fmt='$#,##0.00', color=d_color)
        if row_fill:
            for ci in range(1, 6):
                ws_sa.cell(row=sa_row, column=ci).fill = row_fill
        sa_row += 1

    # Totals row
    tot_sp = sum(SP500_SECTOR_WEIGHTS.values())
    tot_pt = sum(port_sw.get(s, 0.0) for s in SP500_SECTOR_WEIGHTS)
    border_top = Border(top=Side(style="medium", color=NAVY_CODE))
    for ci in range(1, 6):
        ws_sa.cell(row=sa_row, column=ci).border = border_top
    val(ws_sa, sa_row, 1, "Totals", bold=True)
    val(ws_sa, sa_row, 2, tot_sp, fmt="0.00%", bold=True)
    val(ws_sa, sa_row, 3, tot_pt, fmt="0.00%", bold=True)

    # --- Section B: Sector Returns Comparison ---
    sa_ret_start = sa_row + 3
    sa_ret_title = ws_sa.cell(row=sa_ret_start, column=1, value="SECTOR RETURNS COMPARISON")
    sa_ret_title.font = Font(name="Arial", bold=True, size=10, color=WHITE_CODE)
    sa_ret_title.fill = PatternFill("solid", start_color=NAVY_CODE, fgColor=NAVY_CODE)
    ws_sa.merge_cells(f"A{sa_ret_start}:F{sa_ret_start}")
    for ci in range(2, 7):
        ws_sa.cell(row=sa_ret_start, column=ci).fill = PatternFill("solid", start_color=NAVY_CODE, fgColor=NAVY_CODE)

    sa_ret_hdrs = ["ETF Ticker", "Sector", "6-Month Return", "1-Year Return", f"1-Year {portfolio_short_name}", "Difference"]
    sa_ret_widths = [12, 24, 16, 16, 14, 14]
    hdr_row = sa_ret_start + 1
    for ci, (h, w) in enumerate(zip(sa_ret_hdrs, sa_ret_widths), start=1):
        hdr(ws_sa, hdr_row, ci, h)
        cur_w = ws_sa.column_dimensions[get_column_letter(ci)].width
        if w > cur_w:
            ws_sa.column_dimensions[get_column_letter(ci)].width = w

    try:
        etf_prices_xl = _fetch_sector_etf_prices()
    except Exception:
        etf_prices_xl = pd.DataFrame()

    sa_dr = hdr_row + 1
    for sector, etf_ticker in SECTOR_ETF_MAP.items():
        if etf_ticker not in etf_prices_xl.columns:
            continue
        etf_s = etf_prices_xl[etf_ticker].dropna()
        r6 = _period_return(etf_s, 6)
        r12 = _period_return(etf_s, 12)
        dcm12 = _sector_portfolio_return(md.prices, md.holdings, sector, 12)
        diff_r = (dcm12 - r12) if (dcm12 is not None and r12 is not None) else None

        row_fill = PatternFill("solid", start_color=LIGHT_CODE, fgColor=LIGHT_CODE) if sa_dr % 2 == 0 else None

        val(ws_sa, sa_dr, 1, etf_ticker, bold=True)
        val(ws_sa, sa_dr, 2, sector)
        for ci, rv in [(3, r6), (4, r12), (5, dcm12), (6, diff_r)]:
            if rv is not None:
                rc = "27AE60" if rv >= 0 else "E74C3C"
                val(ws_sa, sa_dr, ci, rv, fmt="0.0%", color=rc)
            else:
                val(ws_sa, sa_dr, ci, "N/A")
        if row_fill:
            for ci in range(1, 7):
                ws_sa.cell(row=sa_dr, column=ci).fill = row_fill
        sa_dr += 1

    # ── Monte Carlo Summary ────────────────────────────────────────────────────
    ws5 = wb.create_sheet("Monte Carlo Summary")
    mc_rows = [
        ("Simulation Parameters", ""),
        ("Number of Paths",  sim.n_paths),
        ("Simulation Days",  sim.simulation_days),
        ("Initial Value",    f"${sim.initial_value:,.2f}"),
        ("", ""),
        ("Terminal Portfolio Value Percentiles", ""),
        ("5th Percentile (Worst Case)",  f"${sim.p05_terminal:,.2f}"),
        ("25th Percentile",              f"${sim.p25_terminal:,.2f}"),
        ("Median (50th Percentile)",     f"${sim.median_terminal:,.2f}"),
        ("75th Percentile",              f"${sim.p75_terminal:,.2f}"),
        ("95th Percentile (Best Case)",  f"${sim.p95_terminal:,.2f}"),
        ("", ""),
        ("Loss Probabilities", ""),
        ("P(loss > 10%)",     f"{sim.prob_loss_10:.2%}"),
        ("P(loss > 20%)",     f"{sim.prob_loss_20:.2%}"),
        ("P(loss > 30%)",     f"{sim.prob_loss_30:.2%}"),
        ("P(finish positive)", f"{sim.prob_positive:.2%}"),
        ("", ""),
        ("1-Year VaR from Simulation", ""),
        ("MC VaR 95%",   f"${abs(sim.var_95):,.2f}"),
        ("MC VaR 99%",   f"${abs(sim.var_99):,.2f}"),
        ("MC CVaR 95%",  f"${abs(sim.cvar_95):,.2f}"),
        ("MC CVaR 99%",  f"${abs(sim.cvar_99):,.2f}"),
    ]
    ws5.column_dimensions["A"].width = 35
    ws5.column_dimensions["B"].width = 22
    for r, (k, v) in enumerate(mc_rows, start=1):
        if not k:
            ws5.row_dimensions[r].height = 8
            continue
        if not k.replace(" ", "").replace("(", "").replace(")", "").replace("%", "").isupper() or k == "":
            ws5.cell(row=r, column=1, value=k).font = Font(name="Arial", size=10)
            c = ws5.cell(row=r, column=2, value=v)
            c.font = Font(name="Arial", size=10, bold=True, color=NAVY_CODE)
            c.alignment = Alignment(horizontal="right")
        else:
            c = ws5.cell(row=r, column=1, value=k)
            c.font = Font(name="Arial", bold=True, size=10, color=WHITE_CODE)
            c.fill = PatternFill("solid", start_color=NAVY_CODE, fgColor=NAVY_CODE)
            ws5.merge_cells(f"A{r}:B{r}")

    # ── Last Updated ───────────────────────────────────────────────────────────
    ws6 = wb.create_sheet("Last Updated")
    ws6["A1"] = "Last Model Update"
    ws6["A1"].font = Font(name="Arial", bold=True, size=12, color=WHITE_CODE)
    ws6["A1"].fill = PatternFill("solid", start_color=NAVY_CODE, fgColor=NAVY_CODE)
    ws6.column_dimensions["A"].width = 35
    ws6.column_dimensions["B"].width = 50
    info_rows = [
        ("Last Updated", datetime.now().strftime("%Y-%m-%d %H:%M:%S")),
        ("Tickers Used", ", ".join(sorted(md.portfolio_tickers))),
        ("Failed Tickers", ", ".join(md.quality.failed_tickers) or "None"),
        ("Price Data Range", f"{md.quality.date_range[0]} → {md.quality.date_range[1]}"),
        ("Warnings", str(len(md.quality.log_lines))),
    ]
    for r, (k, v) in enumerate(info_rows, start=2):
        ws6.cell(row=r, column=1, value=k).font = Font(name="Arial", size=10, bold=True)
        c = ws6.cell(row=r, column=2, value=v)
        c.font = Font(name="Arial", size=10)
        c.alignment = Alignment(wrap_text=True)

    wb.save(str(excel_path))
    log.info(f"Excel results saved → {excel_path}")
